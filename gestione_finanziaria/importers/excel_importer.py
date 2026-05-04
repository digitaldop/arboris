"""
Parser Excel configurabile per estratti conto.

Normalizza il primo foglio Excel in righe testuali e riusa la stessa
configurazione a colonne del parser CSV. Supporta nativamente XLSX/XLSM,
XML Spreadsheet 2003 e semplici file HTML con estensione XLS. I vecchi XLS
binari sono supportati quando la dipendenza opzionale ``xlrd`` e' installata.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal
from html.parser import HTMLParser
from typing import Dict, Iterator, List, Optional
from xml.etree import ElementTree as ET
from zipfile import BadZipFile

from openpyxl import load_workbook

from .base import BaseParser, ParsedMovimento
from .csv_importer import (
    CsvImportDetection,
    CsvImporter,
    CsvImporterConfig,
    _build_header_map,
    _column_value,
    _find_column,
)


EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")
OOXML_SIGNATURE = b"PK\x03\x04"
OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def is_probable_excel_file(raw_bytes: bytes, filename: str = "") -> bool:
    lower_name = (filename or "").lower()
    stripped = (raw_bytes or b"").lstrip()[:512].lower()
    if lower_name.endswith(EXCEL_EXTENSIONS):
        return True
    if (raw_bytes or b"").startswith(OOXML_SIGNATURE) or (raw_bytes or b"").startswith(OLE_SIGNATURE):
        return True
    return (
        stripped.startswith(b"<html")
        or stripped.startswith(b"<!doctype html")
        or stripped.startswith(b"<table")
        or b"urn:schemas-microsoft-com:office:spreadsheet" in stripped
        or b"<workbook" in stripped
    )


def _cell_to_string(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, Decimal):
        text = format(value, "f")
        return text.rstrip("0").rstrip(".").replace(".", ",") if "." in text else text
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = f"{value:.10f}".rstrip("0").rstrip(".")
        return text.replace(".", ",")
    return str(value).strip()


def _read_ooxml_rows(raw_bytes: bytes) -> List[List[str]]:
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except BadZipFile as exc:
        raise ValueError("Il file Excel non sembra un XLSX valido.") from exc

    worksheet = workbook.active
    rows: List[List[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        values = [_cell_to_string(value) for value in row]
        while values and not values[-1]:
            values.pop()
        if any(values):
            rows.append(values)
    return rows


def _read_legacy_xls_rows(raw_bytes: bytes) -> List[List[str]]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise ValueError(
            "Il file XLS e' nel vecchio formato binario. Installa la dipendenza xlrd oppure esporta dalla banca in XLSX."
        ) from exc

    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    if not workbook.nsheets:
        return []
    sheet = workbook.sheet_by_index(0)
    rows: List[List[str]] = []
    for row_idx in range(sheet.nrows):
        values: List[str] = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
            values.append(_cell_to_string(value))
        while values and not values[-1]:
            values.pop()
        if any(values):
            rows.append(values)
    return rows


def _read_xml_spreadsheet_rows(raw_bytes: bytes) -> List[List[str]]:
    root = ET.fromstring(raw_bytes)
    rows: List[List[str]] = []
    row_index = 1
    for row_el in root.iter():
        if not row_el.tag.lower().endswith("row"):
            continue
        values: List[str] = []
        for cell_el in list(row_el):
            if not cell_el.tag.lower().endswith("cell"):
                continue
            index_value = (
                cell_el.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
                or cell_el.attrib.get("ss:Index")
                or cell_el.attrib.get("Index")
            )
            if index_value:
                try:
                    while len(values) < int(index_value) - 1:
                        values.append("")
                except ValueError:
                    pass
            data_text = ""
            for data_el in cell_el.iter():
                if data_el is not cell_el and data_el.tag.lower().endswith("data"):
                    data_text = "".join(data_el.itertext()).strip()
                    break
            values.append(data_text)
        while values and not values[-1]:
            values.pop()
        if any(values):
            rows.append(values)
        row_index += 1
    return rows


class _HtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows: List[List[str]] = []
        self._current_row: Optional[List[str]] = None
        self._current_cell: Optional[List[str]] = None
        self._table_depth = 0

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table":
            self._table_depth += 1
        if self._table_depth <= 0:
            return
        if tag == "tr":
            self._current_row = []
        elif tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []

    def handle_endtag(self, tag):
        tag = tag.lower()
        if self._table_depth <= 0:
            return
        if tag in {"td", "th"} and self._current_row is not None and self._current_cell is not None:
            self._current_row.append(" ".join("".join(self._current_cell).split()))
            self._current_cell = None
        elif tag == "tr" and self._current_row is not None:
            if any(self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None
        elif tag == "table":
            self._table_depth -= 1

    def handle_data(self, data):
        if self._current_cell is not None:
            self._current_cell.append(data)


def _read_html_table_rows(raw_bytes: bytes) -> List[List[str]]:
    text = raw_bytes.decode("utf-8", errors="ignore")
    parser = _HtmlTableParser()
    parser.feed(text)
    return parser.rows


def read_excel_rows(raw_bytes: bytes) -> List[List[str]]:
    stripped = (raw_bytes or b"").lstrip()[:512].lower()
    if not raw_bytes:
        return []
    if raw_bytes.startswith(OLE_SIGNATURE):
        return _read_legacy_xls_rows(raw_bytes)
    if raw_bytes.startswith(OOXML_SIGNATURE):
        return _read_ooxml_rows(raw_bytes)
    if stripped.startswith(b"<html") or stripped.startswith(b"<!doctype html") or stripped.startswith(b"<table"):
        return _read_html_table_rows(raw_bytes)
    if b"urn:schemas-microsoft-com:office:spreadsheet" in stripped or b"<workbook" in stripped:
        return _read_xml_spreadsheet_rows(raw_bytes)
    return _read_ooxml_rows(raw_bytes)


def _row_has_known_headers(row: List[str]) -> bool:
    header_map = _build_header_map(row)
    return any(
        item in header_map
        for item in {
            "operazione",
            "data operazione",
            "data contabile",
            "data",
            "importo",
            "descrizione",
            "causale",
            "dare",
            "avere",
            "uscite",
            "entrate",
        }
    )


def _detect_header_row(rows: List[List[str]]) -> Optional[int]:
    for idx, row in enumerate(rows[:25]):
        if _row_has_known_headers(row):
            return idx
    return None


def detect_excel_import_config(raw_bytes: bytes) -> CsvImportDetection:
    rows = read_excel_rows(raw_bytes)
    if not rows:
        raise ValueError("Il file Excel non contiene righe leggibili.")

    header_idx = _detect_header_row(rows)
    if header_idx is None:
        detection = CsvImportDetection(
            config=CsvImporterConfig(ha_intestazione=True),
            formato_rilevato="Excel",
            confidenza=10,
        )
        detection.avvisi.append(
            "Intestazione non riconosciuta: potrebbe essere necessario usare le opzioni avanzate."
        )
        return detection

    header = rows[header_idx]
    header_map = _build_header_map(header)
    config = CsvImporterConfig(
        delimiter="",
        encoding="utf-8-sig",
        ha_intestazione=True,
        righe_da_saltare=header_idx,
        formato_data="%d/%m/%Y",
        separatore_decimale=",",
        separatore_migliaia=".",
    )
    detection = CsvImportDetection(config=config, formato_rilevato="Excel", confidenza=40)

    field_candidates = {
        "colonna_data_contabile": ["Operazione", "Data operazione", "Data contabile", "Data contabilizzazione", "Data"],
        "colonna_data_valuta": ["Data valuta", "Valuta"],
        "colonna_importo": ["Importo", "Importo movimento", "Amount"],
        "colonna_entrate": ["Entrate", "Avere", "Accrediti", "Importo avere"],
        "colonna_uscite": ["Uscite", "Dare", "Addebiti", "Importo dare"],
        "colonna_controparte": ["Controparte", "Beneficiario", "Ordinante"],
        "colonna_iban_controparte": ["IBAN controparte", "Iban ordinante", "Iban beneficiario"],
        "colonna_transaction_id": ["Identificativo End to End", "End To End Id", "ID transazione", "Transaction ID"],
    }

    for config_field, candidates in field_candidates.items():
        column = _find_column(header_map, candidates)
        if column:
            setattr(config, config_field, column)
            detection.colonne_rilevate[config_field] = column
            detection.confidenza += 8

    if config.colonna_importo is not None:
        config.colonna_entrate = None
        config.colonna_uscite = None

    description_columns = []
    for candidate in [
        "Informazioni di riconciliazione",
        "Descrizione",
        "Causale descrizione",
        "Descrizione operazione",
        "Causale",
    ]:
        column = _find_column(header_map, [candidate])
        if column and column not in description_columns:
            description_columns.append(column)

    if description_columns:
        config.colonna_descrizione = description_columns[0]
        config.colonne_descrizione_extra = description_columns[1:]
        detection.colonne_rilevate["colonna_descrizione"] = ", ".join(description_columns)
        detection.confidenza += 10

    if all(_find_column(header_map, [col]) for col in ["Rag. Soc./ Intestatario", "ABI", "CAB", "Conto"]):
        detection.formato_rilevato = "Excel CBI"
        detection.confidenza += 15

    metadata_row = rows[header_idx + 1] if len(rows) > header_idx + 1 else []
    detection.abi = _column_value(metadata_row, _find_column(header_map, ["ABI"]), header_map)
    detection.cab = _column_value(metadata_row, _find_column(header_map, ["CAB"]), header_map)
    detection.numero_conto = _column_value(metadata_row, _find_column(header_map, ["Conto"]), header_map)
    detection.intestatario = _column_value(metadata_row, _find_column(header_map, ["Rag. Soc./ Intestatario"]), header_map)

    mancanti = config.descrizione_campi_richiesti()
    if mancanti:
        detection.avvisi.append(
            "Non sono riuscito a rilevare automaticamente: " + ", ".join(mancanti) + "."
        )
    detection.confidenza = min(detection.confidenza, 100)
    return detection


class ExcelImporter(BaseParser):
    nome_formato = "excel"

    def __init__(self, config: CsvImporterConfig):
        self.config = config

    def parse(self, raw_bytes: bytes) -> Iterator[ParsedMovimento]:
        rows = read_excel_rows(raw_bytes)
        if not rows:
            return iter(())
        csv_text = "\n".join(
            ",".join('"' + re.sub(r'"', '""', cell or "") + '"' for cell in row)
            for row in rows
        )
        csv_config = CsvImporterConfig(**self.config.__dict__)
        csv_config.delimiter = ","
        csv_config.encoding = "utf-8-sig"
        return CsvImporter(csv_config).parse(csv_text.encode("utf-8-sig"))
