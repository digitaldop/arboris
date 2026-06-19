import json
import os
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
import shutil
import tempfile
from unittest import skip
from unittest.mock import Mock, patch

from django.contrib.auth.models import User
from django.contrib.messages.storage.fallback import FallbackStorage
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.http import HttpResponse
from django.core.management import call_command
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from anagrafica.models import Familiare, RelazioneFamiliare, Studente, StudenteFamiliare
from economia.models import CondizioneIscrizione, Iscrizione, RataIscrizione, StatoIscrizione, TariffaCondizioneIscrizione
from gestione_amministrativa.models import BustaPagaDipendente, Dipendente
from scuola.models import AnnoScolastico, Classe
from sistema.models import LivelloPermesso, SistemaUtentePermessi

from .models import (
    CategoriaFinanziaria,
    CanaleMovimento,
    CondizioneRegolaCategorizzazione,
    ConnessioneBancaria,
    ContoBancario,
    DocumentoFornitore,
    DocumentoFornitoreImportAlias,
    EsitoSincronizzazione,
    FattureInCloudConnessione,
    FattureInCloudSyncLog,
    FrequenzaVoceBudget,
    Fornitore,
    FonteSaldo,
    MetodoPagamentoFornitore,
    MovimentoFinanziario,
    NotificaFinanziaria,
    NotificaFinanziariaLettura,
    PagamentoFornitore,
    PianoRatealeSpesa,
    PianificazioneSincronizzazione,
    OrigineDocumentoFornitore,
    OrigineMovimento,
    ProviderBancario,
    RiconciliazioneRataMovimento,
    RegolaCategorizzazione,
    SaldoConto,
    ScadenzaPagamentoFornitore,
    SegnoMovimento,
    SincronizzazioneLog,
    StatoConnessioneBancaria,
    StatoRiconciliazione,
    StatoDocumentoFornitore,
    StatoScadenzaFornitore,
    SpesaOperativa,
    TipoCategoriaFinanziaria,
    TipoContoFinanziario,
    TipoPianoRatealeSpesa,
    TipoProviderBancario,
    TipoSpesaOperativa,
    TipoDocumentoFornitore,
    TipoVoceBudget,
    VoceBudgetRicorrente,
)


def crea_categoria_spesa_test(nome, **kwargs):
    kwargs.setdefault("tipo", TipoCategoriaFinanziaria.SPESA)
    return CategoriaFinanziaria.objects.create(nome=nome, **kwargs)


from .fatture_in_cloud import (
    FattureInCloudClient,
    FattureInCloudError,
    authorization_url,
    has_oauth_credentials,
    importa_documento_fatture_in_cloud,
    sincronizza_fatture_in_cloud,
)
from .importers import CsvImporter, CsvImporterConfig, ExcelImporter, detect_csv_import_config, detect_excel_import_config
from .importers.service import importa_movimenti_da_file
from .forms import DocumentoFornitoreForm
from .providers.base import ProviderAccount, ProviderConnectionInfo, ProviderTransaction
from .providers.enablebanking import EnableBankingAdapter, EnableBankingCredentials
from .services import (
    annulla_pagamento_fornitore,
    anteprima_riconcilia_fornitori_automaticamente,
    applica_regole_a_movimento,
    importo_movimento_disponibile_fornitori,
    importo_rata_residuo,
    applica_anteprima_riconciliazione_fornitori,
    applica_proposta_riconciliazione,
    build_home_financial_dashboard_data,
    build_budgeting_dashboard_data,
    calcola_hash_deduplica_movimento,
    crea_proposta_riconciliazione,
    proposte_riconciliazione_da_movimento,
    proposte_riconciliazione_da_rata,
    proposte_riconciliazione_da_scadenza_fornitore,
    riconcilia_movimento_con_scadenze_fornitore,
    riconcilia_movimento_con_scadenza_fornitore,
    riconcilia_movimento_con_rate,
    registra_pagamento_fornitore,
    trova_movimenti_cumulativi_candidati_per_rate,
    trova_movimenti_cumulativi_candidati_per_scadenza_fornitore,
    trova_scadenze_fornitori_cumulative_candidate,
    trova_scadenze_fornitori_candidate,
    trova_movimenti_candidati_per_rate,
    trova_rate_cumulative_candidate,
    trova_rate_candidate,
)


CBI_CSV_SAMPLE = (
    '"Rag. Soc./ Intestatario";"ABI";"CAB";"Conto";"Operazione";"Valuta";"Importo";"Causale";'
    '"Causale Interna";"Descrizione";"Identificativo End to End";"Informazioni di riconciliazione"\n'
    '"IL SOLE E L\'ALTRE STELLE SRL IMPRESA SOCIALE";"05034";"37060";"000000003228";"24/04/2026";'
    '"24/04/2026";"300,00";"48";"0";"BONIF. VS. FAVORE - YYY24042026 GHEDUZZI";"NOTPROVIDED ";'
    '"Iscrizione 4 classe 2026-2027 Gheduzzi Sofia "\n'
    '"IL SOLE E L\'ALTRE STELLE SRL IMPRESA SOCIALE";"05034";"37060";"000000003228";"24/04/2026";'
    '"24/04/2026";"-24,40";"50";"C";"ADDEBITO DIRETTO SDD - PayPal Europe";"";""\n'
)


class Psd2SchedulerTests(TestCase):
    @patch("gestione_finanziaria.scheduler.sincronizza_conto_psd2")
    @patch("gestione_finanziaria.scheduler.conti_target")
    def test_force_sync_runs_even_when_automatic_schedule_is_inactive(
        self,
        mock_conti_target,
        mock_sincronizza_conto,
    ):
        from gestione_finanziaria.scheduler import maybe_run_scheduled_sync

        PianificazioneSincronizzazione.objects.update_or_create(
            pk=1,
            defaults={
                "attivo": False,
                "sync_saldo": True,
                "sync_movimenti": True,
                "giorni_storico": 14,
                "in_corso": False,
            },
        )
        conto = Mock(nome_conto="Conto PSD2")
        mock_conti_target.return_value = [conto]
        mock_sincronizza_conto.return_value = Mock(
            esito=EsitoSincronizzazione.OK,
            messaggio="ok",
        )

        risultato = maybe_run_scheduled_sync(force=True)

        self.assertIsNotNone(risultato)
        risultato.refresh_from_db()
        self.assertFalse(risultato.in_corso)
        self.assertEqual(risultato.conti_sincronizzati, 1)
        self.assertEqual(risultato.conti_in_errore, 0)
        self.assertIsNotNone(risultato.ultimo_run_at)
        mock_sincronizza_conto.assert_called_once_with(
            conto,
            sync_saldo=True,
            sync_movimenti=True,
            giorni_storico=14,
        )

    @patch("gestione_finanziaria.scheduler.sincronizza_conto_psd2")
    def test_force_sync_keeps_in_progress_lock(self, mock_sincronizza_conto):
        from gestione_finanziaria.scheduler import maybe_run_scheduled_sync

        PianificazioneSincronizzazione.objects.update_or_create(
            pk=1,
            defaults={
                "attivo": False,
                "in_corso": True,
                "avviato_at": timezone.now(),
            },
        )

        risultato = maybe_run_scheduled_sync(force=True)

        self.assertIsNone(risultato)
        mock_sincronizza_conto.assert_not_called()

    @patch("gestione_finanziaria.scheduler.conti_target", side_effect=RuntimeError("boom globale"))
    def test_scheduled_sync_releases_lock_when_global_error_occurs(self, mock_conti_target):
        from gestione_finanziaria.scheduler import maybe_run_scheduled_sync

        PianificazioneSincronizzazione.objects.update_or_create(
            pk=1,
            defaults={
                "attivo": True,
                "in_corso": False,
                "sync_saldo": True,
                "sync_movimenti": True,
                "giorni_storico": 14,
            },
        )

        risultato = maybe_run_scheduled_sync(force=True)

        self.assertIsNotNone(risultato)
        risultato.refresh_from_db()
        self.assertFalse(risultato.in_corso)
        self.assertEqual(risultato.conti_sincronizzati, 0)
        self.assertEqual(risultato.conti_in_errore, 1)
        self.assertEqual(risultato.ultimo_esito, EsitoSincronizzazione.ERRORE)
        self.assertIn("Errore imprevisto scheduler - boom globale", risultato.ultimo_messaggio)
        mock_conti_target.assert_called_once()

    def test_background_scheduler_skips_management_commands(self):
        from gestione_finanziaria.background_scheduler import should_start_background_scheduler

        with patch.object(sys, "argv", ["manage.py", "migrate"]):
            enabled, reason = should_start_background_scheduler()

        self.assertFalse(enabled)
        self.assertEqual(reason, "comando di management")

    def test_background_scheduler_starts_for_web_process_by_default(self):
        from gestione_finanziaria.background_scheduler import should_start_background_scheduler

        with patch.object(sys, "argv", ["gunicorn", "arboris.wsgi:application"]):
            with patch.dict(os.environ, {}, clear=True):
                enabled, reason = should_start_background_scheduler()

        self.assertTrue(enabled)
        self.assertEqual(reason, "processo web")

    def test_background_scheduler_runs_check_before_first_sleep(self):
        from gestione_finanziaria import background_scheduler

        with patch.object(background_scheduler, "_run_due_syncs") as mock_run_due_syncs:
            with patch.object(background_scheduler.time, "sleep", side_effect=RuntimeError("stop")):
                with self.assertRaises(RuntimeError):
                    background_scheduler._scheduler_loop(300)

        mock_run_due_syncs.assert_called_once()

    @patch("gestione_finanziaria.middleware.trigger_due_sync_check_async")
    @patch("gestione_finanziaria.middleware.start_background_scheduler_once")
    def test_schedule_middleware_triggers_async_check_without_blocking(
        self,
        mock_start_background_scheduler,
        mock_trigger_due_sync_check,
    ):
        from gestione_finanziaria.middleware import SincronizzazionePsd2ScheduleMiddleware

        request = RequestFactory().get("/gestione-finanziaria/pianificazione-sincronizzazione/")
        middleware = SincronizzazionePsd2ScheduleMiddleware(lambda request: HttpResponse("ok"))

        response = middleware(request)

        self.assertEqual(response.status_code, 200)
        mock_start_background_scheduler.assert_called_once()
        mock_trigger_due_sync_check.assert_called_once()

    @patch("gestione_finanziaria.views.background_scheduler_status")
    def test_pianificazione_sincronizzazione_renderizza_layout_moderno(self, mock_background_status):
        from gestione_finanziaria.views import pianificazione_sincronizzazione

        mock_background_status.return_value = {
            "enabled": True,
            "thread_alive": True,
            "interval_minutes": 5,
            "reason": "processo web",
        }
        provider = ProviderBancario.objects.create(
            nome="Enable Banking Layout",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM",
            external_connection_id="session-layout",
        )
        ContoBancario.objects.create(
            nome_conto="Banco BPM PSD2",
            provider=provider,
            connessione=connessione,
            external_account_id="account-layout",
            iban="IT67C0503437060000000003228",
            attivo=True,
        )
        PianificazioneSincronizzazione.objects.update_or_create(
            pk=1,
            defaults={
                "attivo": True,
                "intervallo_ore": 12,
                "sync_saldo": True,
                "sync_movimenti": True,
                "giorni_storico": 90,
            },
        )

        request = RequestFactory().get("/gestione-finanziaria/pianificazione-sincronizzazione/")
        request.session = self.client.session
        request._messages = FallbackStorage(request)
        request.user = User.objects.create_user("sync-layout@example.com")

        response = pianificazione_sincronizzazione(request)
        content = response.content.decode("utf-8")

        self.assertIn("finance-page-head", content)
        self.assertIn("finance-guide-panel", content)
        self.assertIn("finance-two-column-layout", content)
        self.assertIn("finance-sync-summary", content)
        self.assertIn("Banco BPM PSD2", content)

    @patch("gestione_finanziaria.providers.adapter_for_provider")
    def test_psd2_account_sync_uses_real_time_module_for_duration(self, mock_adapter_for_provider):
        from gestione_finanziaria.services import sincronizza_conto_psd2

        provider = ProviderBancario.objects.create(
            nome="Provider PSD2 test",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "test"},
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Connessione test",
            external_connection_id="session-test",
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto PSD2 test",
            provider=provider,
            connessione=connessione,
            external_account_id="account-test",
            attivo=True,
        )
        adapter = Mock()
        adapter.saldo_conto.return_value = []
        adapter.movimenti_conto.return_value = []
        mock_adapter_for_provider.return_value = adapter

        log = sincronizza_conto_psd2(conto)

        self.assertEqual(log.esito, EsitoSincronizzazione.OK)
        self.assertEqual(log.movimenti_inseriti, 0)
        self.assertIsNotNone(log.durata_millisecondi)
        adapter.saldo_conto.assert_called_once_with("account-test")
        adapter.movimenti_conto.assert_called_once()

    @patch("gestione_finanziaria.providers.adapter_for_provider")
    def test_psd2_account_sync_marks_enablebanking_expired_session(self, mock_adapter_for_provider):
        from gestione_finanziaria.providers.enablebanking import EnableBankingSessionExpired
        from gestione_finanziaria.services import sincronizza_conto_psd2

        provider = ProviderBancario.objects.create(
            nome="Enable Banking scadenza",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM",
            external_connection_id="session-expired",
            stato=StatoConnessioneBancaria.ATTIVA,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Banco BPM",
            provider=provider,
            connessione=connessione,
            external_account_id="account-expired",
            attivo=True,
        )
        adapter = Mock()
        adapter.saldo_conto.side_effect = EnableBankingSessionExpired("Session is expired")
        mock_adapter_for_provider.return_value = adapter

        log = sincronizza_conto_psd2(conto)

        connessione.refresh_from_db()
        self.assertEqual(log.esito, EsitoSincronizzazione.ERRORE)
        self.assertEqual(connessione.stato, StatoConnessioneBancaria.SCADUTA)
        self.assertIn("Consenso PSD2 scaduto", connessione.ultimo_errore)
        self.assertIn("rinnova il consenso", log.messaggio)
        adapter.movimenti_conto.assert_not_called()

    def test_scheduler_excludes_expired_psd2_connections(self):
        from gestione_finanziaria.scheduler import conti_target

        provider = ProviderBancario.objects.create(
            nome="Provider PSD2 target",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        connessione_attiva = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Connessione attiva",
            stato=StatoConnessioneBancaria.ATTIVA,
        )
        connessione_scaduta = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Connessione scaduta",
            stato=StatoConnessioneBancaria.SCADUTA,
        )
        conto_attivo = ContoBancario.objects.create(
            nome_conto="Conto attivo",
            provider=provider,
            connessione=connessione_attiva,
            external_account_id="account-active",
            attivo=True,
        )
        conto_scaduto = ContoBancario.objects.create(
            nome_conto="Conto scaduto",
            provider=provider,
            connessione=connessione_scaduta,
            external_account_id="account-expired",
            attivo=True,
        )

        self.assertIn(conto_attivo, conti_target())
        self.assertNotIn(conto_scaduto, conti_target())

    @patch("gestione_finanziaria.providers.adapter_for_provider")
    def test_psd2_account_sync_crea_notifica_per_nuovo_movimento(self, mock_adapter_for_provider):
        from gestione_finanziaria.services import sincronizza_conto_psd2

        provider = ProviderBancario.objects.create(
            nome="Provider PSD2 notifiche",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "test"},
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Connessione notifiche",
            external_connection_id="session-notify",
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto notifiche PSD2",
            provider=provider,
            connessione=connessione,
            external_account_id="account-notify",
            attivo=True,
        )
        adapter = Mock()
        adapter.saldo_conto.return_value = []
        adapter.movimenti_conto.return_value = [
            ProviderTransaction(
                data_contabile=date(2026, 5, 20),
                importo=Decimal("-42.50"),
                descrizione="Pagamento cancelleria",
                controparte="Cartoleria",
                provider_transaction_id="tx-notify-1",
            )
        ]
        mock_adapter_for_provider.return_value = adapter

        log = sincronizza_conto_psd2(conto)

        self.assertEqual(log.movimenti_inseriti, 1)
        movimento = MovimentoFinanziario.objects.get(provider_transaction_id="tx-notify-1")
        notifica = NotificaFinanziaria.objects.get(movimento_finanziario=movimento)
        self.assertEqual(notifica.tipo, "movimento_bancario")
        self.assertEqual(notifica.chiave_deduplica, f"movimento-bancario:{movimento.pk}")
        self.assertIn("Nuovo movimento bancario", notifica.titolo)
        self.assertIn("Pagamento cancelleria", notifica.messaggio)

    @patch("gestione_finanziaria.providers.adapter_for_provider")
    def test_psd2_account_sync_salta_movimento_esistente_senza_hash_normalizzato(self, mock_adapter_for_provider):
        from gestione_finanziaria.services import sincronizza_conto_psd2

        provider = ProviderBancario.objects.create(
            nome="Provider PSD2 dedup",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "test"},
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto PSD2 dedup",
            provider=provider,
            external_account_id="account-dedup",
            attivo=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.IMPORT_FILE,
            data_contabile=date(2026, 6, 3),
            importo=Decimal("-18.90"),
            descrizione="COMM.SU BONIFICI AREA SEPA",
            controparte="Banca Test",
            incide_su_saldo_banca=True,
        )
        adapter = Mock()
        adapter.saldo_conto.return_value = []
        adapter.movimenti_conto.return_value = [
            ProviderTransaction(
                data_contabile=date(2026, 6, 3),
                importo=Decimal("-18.90"),
                descrizione="Comm su bonifici   area sepa",
                controparte="Banca Test",
            )
        ]
        mock_adapter_for_provider.return_value = adapter

        log = sincronizza_conto_psd2(conto)

        self.assertEqual(log.movimenti_inseriti, 0)
        self.assertEqual(MovimentoFinanziario.objects.filter(conto=conto).count(), 1)


class MovimentoNotificationTests(TestCase):
    def test_importa_movimenti_da_file_crea_notifica_per_nuovo_movimento(self):
        provider = ProviderBancario.objects.create(
            nome="Import notifiche",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto import notifiche",
            provider=provider,
            attivo=True,
        )
        raw_csv = (
            "Data;Importo;Descrizione\n"
            "20/05/2026;-18,90;Commissioni bancarie\n"
        ).encode("utf-8")
        config = CsvImporterConfig(
            delimiter=";",
            ha_intestazione=True,
            colonna_data_contabile="Data",
            colonna_importo="Importo",
            colonna_descrizione="Descrizione",
        )

        risultato = importa_movimenti_da_file(
            parser=CsvImporter(config),
            raw_bytes=raw_csv,
            conto=conto,
            provider=provider,
            nome_file="movimenti.csv",
            riconcilia_automaticamente=False,
        )

        self.assertEqual(risultato.inseriti, 1)
        movimento = MovimentoFinanziario.objects.get(conto=conto)
        notifica = NotificaFinanziaria.objects.get(movimento_finanziario=movimento)
        self.assertEqual(notifica.tipo, "movimento_bancario")
        self.assertEqual(notifica.url, reverse("lista_movimenti_finanziari"))
        self.assertIn("import estratto conto", notifica.titolo)
        self.assertIn("Commissioni bancarie", notifica.messaggio)

    def test_importa_movimenti_da_file_salta_movimento_esistente_senza_hash_normalizzato(self):
        provider = ProviderBancario.objects.create(
            nome="Import dedup storico",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto dedup storico",
            provider=provider,
            attivo=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.IMPORT_FILE,
            data_contabile=date(2026, 6, 3),
            importo=Decimal("-18.90"),
            descrizione="COMM.SU BONIFICI AREA SEPA",
            controparte="Banca Test",
            incide_su_saldo_banca=True,
        )
        raw_csv = (
            "Data;Importo;Descrizione;Controparte\n"
            "03/06/2026;-18,90;Comm su bonifici   area sepa;Banca Test\n"
        ).encode("utf-8")
        config = CsvImporterConfig(
            delimiter=";",
            ha_intestazione=True,
            colonna_data_contabile="Data",
            colonna_importo="Importo",
            colonna_descrizione="Descrizione",
            colonna_controparte="Controparte",
        )

        risultato = importa_movimenti_da_file(
            parser=CsvImporter(config),
            raw_bytes=raw_csv,
            conto=conto,
            provider=provider,
            nome_file="dedup.csv",
        )

        self.assertEqual(risultato.inseriti, 0)
        self.assertEqual(risultato.duplicati, 1)
        self.assertEqual(MovimentoFinanziario.objects.filter(conto=conto).count(), 1)

    def test_importa_movimenti_da_file_salta_duplicati_equivalenti_nello_stesso_file(self):
        provider = ProviderBancario.objects.create(
            nome="Import dedup interno",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto dedup interno",
            provider=provider,
            attivo=True,
        )
        raw_csv = (
            "Data;Importo;Descrizione;Controparte\n"
            "03/06/2026;-18,90;COMM.SU BONIFICI AREA SEPA;Banca Test\n"
            "03/06/2026;-18,90;Comm su bonifici   area sepa;Banca Test\n"
        ).encode("utf-8")
        config = CsvImporterConfig(
            delimiter=";",
            ha_intestazione=True,
            colonna_data_contabile="Data",
            colonna_importo="Importo",
            colonna_descrizione="Descrizione",
            colonna_controparte="Controparte",
        )

        risultato = importa_movimenti_da_file(
            parser=CsvImporter(config),
            raw_bytes=raw_csv,
            conto=conto,
            provider=provider,
            nome_file="dedup-interno.csv",
        )

        self.assertEqual(risultato.inseriti, 1)
        self.assertEqual(risultato.duplicati, 1)
        self.assertEqual(MovimentoFinanziario.objects.filter(conto=conto).count(), 1)


class DocumentoFornitoreFormsetTests(TestCase):
    def test_popup_scadenze_movimento_field_keeps_all_bank_movements_available(self):
        from gestione_finanziaria.views import (
            _documento_fornitore_formset_class,
            _documento_fornitore_formset_kwargs,
        )

        movimenti = [
            MovimentoFinanziario(
                data_contabile=date(2026, 1, 1) + timedelta(days=index),
                importo=Decimal("-10.00"),
                descrizione=f"Movimento bancario {index}",
                incide_su_saldo_banca=True,
            )
            for index in range(125)
        ]
        MovimentoFinanziario.objects.bulk_create(movimenti)

        formset = _documento_fornitore_formset_class()(
            instance=DocumentoFornitore(),
            **_documento_fornitore_formset_kwargs(compact_movimenti=True),
        )

        queryset = formset.forms[0].fields["movimento_finanziario"].queryset
        self.assertEqual(queryset.count(), 125)


class Psd2ConnectionImportTests(TestCase):
    def _request(self):
        request = RequestFactory().get("/")
        request.session = self.client.session
        request._messages = FallbackStorage(request)
        return request

    def test_gocardless_movimenti_conto_ignora_transazioni_pending(self):
        from .providers.gocardless import GoCardlessBadAdapter, GoCardlessCredentials

        adapter = GoCardlessBadAdapter(GoCardlessCredentials(secret_id="id", secret_key="key"))
        adapter._get = Mock(
            return_value={
                "transactions": {
                    "booked": [
                        {
                            "transactionAmount": {"amount": "700.00", "currency": "EUR"},
                            "bookingDate": "2026-06-13",
                            "valueDate": "2026-06-13",
                            "remittanceInformationUnstructured": (
                                "BONIF. VS. FAVORE - BON.DA LABRIOLA FRANCESCO Caparra"
                            ),
                            "debtorName": "LABRIOLA FRANCESCO",
                            "transactionId": "booked-700",
                        }
                    ],
                    "pending": [
                        {
                            "transactionAmount": {"amount": "700.00", "currency": "EUR"},
                            "bookingDate": "2026-06-12",
                            "remittanceInformationUnstructured": "BONIF. VS. FAVORE",
                            "transactionId": "pending-700",
                        }
                    ],
                }
            }
        )

        movimenti = adapter.movimenti_conto("account-bpm")

        self.assertEqual(len(movimenti), 1)
        self.assertEqual(movimenti[0].provider_transaction_id, "booked-700")
        self.assertEqual(movimenti[0].descrizione, "BONIF. VS. FAVORE - BON.DA LABRIOLA FRANCESCO Caparra")

    def test_lista_connessioni_renderizza_layout_moderno_e_pulsante_rinnovo(self):
        from .views import lista_connessioni_bancarie

        provider = ProviderBancario.objects.create(
            nome="Enable Banking UI",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM",
            external_institution_id="Banco BPM|IT",
            stato=StatoConnessioneBancaria.SCADUTA,
            ultimo_errore="Consenso scaduto",
        )

        response = lista_connessioni_bancarie(self._request())
        content = response.content.decode("utf-8")

        self.assertIn("finance-page-head", content)
        self.assertIn("finance-modern-table", content)
        self.assertIn('class="btn btn-secondary btn-sm btn-icon-text"', content)
        self.assertIn("Rinnova consenso", content)
        self.assertIn("Consenso scaduto", content)

    def test_finalizzazione_psd2_mantiene_connessioni_multiple_stessa_banca(self):
        from .views import _finalizza_connessione_psd2

        provider = ProviderBancario.objects.create(
            nome="Enable Banking test multi",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        connessione_prima = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM - Conto principale",
            external_institution_id="Banco BPM|IT",
            external_connection_id="session-1",
        )
        connessione_seconda = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM - Secondo accesso",
            external_institution_id="Banco BPM|IT",
            external_connection_id="session-2",
        )
        account = ProviderAccount(
            external_account_id="account-shared-id",
            iban="IT60X0542811101000000123456",
            currency="EUR",
            owner_name="Scuola test",
            name="Conto operativo",
            account_type="CACC",
            identification_hash="hash-stesso-conto",
        )

        _finalizza_connessione_psd2(self._request(), connessione_prima, Mock(lista_conti=Mock(return_value=[account])))
        _finalizza_connessione_psd2(self._request(), connessione_seconda, Mock(lista_conti=Mock(return_value=[account])))

        conti = ContoBancario.objects.filter(
            provider=provider,
            external_account_id="account-shared-id",
        ).order_by("connessione_id")
        self.assertEqual(conti.count(), 2)
        self.assertEqual(conti[0].connessione, connessione_prima)
        self.assertEqual(conti[1].connessione, connessione_seconda)

    def test_finalizzazione_psd2_rinnovo_riusa_conto_esistente_se_account_id_cambia(self):
        from .views import _finalizza_connessione_psd2

        provider = ProviderBancario.objects.create(
            nome="Enable Banking test rinnovo",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM",
            external_institution_id="Banco BPM|IT",
            external_connection_id="session-renewed",
            stato=StatoConnessioneBancaria.SCADUTA,
        )
        conto = ContoBancario.objects.create(
            provider=provider,
            connessione=connessione,
            nome_conto="Banco BPM",
            external_account_id="old-enablebanking-id",
        )
        account = ProviderAccount(
            external_account_id="new-enablebanking-id",
            iban="IT67C0503437060000000003228",
            currency="EUR",
            owner_name="Scuola test",
            name="CC013228",
            account_type="CACC",
        )

        _finalizza_connessione_psd2(self._request(), connessione, Mock(lista_conti=Mock(return_value=[account])))

        self.assertEqual(ContoBancario.objects.filter(connessione=connessione).count(), 1)
        conto.refresh_from_db()
        self.assertEqual(conto.external_account_id, "new-enablebanking-id")
        self.assertEqual(conto.iban, "IT67C0503437060000000003228")
        self.assertEqual(conto.nome_conto, "Banco BPM")

    def test_finalizzazione_psd2_rinnovo_riconosce_conto_da_iban_anche_con_piu_conti(self):
        from .views import _finalizza_connessione_psd2

        provider = ProviderBancario.objects.create(
            nome="Enable Banking test iban",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM",
            external_institution_id="Banco BPM|IT",
            external_connection_id="session-iban",
        )
        conto = ContoBancario.objects.create(
            provider=provider,
            connessione=connessione,
            nome_conto="Banco BPM operativo",
            external_account_id="old-account-id",
            iban="IT67C0503437060000000003228",
        )
        ContoBancario.objects.create(
            provider=provider,
            connessione=connessione,
            nome_conto="Banco BPM secondario",
            external_account_id="second-account-id",
            iban="IT67C0503437060000000009999",
        )
        account = ProviderAccount(
            external_account_id="new-account-id",
            iban="IT67 C050 3437 0600 0000 0003 228",
            currency="EUR",
            owner_name="Scuola test",
            name="CC013228",
            account_type="CACC",
        )

        _finalizza_connessione_psd2(self._request(), connessione, Mock(lista_conti=Mock(return_value=[account])))

        self.assertEqual(ContoBancario.objects.filter(connessione=connessione).count(), 2)
        conto.refresh_from_db()
        self.assertEqual(conto.external_account_id, "new-account-id")
        self.assertEqual(conto.nome_conto, "Banco BPM operativo")
        self.assertFalse(ContoBancario.objects.filter(connessione=connessione, external_account_id="old-account-id").exists())

    def test_finalizzazione_psd2_importa_carte_enablebanking_come_prepagate(self):
        from .views import _finalizza_connessione_psd2

        provider = ProviderBancario.objects.create(
            nome="Enable Banking test card",
            tipo=TipoProviderBancario.PSD2,
            configurazione={"adapter": "enablebanking"},
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Unicredit - Carte",
            external_institution_id="Unicredit|IT",
            external_connection_id="session-card",
        )
        account = ProviderAccount(
            external_account_id="card-account-id",
            currency="EUR",
            owner_name="Scuola test",
            name="Carta scuola",
            account_type="CARD",
            account_product="Carta prepagata business",
            identification_hash="hash-carta-prepagata",
        )

        _finalizza_connessione_psd2(self._request(), connessione, Mock(lista_conti=Mock(return_value=[account])))

        conto = ContoBancario.objects.get(external_account_id="card-account-id")
        self.assertEqual(conto.connessione, connessione)
        self.assertEqual(conto.tipo_conto, TipoContoFinanziario.CARTA_PREPAGATA)
        self.assertEqual(conto.external_account_type, "CARD")
        self.assertEqual(conto.external_account_product, "Carta prepagata business")
        self.assertEqual(conto.external_account_hash, "hash-carta-prepagata")
        self.assertEqual(conto.banca, "Unicredit")

    @patch("gestione_finanziaria.views.adapter_for_provider")
    def test_rinnova_connessione_psd2_riusa_connessione_esistente(self, mock_adapter_for_provider):
        from .views import rinnova_connessione_psd2

        provider = ProviderBancario.objects.create(
            nome="Provider rinnovo PSD2",
            tipo=TipoProviderBancario.PSD2,
            configurazione={
                "adapter": "gocardless_bad",
                "secret_id": "secret-id",
                "secret_key_cifrata": "secret-key",
            },
        )
        connessione = ConnessioneBancaria.objects.create(
            provider=provider,
            etichetta="Banco BPM",
            external_institution_id="Banco BPM|IT",
            external_connection_id="session-old",
            stato=StatoConnessioneBancaria.SCADUTA,
            ultimo_errore="Consenso scaduto",
        )
        adapter = Mock()
        expires_at = timezone.now() + timedelta(days=90)
        adapter.crea_connessione.return_value = ProviderConnectionInfo(
            external_connection_id=f"arboris-{connessione.pk}",
            authorization_url="https://bank.example/auth",
            institution_id="Banco BPM|IT",
            expires_at=expires_at,
        )
        mock_adapter_for_provider.return_value = adapter

        request = RequestFactory().post("/")
        request.session = self.client.session
        request._messages = FallbackStorage(request)

        response = rinnova_connessione_psd2(request, connessione.pk)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "https://bank.example/auth")
        connessione.refresh_from_db()
        self.assertEqual(connessione.external_connection_id, f"arboris-{connessione.pk}")
        self.assertEqual(connessione.stato, StatoConnessioneBancaria.SCADUTA)
        self.assertEqual(connessione.ultimo_errore, "")
        self.assertEqual(connessione.consenso_scadenza, expires_at)
        adapter.crea_connessione.assert_called_once()

    def test_enablebanking_usa_account_details_restituiti_da_post_sessions(self):
        adapter = EnableBankingAdapter(
            EnableBankingCredentials(
                app_id="app-test",
                private_key_pem="private-key-non-usata",
            )
        )

        def fake_request(method, path, **_kwargs):
            if method == "POST" and path == "/sessions":
                return {
                    "session_id": "session-rich",
                    "accounts": [
                        {
                            "uid": "card-account-id",
                            "account_id": {"iban": "IT60X0542811101000000123456"},
                            "cash_account_type": "CARD",
                            "product": "Carta prepagata",
                            "details": "Prepaid business",
                            "name": "Scuola test",
                            "currency": "EUR",
                            "identification_hash": "hash-card",
                        }
                    ],
                }
            raise AssertionError("La lista conti deve usare i dati gia' restituiti da /sessions.")

        adapter._request = fake_request

        session_id = adapter.scambia_codice_sessione("auth-code")
        accounts = adapter.lista_conti(session_id)

        self.assertEqual(session_id, "session-rich")
        self.assertEqual(len(accounts), 1)
        self.assertEqual(accounts[0].external_account_id, "card-account-id")
        self.assertEqual(accounts[0].account_type, "CARD")
        self.assertEqual(accounts[0].account_product, "Carta prepagata - Prepaid business")
        self.assertEqual(accounts[0].identification_hash, "hash-card")

    def test_enablebanking_arricchisce_sessione_minima_con_account_details(self):
        adapter = EnableBankingAdapter(
            EnableBankingCredentials(
                app_id="app-test",
                private_key_pem="private-key-non-usata",
            )
        )

        def fake_request(method, path, **_kwargs):
            if method == "GET" and path == "/sessions/session-minimal":
                return {
                    "accounts": ["card-account-id"],
                    "accounts_data": [
                        {
                            "uid": "card-account-id",
                            "identification_hash": "hash-card",
                        }
                    ],
                }
            if method == "GET" and path == "/accounts/card-account-id/details":
                return {
                    "uid": "card-account-id",
                    "cash_account_type": "CARD",
                    "product": "Carta prepagata",
                    "details": "Prepaid business",
                    "currency": "EUR",
                    "identification_hash": "hash-card",
                }
            raise AssertionError(f"Chiamata inattesa: {method} {path}")

        adapter._request = fake_request

        accounts = adapter.lista_conti("session-minimal")

        self.assertEqual(len(accounts), 1)
        self.assertEqual(accounts[0].account_type, "CARD")
        self.assertEqual(accounts[0].account_product, "Carta prepagata - Prepaid business")
        self.assertEqual(accounts[0].identification_hash, "hash-card")


class MovimentoCategoriaInlineTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="finanza-inline@example.com",
            email="finanza-inline@example.com",
            password="Password123!",
        )
        SistemaUtentePermessi.objects.create(
            user=self.user,
            permesso_gestione_finanziaria=LivelloPermesso.GESTIONE,
        )
        self.client.force_login(self.user)

    def test_aggiorna_categoria_movimento_da_lista(self):
        categoria_iniziale = CategoriaFinanziaria.objects.create(
            nome="Da rivedere",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        categoria_nuova = CategoriaFinanziaria.objects.create(
            nome="Commissioni bancarie",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 16),
            importo=Decimal("-12.00"),
            descrizione="Spese bancarie",
            categoria=categoria_iniziale,
            categorizzazione_automatica=True,
        )

        response = self.client.post(
            reverse("aggiorna_categoria_movimento", args=[movimento.pk]),
            {"categoria": str(categoria_nuova.pk)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["category_id"], str(categoria_nuova.pk))
        movimento.refresh_from_db()
        self.assertEqual(movimento.categoria, categoria_nuova)
        self.assertFalse(movimento.categorizzazione_automatica)
        self.assertEqual(movimento.categorizzato_da, self.user)
        self.assertIsNotNone(movimento.categorizzato_il)

    def test_aggiorna_nome_conto_bancario_da_lista(self):
        conto = ContoBancario.objects.create(nome_conto="01e65d55-d434-4928-bd80-c51ea68d4d07")

        response = self.client.post(
            reverse("aggiorna_nome_conto_bancario", args=[conto.pk]),
            {"nome_conto": "Banco BPM - Principale"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["account_id"], str(conto.pk))
        self.assertEqual(payload["account_name"], "Banco BPM - Principale")
        conto.refresh_from_db()
        self.assertEqual(conto.nome_conto, "Banco BPM - Principale")

    def test_aggiorna_nome_conto_bancario_rifiuta_nome_vuoto(self):
        conto = ContoBancario.objects.create(nome_conto="Conto operativo")

        response = self.client.post(
            reverse("aggiorna_nome_conto_bancario", args=[conto.pk]),
            {"nome_conto": "   "},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        conto.refresh_from_db()
        self.assertEqual(conto.nome_conto, "Conto operativo")

    def test_lista_movimenti_usa_click_destro_per_rinominare_conto(self):
        conto = ContoBancario.objects.create(nome_conto="Conto PSD2 grezzo")
        MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 5, 16),
            importo=Decimal("-12.00"),
            descrizione="Spese bancarie",
        )

        response = self.client.get(reverse("lista_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "data-movement-account-cell")
        self.assertContains(response, reverse("aggiorna_nome_conto_bancario", args=[conto.pk]))
        self.assertContains(response, "data-account-name")
        self.assertNotContains(response, "finance-account-edit-link")

    def test_lista_movimenti_mostra_pulsante_pulizia_duplicati(self):
        response = self.client.get(reverse("lista_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pulisci duplicati")
        self.assertContains(response, reverse("pulizia_duplicati_movimenti_finanziari"))

    def test_pulizia_duplicati_movimenti_mostra_gruppi_preselezionati(self):
        conto = ContoBancario.objects.create(nome_conto="Conto duplicati")
        movimento_keep = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 10),
            importo=Decimal("-42.00"),
            descrizione="Pagamento POS",
            controparte="Cartoleria",
        )
        movimento_duplicato = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 10),
            importo=Decimal("-42.00"),
            descrizione="Pagamento POS",
            controparte="Cartoleria",
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 11),
            importo=Decimal("-42.00"),
            descrizione="Pagamento POS",
            controparte="Cartoleria",
        )

        response = self.client.get(reverse("pulizia_duplicati_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Gruppo duplicati #1")
        self.assertContains(response, "Conservato")
        self.assertContains(response, "Duplicato")
        self.assertContains(response, "stessa data, importo, descrizione, controparte e IBAN")
        content = response.content.decode()
        self.assertIn(f'value="{movimento_keep.pk}" data-bulk-checkbox data-duplicate-delete disabled', content)
        self.assertIn(f'value="{movimento_duplicato.pk}" data-bulk-checkbox data-duplicate-delete checked', content)

    def test_pulizia_duplicati_movimenti_propone_causale_generica_con_data_vicina(self):
        conto = ContoBancario.objects.create(nome_conto="Banco BPM")
        movimento_generico = MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 12),
            importo=Decimal("700.00"),
            descrizione="BONIF. VS. FAVORE",
            incide_su_saldo_banca=True,
        )
        movimento_dettagliato = MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 13),
            importo=Decimal("700.00"),
            descrizione="BONIF. VS. FAVORE - BON.DA LABRIOLA FRANCESCO Caparra iscrizione A.S. 2026/27",
            incide_su_saldo_banca=True,
        )

        response = self.client.get(reverse("pulizia_duplicati_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "causale assente o generica")
        content = response.content.decode()
        self.assertIn(f'value="{movimento_generico.pk}" data-bulk-checkbox data-duplicate-delete checked', content)
        self.assertIn(f'value="{movimento_dettagliato.pk}" data-bulk-checkbox data-duplicate-delete disabled', content)

    def test_pulizia_duplicati_movimenti_propone_import_file_e_banca_simili(self):
        conto = ContoBancario.objects.create(nome_conto="Banco BPM")
        movimento_importato = MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.IMPORT_FILE,
            data_contabile=date(2026, 6, 12),
            importo=Decimal("-42.00"),
            descrizione="Pagamento POS 1234",
            hash_deduplica="hash-import-pos-1234",
            incide_su_saldo_banca=True,
        )
        movimento_bancario = MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 13),
            importo=Decimal("-42.00"),
            descrizione="Transazione carta 1234",
            provider_transaction_id="bank-pos-1234",
            incide_su_saldo_banca=True,
        )

        response = self.client.get(reverse("pulizia_duplicati_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "origine diversa: import file / banca")
        self.assertContains(response, "Import estratto conto")
        self.assertContains(response, "Movimento bancario")
        content = response.content.decode()
        self.assertIn(f'value="{movimento_importato.pk}" data-bulk-checkbox data-duplicate-delete checked', content)
        self.assertIn(f'value="{movimento_bancario.pk}" data-bulk-checkbox data-duplicate-delete disabled', content)

    def test_pulizia_duplicati_movimenti_non_propone_import_banca_ambigui(self):
        conto = ContoBancario.objects.create(nome_conto="Banco BPM")
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.IMPORT_FILE,
            data_contabile=date(2026, 6, 12),
            importo=Decimal("-42.00"),
            descrizione="Pagamento POS 1234",
            incide_su_saldo_banca=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 13),
            importo=Decimal("-42.00"),
            descrizione="Transazione carta 1234",
            incide_su_saldo_banca=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 14),
            importo=Decimal("-42.00"),
            descrizione="Transazione carta 5678",
            incide_su_saldo_banca=True,
        )

        response = self.client.get(reverse("pulizia_duplicati_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "origine diversa: import file / banca")

    def test_pulizia_duplicati_movimenti_non_propone_causale_generica_ambigua(self):
        conto = ContoBancario.objects.create(nome_conto="Banco BPM")
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 12),
            importo=Decimal("700.00"),
            descrizione="BONIF. VS. FAVORE",
            incide_su_saldo_banca=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 13),
            importo=Decimal("700.00"),
            descrizione="BONIF. VS. FAVORE - BON.DA LABRIOLA FRANCESCO Caparra iscrizione A.S. 2026/27",
            incide_su_saldo_banca=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 6, 14),
            importo=Decimal("700.00"),
            descrizione="BONIF. VS. FAVORE - BON.DA ROSSI MARIO Retta giugno",
            incide_su_saldo_banca=True,
        )

        response = self.client.get(reverse("pulizia_duplicati_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "causale assente o generica")

    def test_pulizia_duplicati_movimenti_elimina_solo_i_selezionati(self):
        conto = ContoBancario.objects.create(nome_conto="Conto da pulire")
        MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 10),
            importo=Decimal("-100.00"),
            descrizione="Bonifico fornitore",
            controparte="Fornitore SRL",
            incide_su_saldo_banca=True,
        )
        duplicato_da_eliminare = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 10),
            importo=Decimal("-100.00"),
            descrizione="Bonifico fornitore",
            controparte="Fornitore SRL",
            incide_su_saldo_banca=True,
        )
        duplicato_deselezionato = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 10),
            importo=Decimal("-100.00"),
            descrizione="Bonifico fornitore",
            controparte="Fornitore SRL",
            incide_su_saldo_banca=True,
        )

        response = self.client.post(
            reverse("pulizia_duplicati_movimenti_finanziari"),
            {"selected_ids": [str(duplicato_da_eliminare.pk)]},
        )

        self.assertRedirects(response, reverse("lista_movimenti_finanziari"))
        self.assertFalse(MovimentoFinanziario.objects.filter(pk=duplicato_da_eliminare.pk).exists())
        self.assertTrue(MovimentoFinanziario.objects.filter(pk=duplicato_deselezionato.pk).exists())
        conto.refresh_from_db()
        self.assertEqual(conto.saldo_corrente, Decimal("-200.00"))

    def test_pulizia_duplicati_movimenti_forza_conservato_e_trasferisce_collegamenti(self):
        conto = ContoBancario.objects.create(nome_conto="Conto riconciliato")
        movimento_riconciliato = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 10),
            importo=Decimal("-100.00"),
            descrizione="Bonifico fornitore",
            controparte="Fornitore SRL",
            incide_su_saldo_banca=True,
            stato_riconciliazione=StatoRiconciliazione.RICONCILIATO,
        )
        movimento_da_mantenere = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 6, 10),
            importo=Decimal("-100.00"),
            descrizione="Bonifico fornitore",
            controparte="Fornitore SRL",
            incide_su_saldo_banca=True,
        )
        fornitore = Fornitore.objects.create(denominazione="Fornitore SRL")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FOR-001",
            data_documento=date(2026, 6, 1),
            totale=Decimal("100.00"),
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 6, 10),
            importo_previsto=Decimal("100.00"),
            importo_pagato=Decimal("100.00"),
            data_pagamento=date(2026, 6, 10),
            conto_bancario=conto,
            movimento_finanziario=movimento_riconciliato,
        )
        pagamento = PagamentoFornitore.objects.create(
            scadenza=scadenza,
            movimento_finanziario=movimento_riconciliato,
            data_pagamento=date(2026, 6, 10),
            importo=Decimal("100.00"),
            metodo=MetodoPagamentoFornitore.BANCA,
            conto_bancario=conto,
            creato_da=self.user,
        )

        response = self.client.get(reverse("pulizia_duplicati_movimenti_finanziari"))
        self.assertEqual(response.status_code, 200)
        gruppo = response.context["duplicate_groups"][0]
        self.assertEqual(gruppo["keep"].pk, movimento_riconciliato.pk)
        self.assertContains(response, "forzare il movimento conservato")

        response = self.client.post(
            reverse("pulizia_duplicati_movimenti_finanziari"),
            {
                f"keep_{gruppo['key']}": str(movimento_da_mantenere.pk),
                "selected_ids": [str(movimento_da_mantenere.pk)],
            },
        )

        self.assertRedirects(response, reverse("lista_movimenti_finanziari"))
        self.assertFalse(MovimentoFinanziario.objects.filter(pk=movimento_riconciliato.pk).exists())
        self.assertTrue(MovimentoFinanziario.objects.filter(pk=movimento_da_mantenere.pk).exists())
        movimento_da_mantenere.refresh_from_db()
        scadenza.refresh_from_db()
        pagamento.refresh_from_db()
        self.assertEqual(movimento_da_mantenere.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(scadenza.movimento_finanziario, movimento_da_mantenere)
        self.assertEqual(pagamento.movimento_finanziario, movimento_da_mantenere)

    def test_lista_movimenti_prepara_dropdown_categorie_gerarchico(self):
        padre = CategoriaFinanziaria.objects.create(
            nome="Spese di Gestione",
            tipo=TipoCategoriaFinanziaria.SPESA,
            icona="briefcase",
        )
        figlia = CategoriaFinanziaria.objects.create(
            nome="Utenze e Servizi",
            tipo=TipoCategoriaFinanziaria.SPESA,
            parent=padre,
            icona="bolt",
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 16),
            importo=Decimal("-82.96"),
            descrizione="Pagamento utenza",
            categoria=figlia,
        )

        response = self.client.get(reverse("lista_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="movement-category-options-template"', html=False)
        self.assertContains(response, 'data-category-name="Spese di Gestione"', html=False)
        self.assertContains(response, 'data-category-has-children="1"', html=False)
        self.assertContains(response, 'data-category-name="Utenze e Servizi"', html=False)
        self.assertContains(response, 'data-category-level="1"', html=False)
        self.assertContains(response, 'data-category-parent="Spese di Gestione"', html=False)
        self.assertContains(response, 'data-category-icon="bolt"', html=False)


class CategoriaFinanziariaMaintenanceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="finanza-categorie@example.com",
            email="finanza-categorie@example.com",
            password="Password123!",
        )
        SistemaUtentePermessi.objects.create(
            user=self.user,
            permesso_gestione_finanziaria=LivelloPermesso.GESTIONE,
        )
        self.client.force_login(self.user)

    def _crea_categoria_con_collegamenti(self):
        sorgente = CategoriaFinanziaria.objects.create(
            nome="Utenze duplicate",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        destinazione = CategoriaFinanziaria.objects.create(
            nome="Utenze",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        figlia = CategoriaFinanziaria.objects.create(
            nome="Energia elettrica",
            tipo=TipoCategoriaFinanziaria.SPESA,
            parent=sorgente,
        )
        regola = RegolaCategorizzazione.objects.create(
            nome="Regola utenze",
            condizione_tipo=CondizioneRegolaCategorizzazione.DESCRIZIONE_CONTIENE,
            pattern="bolletta",
            categoria_da_assegnare=sorgente,
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 20),
            importo=Decimal("-80.00"),
            descrizione="Bolletta energia",
            categoria=sorgente,
            categorizzazione_automatica=True,
            regola_categorizzazione=regola,
        )
        fornitore = Fornitore.objects.create(
            denominazione="Fornitore energia",
            categoria_spesa=sorgente,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            categoria_spesa=sorgente,
            numero_documento="FT-1",
            data_documento=date(2026, 5, 21),
        )
        voce_budget = VoceBudgetRicorrente.objects.create(
            nome="Budget utenze",
            tipo=TipoVoceBudget.USCITA,
            categoria=sorgente,
            importo=Decimal("100.00"),
            data_inizio=date(2026, 1, 1),
        )
        piano = PianoRatealeSpesa.objects.create(
            descrizione="Piano utenze",
            categoria=sorgente,
            importo_totale=Decimal("300.00"),
            data_prima_scadenza=date(2026, 6, 1),
        )
        spesa = SpesaOperativa.objects.create(
            descrizione="Rata utenze",
            categoria=sorgente,
            data_scadenza=date(2026, 6, 1),
            importo_previsto=Decimal("100.00"),
        )
        return {
            "sorgente": sorgente,
            "destinazione": destinazione,
            "figlia": figlia,
            "regola": regola,
            "movimento": movimento,
            "fornitore": fornitore,
            "documento": documento,
            "voce_budget": voce_budget,
            "piano": piano,
            "spesa": spesa,
        }

    def test_trasferimento_categoria_sposta_collegamenti(self):
        dati = self._crea_categoria_con_collegamenti()

        get_response = self.client.get(reverse("trasferisci_categoria_finanziaria", args=[dati["sorgente"].pk]))
        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, 'id="id_categoria_destinazione"', html=False)
        self.assertContains(get_response, dati["destinazione"].percorso_label)

        response = self.client.post(
            reverse("trasferisci_categoria_finanziaria", args=[dati["sorgente"].pk]),
            {
                "categoria_destinazione": str(dati["destinazione"].pk),
                "conferma_trasferimento": "1",
            },
        )

        self.assertRedirects(response, reverse("lista_categorie_finanziarie"))
        self.assertTrue(CategoriaFinanziaria.objects.filter(pk=dati["sorgente"].pk).exists())
        for key in ["figlia", "regola", "movimento", "fornitore", "documento", "voce_budget", "piano", "spesa"]:
            dati[key].refresh_from_db()
        self.assertEqual(dati["figlia"].parent, dati["destinazione"])
        self.assertEqual(dati["regola"].categoria_da_assegnare, dati["destinazione"])
        self.assertEqual(dati["movimento"].categoria, dati["destinazione"])
        self.assertEqual(dati["fornitore"].categoria_spesa, dati["destinazione"])
        self.assertEqual(dati["documento"].categoria_spesa, dati["destinazione"])
        self.assertEqual(dati["voce_budget"].categoria, dati["destinazione"])
        self.assertEqual(dati["piano"].categoria, dati["destinazione"])
        self.assertEqual(dati["spesa"].categoria, dati["destinazione"])

    def test_eliminazione_categoria_richiede_conferma_forte(self):
        dati = self._crea_categoria_con_collegamenti()

        response = self.client.post(
            reverse("elimina_categoria_finanziaria", args=[dati["sorgente"].pk]),
            {"conferma_eliminazione": "ELIMINA altra categoria"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(CategoriaFinanziaria.objects.filter(pk=dati["sorgente"].pk).exists())
        dati["movimento"].refresh_from_db()
        self.assertEqual(dati["movimento"].categoria, dati["sorgente"])

    def test_eliminazione_categoria_scollega_dati_ed_elimina_regole(self):
        dati = self._crea_categoria_con_collegamenti()

        response = self.client.post(
            reverse("elimina_categoria_finanziaria", args=[dati["sorgente"].pk]),
            {"conferma_eliminazione": f"ELIMINA {dati['sorgente'].nome}"},
        )

        self.assertRedirects(response, reverse("lista_categorie_finanziarie"))
        self.assertFalse(CategoriaFinanziaria.objects.filter(pk=dati["sorgente"].pk).exists())
        self.assertFalse(RegolaCategorizzazione.objects.filter(pk=dati["regola"].pk).exists())
        for key in ["figlia", "movimento", "fornitore", "documento", "voce_budget", "piano", "spesa"]:
            dati[key].refresh_from_db()
        self.assertIsNone(dati["figlia"].parent)
        self.assertIsNone(dati["movimento"].categoria)
        self.assertFalse(dati["movimento"].categorizzazione_automatica)
        self.assertIsNone(dati["movimento"].regola_categorizzazione)
        self.assertIsNone(dati["fornitore"].categoria_spesa)
        self.assertIsNone(dati["documento"].categoria_spesa)
        self.assertIsNone(dati["voce_budget"].categoria)
        self.assertIsNone(dati["piano"].categoria)
        self.assertIsNone(dati["spesa"].categoria)


class FusioneContiBancariTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="finanza-fusione@example.com",
            email="finanza-fusione@example.com",
            password="Password123!",
        )
        SistemaUtentePermessi.objects.create(
            user=self.user,
            permesso_gestione_finanziaria=LivelloPermesso.GESTIONE,
        )
        self.client.force_login(self.user)

    def test_fusione_conti_sposta_riferimenti_e_disattiva_sorgente(self):
        provider, _created = ProviderBancario.objects.get_or_create(
            nome="Enable Banking Fusione Test",
            defaults={
                "tipo": TipoProviderBancario.PSD2,
                "attivo": True,
            },
        )
        conto_sorgente = ContoBancario.objects.create(
            nome_conto="Banco BPM storico import manuale",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
            attivo=True,
        )
        conto_destinazione = ContoBancario.objects.create(
            nome_conto="Banco BPM PSD2",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
            provider=provider,
            external_account_id="psd2-account-id",
            attivo=True,
        )
        saldo = SaldoConto.objects.create(
            conto=conto_sorgente,
            data_riferimento=timezone.make_aware(datetime(2026, 4, 1, 23, 59)),
            saldo_contabile=Decimal("1000.00"),
            fonte=FonteSaldo.MANUALE,
        )
        movimento = MovimentoFinanziario.objects.create(
            conto=conto_sorgente,
            canale=CanaleMovimento.BANCA,
            data_contabile=date(2026, 4, 2),
            importo=Decimal("-100.00"),
            descrizione="Pagamento fornitore",
            controparte="Fornitore SRL",
            incide_su_saldo_banca=True,
            hash_deduplica=calcola_hash_deduplica_movimento(
                conto_id=conto_sorgente.pk,
                data_contabile=date(2026, 4, 2),
                importo=Decimal("-100.00"),
                descrizione="Pagamento fornitore",
                controparte="Fornitore SRL",
                iban_controparte="",
            ),
        )
        log = SincronizzazioneLog.objects.create(
            conto=conto_sorgente,
            tipo_operazione="import_file",
            esito=EsitoSincronizzazione.OK,
            movimenti_inseriti=1,
        )

        response = self.client.post(
            reverse("fondi_conti_bancari"),
            {
                "azione": "conferma",
                "conto_sorgente": str(conto_sorgente.pk),
                "conto_destinazione": str(conto_destinazione.pk),
                "conferma_operazione": "on",
            },
        )

        self.assertRedirects(response, reverse("lista_conti_bancari"))
        movimento.refresh_from_db()
        saldo.refresh_from_db()
        log.refresh_from_db()
        conto_sorgente.refresh_from_db()
        conto_destinazione.refresh_from_db()

        self.assertEqual(movimento.conto, conto_destinazione)
        self.assertEqual(
            movimento.hash_deduplica,
            calcola_hash_deduplica_movimento(
                conto_id=conto_destinazione.pk,
                data_contabile=movimento.data_contabile,
                importo=movimento.importo,
                descrizione=movimento.descrizione,
                controparte=movimento.controparte,
                iban_controparte=movimento.iban_controparte,
            ),
        )
        self.assertEqual(saldo.conto, conto_destinazione)
        self.assertEqual(log.conto, conto_destinazione)
        self.assertFalse(conto_sorgente.attivo)
        self.assertIn("Fuso nel conto", conto_sorgente.note)
        self.assertEqual(conto_sorgente.saldo_corrente, Decimal("0.00"))
        self.assertEqual(conto_destinazione.saldo_corrente, Decimal("900.00"))

    def test_fusione_conti_mostra_anteprima_con_possibili_duplicati(self):
        conto_sorgente = ContoBancario.objects.create(nome_conto="Conto importato")
        conto_destinazione = ContoBancario.objects.create(nome_conto="Conto PSD2")
        MovimentoFinanziario.objects.create(
            conto=conto_sorgente,
            data_contabile=date(2026, 4, 2),
            importo=Decimal("-100.00"),
            descrizione="Pagamento fornitore",
            controparte="Fornitore SRL",
        )
        MovimentoFinanziario.objects.create(
            conto=conto_destinazione,
            data_contabile=date(2026, 4, 2),
            importo=Decimal("-100.00"),
            descrizione="Pagamento fornitore",
            controparte="Fornitore SRL",
        )

        response = self.client.post(
            reverse("fondi_conti_bancari"),
            {
                "azione": "anteprima",
                "conto_sorgente": str(conto_sorgente.pk),
                "conto_destinazione": str(conto_destinazione.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Possibili duplicati")
        self.assertContains(response, "stessa data, importo e descrizione")

    def test_fusione_conti_con_duplicati_provider_prosegue_saltando_i_movimenti_doppi(self):
        categoria = CategoriaFinanziaria.objects.create(
            nome="Spese bancarie",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        conto_sorgente = ContoBancario.objects.create(nome_conto="Banco BPM duplicato")
        conto_destinazione = ContoBancario.objects.create(nome_conto="Banco BPM PSD2")
        movimento_duplicato_sorgente = MovimentoFinanziario.objects.create(
            conto=conto_sorgente,
            data_contabile=date(2026, 5, 15),
            importo=Decimal("-0.80"),
            descrizione="Commissioni",
            provider_transaction_id="tx-duplicate",
            categoria=categoria,
            stato_riconciliazione=StatoRiconciliazione.RICONCILIATO,
        )
        movimento_duplicato_destinazione = MovimentoFinanziario.objects.create(
            conto=conto_destinazione,
            data_contabile=date(2026, 5, 15),
            importo=Decimal("-0.80"),
            descrizione="Commissioni",
            provider_transaction_id="tx-duplicate",
        )
        movimento_unico = MovimentoFinanziario.objects.create(
            conto=conto_sorgente,
            data_contabile=date(2026, 5, 16),
            importo=Decimal("-100.00"),
            descrizione="Pagamento unico",
            provider_transaction_id="tx-unique",
        )
        movimento_duplicato_sorgente_id = movimento_duplicato_sorgente.pk

        response = self.client.post(
            reverse("fondi_conti_bancari"),
            {
                "azione": "conferma",
                "conto_sorgente": str(conto_sorgente.pk),
                "conto_destinazione": str(conto_destinazione.pk),
                "conferma_operazione": "on",
            },
        )

        self.assertRedirects(response, reverse("lista_conti_bancari"))
        conto_sorgente.refresh_from_db()
        movimento_duplicato_destinazione.refresh_from_db()
        movimento_unico.refresh_from_db()

        self.assertFalse(conto_sorgente.attivo)
        self.assertIn("Movimenti duplicati assorbiti: 1", conto_sorgente.note)
        self.assertFalse(MovimentoFinanziario.objects.filter(pk=movimento_duplicato_sorgente_id).exists())
        self.assertEqual(movimento_duplicato_destinazione.conto, conto_destinazione)
        self.assertEqual(movimento_duplicato_destinazione.categoria, categoria)
        self.assertEqual(movimento_duplicato_destinazione.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(movimento_unico.conto, conto_destinazione)
        self.assertEqual(MovimentoFinanziario.objects.filter(conto=conto_destinazione).count(), 2)

    def test_fusione_conti_richiede_conti_diversi(self):
        conto = ContoBancario.objects.create(nome_conto="Conto operativo")

        response = self.client.post(
            reverse("fondi_conti_bancari"),
            {
                "azione": "anteprima",
                "conto_sorgente": str(conto.pk),
                "conto_destinazione": str(conto.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Scegli due conti diversi")


class MovimentoRiconciliazioneLayoutTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username="admin-riconciliazione@example.com",
            email="admin-riconciliazione@example.com",
            password="Password123!",
        )
        self.client.force_login(self.user)

    def _crea_rata_corrente_con_familiare(self, *, importo=Decimal("300.00")):
        relazione = RelazioneFamiliare.objects.create(relazione="Genitore")
        familiare = Familiare.objects.create(
            relazione_familiare=relazione,
            nome="Mario",
            cognome="Rossi",
        )
        studente = Studente.objects.create(nome="Luca", cognome="Rossi")
        StudenteFamiliare.objects.create(
            studente=studente,
            familiare=familiare,
            relazione_familiare=relazione,
            attivo=True,
        )
        anno = AnnoScolastico.objects.create(
            nome_anno_scolastico="2025/2026",
            data_inizio=date(2025, 9, 1),
            data_fine=date(2026, 6, 30),
        )
        stato_iscrizione = StatoIscrizione.objects.create(
            stato_iscrizione="Attiva",
            ordine=1,
            attiva=True,
        )
        condizione = CondizioneIscrizione.objects.create(
            anno_scolastico=anno,
            nome_condizione_iscrizione="Retta standard",
            numero_mensilita_default=10,
            mese_prima_retta=9,
            giorno_scadenza_rate=10,
        )
        iscrizione = Iscrizione.objects.create(
            studente=studente,
            anno_scolastico=anno,
            stato_iscrizione=stato_iscrizione,
            condizione_iscrizione=condizione,
            data_iscrizione=date(2025, 9, 1),
        )
        return RataIscrizione.objects.create(
            iscrizione=iscrizione,
            numero_rata=1,
            mese_riferimento=9,
            anno_riferimento=2025,
            importo_dovuto=importo,
            importo_finale=importo,
            data_scadenza=date(2025, 9, 10),
        )

    def test_lista_riconciliazione_usa_layout_finanziario_moderno(self):
        conto = ContoBancario.objects.create(nome_conto="Conto operativo")
        movimento = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 5, 16),
            importo=Decimal("410.00"),
            descrizione="Bonifico retta Scamporlino",
            controparte="Scamporlino",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        response = self.client.get(reverse("lista_movimenti_da_riconciliare"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "finance-page-head")
        self.assertContains(response, "finance-reconciliation-summary")
        self.assertContains(response, "finance-filter-toolbar")
        self.assertContains(response, "finance-reconciliation-list-panel")
        self.assertContains(response, "finance-reconciliation-list-table")
        self.assertContains(
            response,
            f'data-row-href="{reverse("riconcilia_movimento", args=[movimento.pk])}?next=/gestione-finanziaria/riconciliazione/"',
        )
        self.assertContains(response, 'data-floating-text="Collega pagamento"')

    def test_riconciliazione_movimento_usa_layout_finanziario_moderno(self):
        conto = ContoBancario.objects.create(nome_conto="Conto operativo")
        movimento = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 5, 16),
            importo=Decimal("410.00"),
            descrizione="Bonifico retta Scamporlino",
            controparte="Scamporlino",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        response = self.client.get(reverse("riconcilia_movimento", args=[movimento.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "finance-reconciliation-page-head")
        self.assertContains(response, "finance-reconciliation-summary")
        self.assertContains(response, "finance-reconciliation-candidates-panel")
        self.assertContains(response, "finance-reconciliation-action-bar")
        self.assertContains(response, "formnovalidate")

    def test_riconciliazione_movimento_conserva_pagina_di_ritorno(self):
        conto = ContoBancario.objects.create(nome_conto="Conto operativo")
        movimento = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 5, 16),
            importo=Decimal("410.00"),
            descrizione="Bonifico retta Scamporlino",
            controparte="Scamporlino",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        return_url = reverse("lista_movimenti_finanziari")

        response = self.client.get(reverse("riconcilia_movimento", args=[movimento.pk]), {"next": return_url})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'href="{return_url}"')
        self.assertContains(response, f'name="next" value="{return_url}"')

    def test_riconciliazione_movimento_post_torna_alla_pagina_origine(self):
        conto = ContoBancario.objects.create(nome_conto="Conto operativo")
        movimento = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 5, 16),
            importo=Decimal("410.00"),
            descrizione="Bonifico retta Scamporlino",
            controparte="Scamporlino",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        return_url = reverse("lista_movimenti_finanziari")

        response = self.client.post(
            reverse("riconcilia_movimento", args=[movimento.pk]),
            {"azione": "ignora", "next": return_url},
        )

        self.assertRedirects(response, return_url, fetch_redirect_response=False)

    def test_rate_candidate_include_movimento_parziale_da_movimento(self):
        rata = self._crea_rata_corrente_con_familiare(importo=Decimal("300.00"))
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 11),
            importo=Decimal("100.00"),
            descrizione="Acconto retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidati = trova_rate_candidate(movimento)

        self.assertIn(rata.pk, [candidato.rata.pk for candidato in candidati])
        candidato = next(candidato for candidato in candidati if candidato.rata.pk == rata.pk)
        self.assertTrue(any("parziale" in motivazione for motivazione in candidato.motivazioni))

    def test_proposte_riconciliazione_movimento_normalizzano_rata(self):
        rata = self._crea_rata_corrente_con_familiare(importo=Decimal("300.00"))
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("300.00"),
            descrizione="Bonifico retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        proposte = proposte_riconciliazione_da_movimento(movimento)

        proposta = next(
            proposta
            for proposta in proposte
            if proposta.kind == "rate" and proposta.tipo == "singola" and proposta.targets == [rata]
        )
        self.assertEqual(proposta.direction, "movimento_to_targets")
        self.assertEqual(proposta.source.rata, rata)
        self.assertEqual(proposta.movimenti, [movimento])
        self.assertEqual(proposta.importo_totale, Decimal("300.00"))
        self.assertEqual(proposta.allocazioni[0].movimento, movimento)
        self.assertEqual(proposta.allocazioni[0].target, rata)
        self.assertEqual(proposta.allocazioni[0].target_tipo, "rata")
        self.assertIn("movimento_to_targets|rate|singola", proposta.key)

    def test_applica_proposta_riconciliazione_rate_accetta_piu_movimenti(self):
        rata = self._crea_rata_corrente_con_familiare(importo=Decimal("300.00"))
        primo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("100.00"),
            descrizione="Acconto retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        secondo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 12),
            importo=Decimal("200.00"),
            descrizione="Saldo retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        proposta = crea_proposta_riconciliazione(
            kind="rate",
            direction="target_to_movimenti",
            tipo="cumulativa",
            allocazioni=[
                (primo_movimento, rata, "rata", Decimal("100.00")),
                (secondo_movimento, rata, "rata", Decimal("200.00")),
            ],
        )

        risultato = applica_proposta_riconciliazione(proposta, utente=self.user)

        rata.refresh_from_db()
        primo_movimento.refresh_from_db()
        secondo_movimento.refresh_from_db()
        self.assertEqual(risultato.importo_totale, Decimal("300.00"))
        self.assertEqual({movimento.pk for movimento in risultato.movimenti}, {primo_movimento.pk, secondo_movimento.pk})
        self.assertEqual(risultato.targets, [rata])
        self.assertTrue(rata.pagata)
        self.assertEqual(rata.importo_pagato, Decimal("300.00"))
        self.assertEqual(primo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(secondo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_riconcilia_rata_accetta_piu_movimenti_candidati(self):
        rata = self._crea_rata_corrente_con_familiare(importo=Decimal("300.00"))
        primo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("100.00"),
            descrizione="Acconto retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        secondo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 12),
            importo=Decimal("200.00"),
            descrizione="Saldo retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidati = trova_movimenti_cumulativi_candidati_per_rate(rata, [rata])

        self.assertEqual(len(candidati), 1)
        self.assertEqual(
            {movimento.pk for movimento in candidati[0].movimenti},
            {primo_movimento.pk, secondo_movimento.pk},
        )
        proposte = proposte_riconciliazione_da_rata(rata, [rata])
        proposta_cumulativa = next(proposta for proposta in proposte if proposta.tipo == "cumulativa")
        self.assertEqual(proposta_cumulativa.kind, "rate")
        self.assertEqual(proposta_cumulativa.direction, "target_to_movimenti")
        self.assertEqual(proposta_cumulativa.targets, [rata])
        self.assertEqual(
            {movimento.pk for movimento in proposta_cumulativa.movimenti},
            {primo_movimento.pk, secondo_movimento.pk},
        )
        self.assertEqual(proposta_cumulativa.importo_totale, Decimal("300.00"))
        get_response = self.client.get(reverse("riconcilia_rata_iscrizione", kwargs={"pk": rata.pk}))
        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, "movimenti_cumulativi_0_ids")

        response = self.client.post(
            reverse("riconcilia_rata_iscrizione", kwargs={"pk": rata.pk}),
            {
                "azione": "collega_movimenti_cumulativa:0",
                "movimenti_cumulativi_0_ids": [str(primo_movimento.pk), str(secondo_movimento.pk)],
                f"importo_movimento_0_{primo_movimento.pk}": "100.00",
                f"importo_movimento_0_{secondo_movimento.pk}": "200.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        rata.refresh_from_db()
        primo_movimento.refresh_from_db()
        secondo_movimento.refresh_from_db()
        self.assertTrue(rata.pagata)
        self.assertEqual(rata.importo_pagato, Decimal("300.00"))
        self.assertEqual(primo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(secondo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_riconcilia_rata_accetta_movimento_inferiore_al_residuo_come_parziale(self):
        rata = self._crea_rata_corrente_con_familiare(importo=Decimal("480.00"))
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("445.00"),
            descrizione="Quota retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        response = self.client.post(
            reverse("riconcilia_rata_iscrizione", kwargs={"pk": rata.pk}),
            {
                "movimento_pk": str(movimento.pk),
                f"importo_rata_{rata.pk}": "480,00",
            },
        )

        self.assertEqual(response.status_code, 302)
        rata.refresh_from_db()
        movimento.refresh_from_db()
        link = RiconciliazioneRataMovimento.objects.get(rata=rata, movimento=movimento)
        self.assertEqual(link.importo, Decimal("445.00"))
        self.assertEqual(rata.importo_pagato, Decimal("445.00"))
        self.assertFalse(rata.pagata)
        self.assertEqual(importo_rata_residuo(rata), Decimal("35.00"))
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

        secondo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 12),
            importo=Decimal("445.00"),
            descrizione="Seconda quota retta settembre Mario Rossi",
            controparte="Mario Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        response = self.client.post(
            reverse("riconcilia_rata_iscrizione", kwargs={"pk": rata.pk}),
            {
                "movimento_pk": str(secondo_movimento.pk),
                f"importo_rata_{rata.pk}": "35,00",
            },
        )

        self.assertEqual(response.status_code, 302)
        rata.refresh_from_db()
        secondo_movimento.refresh_from_db()
        secondo_link = RiconciliazioneRataMovimento.objects.get(rata=rata, movimento=secondo_movimento)
        self.assertEqual(secondo_link.importo, Decimal("35.00"))
        self.assertEqual(rata.importo_pagato, Decimal("480.00"))
        self.assertTrue(rata.pagata)
        self.assertEqual(secondo_movimento.stato_riconciliazione, StatoRiconciliazione.NON_RICONCILIATO)

    def test_riconciliazione_movimento_in_uscita_mostra_scadenze_fornitore(self):
        fornitore = Fornitore.objects.create(denominazione="Energia Srl", tipo_soggetto="azienda")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.FATTURA,
            numero_documento="E-010",
            data_documento=date(2026, 4, 1),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 4, 30),
            importo_previsto=Decimal("122.00"),
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 29),
            importo=Decimal("-122.00"),
            descrizione="Bonifico Energia Srl fattura E-010",
            controparte="Energia Srl",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        response = self.client.get(reverse("riconcilia_movimento", args=[movimento.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Scadenze fornitore candidate")
        self.assertContains(response, "Energia Srl")
        self.assertNotContains(response, "Rate candidate")


@skip("Legacy test basato sulla tabella anagrafica.Famiglia rimossa.")
class RiconciliazioneRateMatchingTests(TestCase):
    def setUp(self):
        self.stato_relazione = StatoRelazioneFamiglia.objects.create(stato="Iscritta")
        self.relazione_genitore = RelazioneFamiliare.objects.create(relazione="Genitore")
        self.anno = AnnoScolastico.objects.create(
            nome_anno_scolastico="2025/2026",
            data_inizio=date(2025, 9, 1),
            data_fine=date(2026, 6, 30),
        )
        self.classe = Classe.objects.create(nome_classe="Primaria", ordine_classe=1)
        self.stato_iscrizione = StatoIscrizione.objects.create(stato_iscrizione="Iscritto")
        self.condizione = CondizioneIscrizione.objects.create(
            anno_scolastico=self.anno,
            nome_condizione_iscrizione="Retta standard",
            numero_mensilita_default=10,
        )
        TariffaCondizioneIscrizione.objects.create(
            condizione_iscrizione=self.condizione,
            ordine_figlio_da=1,
            retta_annuale=Decimal("1000.00"),
        )

    def _crea_rata(self, *, famiglia_cognome, studente_nome, studente_cognome, genitore_nome, genitore_cognome):
        famiglia = Famiglia.objects.create(
            cognome_famiglia=famiglia_cognome,
            stato_relazione_famiglia=self.stato_relazione,
        )
        Familiare.objects.create(
            famiglia=famiglia,
            relazione_familiare=self.relazione_genitore,
            nome=genitore_nome,
            cognome=genitore_cognome,
        )
        studente = Studente.objects.create(
            famiglia=famiglia,
            nome=studente_nome,
            cognome=studente_cognome,
            data_nascita=date(2020, 5, 5),
        )
        iscrizione = Iscrizione.objects.create(
            studente=studente,
            anno_scolastico=self.anno,
            classe=self.classe,
            stato_iscrizione=self.stato_iscrizione,
            condizione_iscrizione=self.condizione,
            data_iscrizione=date(2025, 9, 1),
            data_fine_iscrizione=date(2026, 6, 30),
        )
        rata = RataIscrizione.objects.create(
            iscrizione=iscrizione,
            famiglia=famiglia,
            numero_rata=1,
            mese_riferimento=9,
            anno_riferimento=2025,
            importo_dovuto=Decimal("100.00"),
            importo_finale=Decimal("100.00"),
            data_scadenza=date(2025, 9, 10),
        )
        return famiglia, studente, rata

    def test_rate_candidate_usa_nominativi_genitori_in_causale(self):
        _, _, rata_corretta = self._crea_rata(
            famiglia_cognome="Rossi",
            studente_nome="Luca",
            studente_cognome="Rossi",
            genitore_nome="Simone",
            genitore_cognome="Rossi",
        )
        _, _, rata_altra_famiglia = self._crea_rata(
            famiglia_cognome="Rossi",
            studente_nome="Anna",
            studente_cognome="Rossi",
            genitore_nome="Paolo",
            genitore_cognome="Rossi",
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("100.00"),
            descrizione="Bonifico retta settembre Simone Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidate_ids = [candidato.rata.pk for candidato in trova_rate_candidate(movimento)]

        self.assertIn(rata_corretta.pk, candidate_ids)
        self.assertNotIn(rata_altra_famiglia.pk, candidate_ids)

    def test_movimenti_candidati_escludono_causali_di_altri_genitori(self):
        _, _, rata = self._crea_rata(
            famiglia_cognome="Rossi",
            studente_nome="Luca",
            studente_cognome="Rossi",
            genitore_nome="Simone",
            genitore_cognome="Rossi",
        )
        movimento_corretto = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("100.00"),
            descrizione="Bonifico retta settembre Simone Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        movimento_altra_famiglia = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("100.00"),
            descrizione="Bonifico retta settembre Paolo Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidate_ids = [
            candidato.movimento.pk
            for candidato in trova_movimenti_candidati_per_rate(rata, [rata])
        ]

        self.assertIn(movimento_corretto.pk, candidate_ids)
        self.assertNotIn(movimento_altra_famiglia.pk, candidate_ids)

    def test_rate_cumulative_candidate_limits_search_on_many_open_rates(self):
        famiglia = Famiglia.objects.create(
            cognome_famiglia="Rossi",
            stato_relazione_famiglia=self.stato_relazione,
        )
        Familiare.objects.create(
            famiglia=famiglia,
            relazione_familiare=self.relazione_genitore,
            nome="Simone",
            cognome="Rossi",
        )
        nomi = ["Luca", "Marta", "Anna", "Pietro", "Giulia", "Marco"]
        for nome in nomi:
            studente = Studente.objects.create(
                famiglia=famiglia,
                nome=nome,
                cognome="Rossi",
                data_nascita=date(2020, 5, 5),
            )
            iscrizione = Iscrizione.objects.create(
                studente=studente,
                anno_scolastico=self.anno,
                classe=self.classe,
                stato_iscrizione=self.stato_iscrizione,
                condizione_iscrizione=self.condizione,
                data_iscrizione=date(2025, 9, 1),
                data_fine_iscrizione=date(2026, 6, 30),
            )
            for index, mese in enumerate(range(9, 13), start=1):
                RataIscrizione.objects.create(
                    iscrizione=iscrizione,
                    famiglia=famiglia,
                    numero_rata=index,
                    mese_riferimento=mese,
                    anno_riferimento=2025,
                    importo_dovuto=Decimal("100.00"),
                    importo_finale=Decimal("100.00"),
                    data_scadenza=date(2025, mese, 10),
                )

        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("200.00"),
            descrizione="Bonifico rette settembre Luca e Marta Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidati = trova_rate_cumulative_candidate(movimento)

        self.assertTrue(candidati)
        self.assertEqual(len(candidati[0].allocazioni), 2)
        self.assertEqual(
            sum(importo for _rata, importo in candidati[0].allocazioni),
            Decimal("200.00"),
        )

    def test_salvataggio_riconciliazione_blocca_movimento_senza_nominativo_compatibile(self):
        _, _, rata = self._crea_rata(
            famiglia_cognome="Rossi",
            studente_nome="Luca",
            studente_cognome="Rossi",
            genitore_nome="Simone",
            genitore_cognome="Rossi",
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 10),
            importo=Decimal("100.00"),
            descrizione="Bonifico retta settembre Paolo Rossi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        with self.assertRaisesMessage(ValidationError, "Controllo di sicurezza"):
            riconcilia_movimento_con_rate(movimento, [(rata, Decimal("100.00"))])

        movimento.refresh_from_db()
        rata.refresh_from_db()
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.NON_RICONCILIATO)
        self.assertFalse(rata.pagata)


class BudgetingGestioneFinanziariaTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="budget@example.com",
            email="budget@example.com",
            password="Password123!",
        )
        SistemaUtentePermessi.objects.create(
            user=self.user,
            permesso_gestione_finanziaria=LivelloPermesso.GESTIONE,
        )
        self.client.force_login(self.user)
        self.today = timezone.localdate()
        AnnoScolastico.objects.create(
            nome_anno_scolastico="2025/2026",
            data_inizio=date(2025, 9, 1),
            data_fine=date(2026, 6, 30),
        )

    def test_budgeting_dashboard_renders_recurring_forecast(self):
        categoria = CategoriaFinanziaria.objects.create(
            nome="Affitto",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        VoceBudgetRicorrente.objects.create(
            nome="Affitto sede",
            tipo=TipoVoceBudget.USCITA,
            categoria=categoria,
            importo=Decimal("1500.00"),
            frequenza=FrequenzaVoceBudget.MENSILE,
            data_inizio=date(self.today.year, self.today.month, 1),
            giorno_previsto=5,
        )

        data = build_budgeting_dashboard_data(today=self.today)
        self.assertEqual(data["current_month"]["ricorrenti_uscite"], Decimal("1500.00"))
        self.assertEqual(data["current_month"]["uscite_previste"], Decimal("1500.00"))

        response = self.client.get(reverse("budgeting_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Budgeting")
        self.assertContains(response, "Affitto sede")
        self.assertContains(response, "EUR 1.500,00")
        self.assertContains(response, "Flusso di cassa")
        self.assertContains(response, "Bilancio mese per mese")
        self.assertContains(response, "Affitto")

    def test_home_dashboard_esclude_ricariche_prepagate_e_trasferimenti(self):
        conto_corrente = ContoBancario.objects.create(
            nome_conto="Conto operativo",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
        )
        prepagata = ContoBancario.objects.create(
            nome_conto="Carta web",
            tipo_conto=TipoContoFinanziario.CARTA_PREPAGATA,
        )
        trasferimenti = CategoriaFinanziaria.objects.create(
            nome="Giroconti",
            tipo=TipoCategoriaFinanziaria.TRASFERIMENTO,
        )
        MovimentoFinanziario.objects.create(
            conto=conto_corrente,
            data_contabile=self.today,
            importo=Decimal("100.00"),
            descrizione="Incasso reale",
        )
        MovimentoFinanziario.objects.create(
            conto=prepagata,
            data_contabile=self.today,
            importo=Decimal("300.00"),
            descrizione="Ricarica carta",
            canale=CanaleMovimento.PREPAGATA,
        )
        MovimentoFinanziario.objects.create(
            conto=prepagata,
            data_contabile=self.today,
            importo=Decimal("-45.00"),
            descrizione="Spesa web",
            canale=CanaleMovimento.PREPAGATA,
        )
        MovimentoFinanziario.objects.create(
            conto=conto_corrente,
            data_contabile=self.today,
            importo=Decimal("-300.00"),
            descrizione="Giroconto verso carta",
            categoria=trasferimenti,
        )

        data = build_home_financial_dashboard_data(today=self.today)

        self.assertEqual(data["monthly"]["totale_entrate"], Decimal("100.00"))
        self.assertEqual(data["monthly"]["totale_uscite"], Decimal("45.00"))
        self.assertEqual(data["monthly"]["saldo"], Decimal("55.00"))
        self.assertEqual(data["monthly"]["movimenti"], 2)

    def test_crea_voce_budget(self):
        response = self.client.post(
            reverse("crea_voce_budget"),
            {
                "nome": "Contributo pubblico previsto",
                "tipo": TipoVoceBudget.ENTRATA,
                "categoria": "",
                "fornitore": "",
                "importo": "500.00",
                "frequenza": FrequenzaVoceBudget.UNA_TANTUM,
                "data_inizio": self.today.strftime("%Y-%m-%d"),
                "data_fine": "",
                "giorno_previsto": str(self.today.day),
                "attiva": "on",
                "note": "Prima ipotesi di budget.",
            },
        )

        self.assertRedirects(response, reverse("budgeting_dashboard"))
        voce = VoceBudgetRicorrente.objects.get(nome="Contributo pubblico previsto")
        self.assertEqual(voce.tipo, TipoVoceBudget.ENTRATA)
        self.assertEqual(voce.importo, Decimal("500.00"))
        self.assertIsNone(voce.mese_previsto)

    def test_crea_voce_budget_popup_chiude_dopo_salvataggio(self):
        response = self.client.post(
            reverse("crea_voce_budget"),
            {
                "popup": "1",
                "nome": "Contributo popup",
                "tipo": TipoVoceBudget.ENTRATA,
                "categoria": "",
                "fornitore": "",
                "importo": "300.00",
                "frequenza": FrequenzaVoceBudget.UNA_TANTUM,
                "data_inizio": self.today.strftime("%Y-%m-%d"),
                "data_fine": "",
                "giorno_previsto": str(self.today.day),
                "attiva": "on",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "popup/popup_close.html")
        self.assertContains(response, "Voce di budget creata correttamente.")
        self.assertTrue(VoceBudgetRicorrente.objects.filter(nome="Contributo popup").exists())

    def test_dashboard_apre_voci_budget_in_popup(self):
        voce = VoceBudgetRicorrente.objects.create(
            nome="Canone da popup",
            tipo=TipoVoceBudget.USCITA,
            importo=Decimal("900.00"),
            frequenza=FrequenzaVoceBudget.MENSILE,
            data_inizio=date(2026, 1, 1),
            giorno_previsto=1,
        )

        response = self.client.get(reverse("budgeting_dashboard"))

        self.assertContains(response, f'{reverse("crea_voce_budget")}?popup=1')
        self.assertContains(response, f'{reverse("modifica_voce_budget", args=[voce.pk])}?popup=1')
        self.assertContains(response, 'data-window-popup="1"')

    def test_voce_budget_inattiva_resta_visibile_e_togglabile(self):
        categoria = CategoriaFinanziaria.objects.create(
            nome="Utenze",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        voce = VoceBudgetRicorrente.objects.create(
            nome="Utenza stimata",
            tipo=TipoVoceBudget.USCITA,
            categoria=categoria,
            importo=Decimal("280.00"),
            frequenza=FrequenzaVoceBudget.MENSILE,
            data_inizio=date(self.today.year, self.today.month, 1),
            giorno_previsto=15,
            attiva=False,
        )

        data = build_budgeting_dashboard_data(today=self.today)
        self.assertEqual(data["current_month"]["ricorrenti_uscite"], Decimal("0.00"))
        self.assertEqual(data["voci_budget_count"], 1)
        self.assertEqual(data["voci_budget_attive_count"], 0)

        response = self.client.get(reverse("budgeting_dashboard"))
        self.assertContains(response, "Utenza stimata")
        self.assertContains(response, "Non attiva")

        response = self.client.post(
            reverse("toggle_voce_budget", args=[voce.pk]),
            {"attiva": "1", "ajax": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        voce.refresh_from_db()
        self.assertTrue(voce.attiva)

        data = build_budgeting_dashboard_data(today=self.today)
        self.assertEqual(data["current_month"]["ricorrenti_uscite"], Decimal("280.00"))

    def test_modifica_voce_budget_precompila_date_e_nasconde_mese_previsto(self):
        voce = VoceBudgetRicorrente.objects.create(
            nome="Canone annuale",
            tipo=TipoVoceBudget.USCITA,
            importo=Decimal("900.00"),
            frequenza=FrequenzaVoceBudget.ANNUALE,
            data_inizio=date(2026, 1, 15),
            data_fine=date(2026, 12, 31),
            giorno_previsto=15,
            mese_previsto=4,
        )

        response = self.client.get(reverse("modifica_voce_budget", args=[voce.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="2026-01-15"')
        self.assertContains(response, 'value="2026-12-31"')
        self.assertNotContains(response, "Mese previsto")
        self.assertContains(response, 'id="add-budget-categoria-btn"')
        self.assertContains(response, 'id="add-budget-fornitore-btn"')

    def test_modifica_voce_budget_popup_usa_template_popup(self):
        voce = VoceBudgetRicorrente.objects.create(
            nome="Canone popup",
            tipo=TipoVoceBudget.USCITA,
            importo=Decimal("900.00"),
            frequenza=FrequenzaVoceBudget.ANNUALE,
            data_inizio=date(2026, 1, 15),
            data_fine=date(2026, 12, 31),
            giorno_previsto=15,
        )

        response = self.client.get(f'{reverse("modifica_voce_budget", args=[voce.pk])}?popup=1')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'body class="popup-page"')
        self.assertContains(response, 'name="popup" value="1"')
        self.assertContains(response, "budget-voice-popup-card")
        self.assertContains(response, "budget-voice-input-shell")
        self.assertContains(response, 'value="2026-01-15"')

    def test_modifica_voce_budget_popup_chiude_dopo_salvataggio(self):
        voce = VoceBudgetRicorrente.objects.create(
            nome="Canone da aggiornare",
            tipo=TipoVoceBudget.USCITA,
            importo=Decimal("900.00"),
            frequenza=FrequenzaVoceBudget.ANNUALE,
            data_inizio=date(2026, 1, 15),
            data_fine=date(2026, 12, 31),
            giorno_previsto=15,
        )

        response = self.client.post(
            reverse("modifica_voce_budget", args=[voce.pk]),
            {
                "popup": "1",
                "nome": "Canone aggiornato",
                "tipo": TipoVoceBudget.USCITA,
                "categoria": "",
                "fornitore": "",
                "importo": "950.00",
                "frequenza": FrequenzaVoceBudget.ANNUALE,
                "data_inizio": "2026-01-15",
                "data_fine": "2026-12-31",
                "giorno_previsto": "15",
                "attiva": "on",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "popup/popup_close.html")
        self.assertContains(response, "Voce di budget aggiornata correttamente.")
        voce.refresh_from_db()
        self.assertEqual(voce.nome, "Canone aggiornato")
        self.assertEqual(voce.importo, Decimal("950.00"))


class FornitoriGestioneFinanziariaTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.media_override = override_settings(MEDIA_ROOT=self.media_root)
        self.media_override.enable()
        self.user = User.objects.create_user(
            username="finanza@example.com",
            email="finanza@example.com",
            password="Password123!",
        )
        SistemaUtentePermessi.objects.create(
            user=self.user,
            permesso_gestione_finanziaria=LivelloPermesso.GESTIONE,
        )
        self.client.force_login(self.user)

    def tearDown(self):
        self.media_override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)
        super().tearDown()

    def _crea_scadenza_pagamento_test(self, *, importo=Decimal("100.00")):
        fornitore = Fornitore.objects.create(denominazione="Beta Servizi")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="BETA-001",
            data_documento=timezone.localdate(),
            totale=importo,
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=timezone.localdate(),
            importo_previsto=importo,
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=timezone.localdate(),
            importo=-importo,
            descrizione="Bonifico Beta Servizi",
            controparte="Beta Servizi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        return scadenza, movimento

    def test_tipo_documento_fornitore_include_proforma_after_fattura(self):
        choices = list(TipoDocumentoFornitore.choices)

        self.assertEqual(choices[0], (TipoDocumentoFornitore.FATTURA, "Fattura"))
        self.assertEqual(choices[1], (TipoDocumentoFornitore.PROFORMA, "Proforma"))

    def test_fatture_scadenze_mostra_pulsante_pulizia_duplicati(self):
        response = self.client.get(reverse("fatture_scadenze_fornitori"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pulisci duplicati")
        self.assertContains(response, reverse("pulizia_duplicati_documenti_fornitori"))

    def test_pulizia_duplicati_documenti_fornitori_mostra_gruppi_preselezionati(self):
        fornitore = Fornitore.objects.create(denominazione="CAMST")
        documento_keep = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="CAMST-001",
            data_documento=date(2026, 1, 31),
            totale=Decimal("122.00"),
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento_keep,
            data_scadenza=date(2026, 2, 28),
            importo_previsto=Decimal("122.00"),
        )
        documento_duplicato = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="CAMST 001",
            data_documento=date(2026, 1, 31),
            totale=Decimal("122.00"),
        )

        response = self.client.get(reverse("pulizia_duplicati_documenti_fornitori"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Gruppo duplicati #1")
        self.assertContains(response, "stesso fornitore, tipo, numero normalizzato e data documento")
        self.assertContains(response, "Conservata")
        self.assertContains(response, "Duplicata")
        content = response.content.decode()
        self.assertIn(f'value="{documento_keep.pk}" data-bulk-checkbox data-duplicate-delete disabled', content)
        self.assertIn(f'value="{documento_duplicato.pk}" data-bulk-checkbox data-duplicate-delete checked', content)

    def test_pulizia_duplicati_documenti_fornitori_usa_numero_fattura_come_segnale_forte(self):
        fornitore = Fornitore.objects.create(denominazione="CAMST")
        documento_keep = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="42/A",
            data_documento=date(2026, 1, 31),
            data_ricezione=date(2026, 2, 1),
            descrizione="Servizio mensa gennaio",
            totale=Decimal("122.00"),
        )
        documento_duplicato = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="42 A",
            data_documento=date(2026, 2, 3),
            data_ricezione=date(2026, 2, 4),
            descrizione="Mensa CAMST",
            totale=Decimal("130.00"),
        )

        response = self.client.get(reverse("pulizia_duplicati_documenti_fornitori"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "stesso fornitore e numero fattura normalizzato")
        content = response.content.decode()
        self.assertIn(f'value="{documento_keep.pk}" data-bulk-checkbox data-duplicate-delete disabled', content)
        self.assertIn(f'value="{documento_duplicato.pk}" data-bulk-checkbox data-duplicate-delete checked', content)

    def test_pulizia_duplicati_documenti_fornitori_confronta_date_scadenza_importo_e_descrizione(self):
        fornitore = Fornitore.objects.create(denominazione="CAMST")
        documento_keep = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FT-100",
            data_documento=date(2026, 1, 31),
            data_ricezione=date(2026, 2, 1),
            descrizione="Servizio mensa gennaio",
            totale=Decimal("122.00"),
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento_keep,
            data_scadenza=date(2026, 2, 28),
            importo_previsto=Decimal("122.00"),
        )
        documento_duplicato = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FT-101",
            data_documento=date(2026, 1, 31),
            data_ricezione=date(2026, 2, 1),
            descrizione="Servizio mensa gennaio",
            totale=Decimal("122.00"),
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento_duplicato,
            data_scadenza=date(2026, 2, 28),
            importo_previsto=Decimal("122.00"),
        )

        response = self.client.get(reverse("pulizia_duplicati_documenti_fornitori"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "stessi dati fattura, ricezione, scadenza, importo e descrizione")
        content = response.content.decode()
        self.assertIn(f'value="{documento_keep.pk}" data-bulk-checkbox data-duplicate-delete disabled', content)
        self.assertIn(f'value="{documento_duplicato.pk}" data-bulk-checkbox data-duplicate-delete checked', content)

    def test_pulizia_duplicati_documenti_fornitori_riconosce_scadenze_duplicate_stessa_fattura(self):
        fornitore = Fornitore.objects.create(denominazione="Duferco Energia Spa")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="00126FT02140128",
            data_documento=date(2026, 6, 24),
            data_ricezione=date(2026, 6, 24),
            descrizione="Fattura energia",
            totale=Decimal("213.47"),
        )
        scadenza_keep = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 7, 9),
            importo_previsto=Decimal("213.47"),
        )
        scadenza_duplicata = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 7, 9),
            importo_previsto=Decimal("213.47"),
        )

        response = self.client.get(reverse("pulizia_duplicati_documenti_fornitori"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Gruppo scadenze duplicate #1")
        self.assertContains(response, "stessa fattura, scadenza e importo")
        content = response.content.decode()
        self.assertIn(f'value="{scadenza_keep.pk}" data-bulk-checkbox data-duplicate-delete disabled', content)
        self.assertIn(f'value="{scadenza_duplicata.pk}" data-bulk-checkbox data-duplicate-delete checked', content)

    def test_pulizia_duplicati_documenti_fornitori_elimina_scadenza_duplicata_senza_cancellare_fattura(self):
        fornitore = Fornitore.objects.create(denominazione="Duferco Energia Spa")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="00126FT02140128",
            data_documento=date(2026, 6, 24),
            data_ricezione=date(2026, 6, 24),
            totale=Decimal("213.47"),
        )
        scadenza_keep = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 7, 9),
            importo_previsto=Decimal("213.47"),
        )
        scadenza_duplicata = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 7, 9),
            importo_previsto=Decimal("213.47"),
        )

        response = self.client.post(
            reverse("pulizia_duplicati_documenti_fornitori"),
            {
                "tipo": "scadenze",
                "selected_ids": [str(scadenza_duplicata.pk)],
            },
        )

        self.assertRedirects(response, reverse("fatture_scadenze_fornitori"))
        self.assertTrue(DocumentoFornitore.objects.filter(pk=documento.pk).exists())
        self.assertTrue(ScadenzaPagamentoFornitore.objects.filter(pk=scadenza_keep.pk).exists())
        self.assertFalse(ScadenzaPagamentoFornitore.objects.filter(pk=scadenza_duplicata.pk).exists())

    def test_pulizia_duplicati_documenti_fornitori_elimina_solo_selezionati(self):
        fornitore = Fornitore.objects.create(denominazione="CAMST")
        documento_keep = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="CAMST-001",
            data_documento=date(2026, 1, 31),
            totale=Decimal("122.00"),
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento_keep,
            data_scadenza=date(2026, 2, 28),
            importo_previsto=Decimal("122.00"),
        )
        documento_duplicato = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="CAMST 001",
            data_documento=date(2026, 1, 31),
            totale=Decimal("122.00"),
        )

        response = self.client.post(
            reverse("pulizia_duplicati_documenti_fornitori"),
            {"selected_ids": [str(documento_duplicato.pk)]},
        )

        self.assertRedirects(response, reverse("fatture_scadenze_fornitori"))
        self.assertTrue(DocumentoFornitore.objects.filter(pk=documento_keep.pk).exists())
        self.assertFalse(DocumentoFornitore.objects.filter(pk=documento_duplicato.pk).exists())
        self.assertTrue(ScadenzaPagamentoFornitore.objects.filter(pk=scadenza.pk, documento=documento_keep).exists())

    def test_pulizia_duplicati_documenti_fornitori_memorizza_alias_fic_eliminato(self):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        fornitore = Fornitore.objects.create(denominazione="Alias Supplier")
        documento_keep = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="ALIAS-1",
            data_documento=date(2026, 1, 31),
            totale=Decimal("122.00"),
            origine=OrigineDocumentoFornitore.FATTURE_IN_CLOUD,
            external_source="fatture_in_cloud",
            external_id="fic-keep",
        )
        documento_duplicato = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="ALIAS 1",
            data_documento=date(2026, 1, 31),
            totale=Decimal("122.00"),
            origine=OrigineDocumentoFornitore.FATTURE_IN_CLOUD,
            external_source="fatture_in_cloud",
            external_id="fic-dup",
        )

        response = self.client.post(
            reverse("pulizia_duplicati_documenti_fornitori"),
            {"selected_ids": [str(documento_duplicato.pk)]},
        )

        self.assertRedirects(response, reverse("fatture_scadenze_fornitori"))
        self.assertTrue(DocumentoFornitore.objects.filter(pk=documento_keep.pk).exists())
        self.assertFalse(DocumentoFornitore.objects.filter(pk=documento_duplicato.pk).exists())
        alias = DocumentoFornitoreImportAlias.objects.get(
            external_source="fatture_in_cloud",
            external_id="fic-dup",
        )
        self.assertEqual(alias.documento, documento_keep)
        self.assertFalse(alias.ignorato)

        result = importa_documento_fatture_in_cloud(
            connessione,
            {
                "id": "fic-dup",
                "type": "expense",
                "description": "Duplicato gia assorbito",
                "invoice_number": "ALIAS 1",
                "date": "2026-01-31",
                "amount_net": "100.00",
                "amount_vat": "22.00",
                "amount_gross": "122.00",
                "entity": {"name": "Alias Supplier"},
                "payments_list": [{"due_date": "2026-02-28", "amount": "122.00"}],
            },
            pending=False,
            utente=self.user,
        )

        self.assertTrue(result["skipped"])
        self.assertEqual(result["skip_reason"], "alias_assorbito")
        self.assertEqual(DocumentoFornitore.objects.count(), 1)
        documento_keep.refresh_from_db()
        self.assertEqual(documento_keep.external_id, "fic-keep")

    def test_pagamento_fornitore_popup_usa_layout_senza_shell_globale(self):
        scadenza, _movimento = self._crea_scadenza_pagamento_test()

        response = self.client.get(f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza.pk})}?popup=1")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="popup-page"', html=False)
        self.assertContains(response, "supplier-payment-shell is-popup")
        self.assertContains(response, "Riconciliazione bancaria")
        self.assertContains(response, "Riconcilia")
        self.assertContains(response, "Aggiungi pagamento")
        self.assertContains(response, "Salva pagamento")
        self.assertContains(response, "Lascia vuoto il movimento bancario")
        self.assertContains(response, 'onclick="window.close()"')
        self.assertContains(response, '<span class="btn-label">Annulla</span>', html=False)
        self.assertContains(response, "Confermi la riconciliazione con questo movimento bancario")
        self.assertNotContains(response, "NAVIGAZIONE")

    def test_pagamento_fornitore_popup_riconcilia_movimento_candidato(self):
        scadenza, movimento = self._crea_scadenza_pagamento_test()

        response = self.client.post(
            f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza.pk})}?popup=1",
            {
                "popup": "1",
                "scadenza": str(scadenza.pk),
                "quick_movimento": str(movimento.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pagamento fornitore registrato correttamente.")
        self.assertContains(response, r"gestione\u002Dfinanziaria/fatture\u002Dscadenze\u002Dfornitori")
        self.assertContains(response, "handleReloadToUrl")
        self.assertContains(response, "popup-close-fallback")
        pagamento = PagamentoFornitore.objects.get(scadenza=scadenza)
        self.assertEqual(pagamento.movimento_finanziario, movimento)
        self.assertEqual(pagamento.metodo, MetodoPagamentoFornitore.BANCA)
        self.assertEqual(pagamento.importo, Decimal("100.00"))
        scadenza.refresh_from_db()
        movimento.refresh_from_db()
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_registra_pagamento_fornitore_ignora_cache_residuo_stantia(self):
        scadenza, movimento = self._crea_scadenza_pagamento_test()
        movimento._arboris_importo_disponibile_cache = abs(movimento.importo)

        registra_pagamento_fornitore(
            scadenza,
            importo=Decimal("100.00"),
            data_pagamento=movimento.data_contabile,
            movimento=movimento,
            metodo=MetodoPagamentoFornitore.BANCA,
            conto=movimento.conto,
            utente=self.user,
        )

        movimento.refresh_from_db()
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_pagamento_fornitore_accetta_piu_righe_manuali(self):
        scadenza, _movimento = self._crea_scadenza_pagamento_test(importo=Decimal("100.00"))
        url = f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza.pk})}?popup=1"

        for index, (data_pagamento, importo, nota) in enumerate([
            ("2026-05-10", "35.00", "Acconto manuale"),
            ("2026-05-15", "65.00", "Saldo manuale"),
        ]):
            response = self.client.post(
                url,
                {
                    "popup": "1",
                    "scadenza": str(scadenza.pk),
                    "movimento_finanziario": "",
                    "data_pagamento": data_pagamento,
                    "importo": importo,
                    "metodo": MetodoPagamentoFornitore.MANUALE,
                    "conto_bancario": "",
                    "note": nota,
                },
            )
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Pagamento fornitore registrato correttamente.")
            if index == 0:
                response = self.client.get(url)
                self.assertContains(response, "Pagamenti già registrati")
                self.assertContains(response, "Registrazione manuale")
                self.assertContains(response, "Nuova riga pagamento")

        pagamenti = PagamentoFornitore.objects.filter(scadenza=scadenza).order_by("data_pagamento")
        self.assertEqual(pagamenti.count(), 2)
        self.assertEqual([pagamento.importo for pagamento in pagamenti], [Decimal("35.00"), Decimal("65.00")])
        self.assertTrue(all(pagamento.movimento_finanziario_id is None for pagamento in pagamenti))
        scadenza.refresh_from_db()
        scadenza.documento.refresh_from_db()
        self.assertEqual(scadenza.importo_pagato, Decimal("100.00"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(scadenza.documento.stato, StatoDocumentoFornitore.PAGATO)

    def test_pagamento_fornitore_accetta_piu_movimenti_bancari(self):
        scadenza, primo_movimento = self._crea_scadenza_pagamento_test(importo=Decimal("100.00"))
        secondo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 12),
            importo=Decimal("-60.00"),
            descrizione="Secondo bonifico Beta Servizi",
            controparte="Beta Servizi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        url = f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza.pk})}?popup=1"

        response = self.client.post(
            url,
            {
                "popup": "1",
                "scadenza": str(scadenza.pk),
                "movimento_finanziario": str(primo_movimento.pk),
                "data_pagamento": "2026-05-10",
                "importo": "40.00",
                "metodo": MetodoPagamentoFornitore.BANCA,
                "conto_bancario": "",
                "note": "Primo pagamento",
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            url,
            {
                "popup": "1",
                "scadenza": str(scadenza.pk),
                "movimento_finanziario": str(secondo_movimento.pk),
                "data_pagamento": "2026-05-12",
                "importo": "60.00",
                "metodo": MetodoPagamentoFornitore.BANCA,
                "conto_bancario": "",
                "note": "Saldo",
            },
        )
        self.assertEqual(response.status_code, 200)

        self.assertEqual(PagamentoFornitore.objects.filter(scadenza=scadenza).count(), 2)
        scadenza.refresh_from_db()
        primo_movimento.refresh_from_db()
        secondo_movimento.refresh_from_db()
        self.assertEqual(scadenza.importo_pagato, Decimal("100.00"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(primo_movimento.stato_riconciliazione, StatoRiconciliazione.NON_RICONCILIATO)
        self.assertEqual(secondo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_applica_proposta_riconciliazione_fornitore_accetta_piu_movimenti(self):
        scadenza, primo_movimento = self._crea_scadenza_pagamento_test(importo=Decimal("100.00"))
        primo_movimento.importo = Decimal("-40.00")
        primo_movimento.descrizione = "Acconto Beta Servizi"
        primo_movimento.save(update_fields=["importo", "descrizione"])
        secondo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=timezone.localdate(),
            importo=Decimal("-60.00"),
            descrizione="Saldo Beta Servizi",
            controparte="Beta Servizi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        proposta = crea_proposta_riconciliazione(
            kind="fornitore",
            direction="target_to_movimenti",
            tipo="cumulativa",
            allocazioni=[
                (primo_movimento, scadenza, "scadenza_fornitore", Decimal("40.00")),
                (secondo_movimento, scadenza, "scadenza_fornitore", Decimal("60.00")),
            ],
        )

        risultato = applica_proposta_riconciliazione(proposta, utente=self.user)

        scadenza.refresh_from_db()
        primo_movimento.refresh_from_db()
        secondo_movimento.refresh_from_db()
        self.assertEqual(risultato.importo_totale, Decimal("100.00"))
        self.assertEqual(len(risultato.pagamenti), 2)
        self.assertEqual(PagamentoFornitore.objects.filter(scadenza=scadenza).count(), 2)
        self.assertEqual(scadenza.importo_pagato, Decimal("100.00"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(primo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(secondo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_pagamento_fornitore_popup_riconcilia_piu_movimenti_candidati(self):
        scadenza, primo_movimento = self._crea_scadenza_pagamento_test(importo=Decimal("100.00"))
        primo_movimento.importo = Decimal("-40.00")
        primo_movimento.descrizione = "Acconto Beta Servizi"
        primo_movimento.save(update_fields=["importo", "descrizione"])
        secondo_movimento = MovimentoFinanziario.objects.create(
            data_contabile=timezone.localdate(),
            importo=Decimal("-60.00"),
            descrizione="Saldo Beta Servizi",
            controparte="Beta Servizi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidati = trova_movimenti_cumulativi_candidati_per_scadenza_fornitore(scadenza)

        self.assertEqual(len(candidati), 1)
        self.assertEqual(
            {movimento.pk for movimento in candidati[0].movimenti},
            {primo_movimento.pk, secondo_movimento.pk},
        )
        proposte = proposte_riconciliazione_da_scadenza_fornitore(scadenza)
        proposta_cumulativa = next(proposta for proposta in proposte if proposta.tipo == "cumulativa")
        self.assertEqual(proposta_cumulativa.kind, "fornitore")
        self.assertEqual(proposta_cumulativa.direction, "target_to_movimenti")
        self.assertEqual(proposta_cumulativa.targets, [scadenza])
        self.assertEqual(
            {movimento.pk for movimento in proposta_cumulativa.movimenti},
            {primo_movimento.pk, secondo_movimento.pk},
        )
        self.assertEqual(proposta_cumulativa.importo_totale, Decimal("100.00"))
        get_response = self.client.get(
            f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza.pk})}?popup=1"
        )
        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, "quick_movimenti_cumulativi")

        response = self.client.post(
            f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza.pk})}?popup=1",
            {
                "popup": "1",
                "quick_movimenti_cumulativi": "0",
                "quick_movimenti_cumulativi_0_ids": [str(primo_movimento.pk), str(secondo_movimento.pk)],
                f"quick_movimenti_cumulativi_0_importo_{primo_movimento.pk}": "40.00",
                f"quick_movimenti_cumulativi_0_importo_{secondo_movimento.pk}": "60.00",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pagamenti fornitore cumulativi registrati correttamente.")
        self.assertEqual(PagamentoFornitore.objects.filter(scadenza=scadenza).count(), 2)
        scadenza.refresh_from_db()
        primo_movimento.refresh_from_db()
        secondo_movimento.refresh_from_db()
        self.assertEqual(scadenza.importo_pagato, Decimal("100.00"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(primo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(secondo_movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_documento_fornitore_popup_annulla_pagamento_usa_popup_gestito(self):
        scadenza, movimento = self._crea_scadenza_pagamento_test()
        pagamento = riconcilia_movimento_con_scadenza_fornitore(movimento, scadenza, utente=self.user)
        documento_url = reverse("modifica_documento_fornitore", kwargs={"pk": scadenza.documento.pk})
        annulla_url = f"{reverse('elimina_pagamento_fornitore', kwargs={'pk': pagamento.pk})}?popup=1"

        response = self.client.get(f"{documento_url}?popup=1")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, annulla_url)
        self.assertContains(response, "reload_url=")
        self.assertContains(response, 'data-window-popup="1"')
        self.assertContains(response, f'data-popup-url="{annulla_url}&reload_url=')
        self.assertContains(response, 'data-popup-window-features="width=760,height=560,resizable=yes,scrollbars=yes"')

    def test_documento_fornitore_popup_mostra_elimina_movimento_in_sola_lettura(self):
        scadenza, movimento = self._crea_scadenza_pagamento_test()
        pagamento = riconcilia_movimento_con_scadenza_fornitore(movimento, scadenza, utente=self.user)
        documento_url = reverse("modifica_documento_fornitore", kwargs={"pk": scadenza.documento.pk})
        annulla_url = f"{reverse('elimina_pagamento_fornitore', kwargs={'pk': pagamento.pk})}?popup=1"

        response = self.client.get(f"{documento_url}?popup=1")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "is-view-mode")
        self.assertContains(response, "Elimina movimento")
        self.assertContains(response, f'href="{annulla_url}&reload_url=')
        content = response.content.decode()
        anchor_start = content.index(f'href="{annulla_url}&reload_url=')
        anchor_end = content.index(">", anchor_start)
        self.assertNotIn("mode-edit-only", content[anchor_start:anchor_end])
        actions_start = content.rfind('data-label="Azioni"', 0, anchor_start)
        payments_start = content.rfind('data-label="Pagamenti"', 0, anchor_start)
        self.assertGreater(actions_start, payments_start)
        actions_end = content.index("</td>", actions_start)
        self.assertIn("Elimina movimento", content[actions_start:actions_end])

    def test_elimina_pagamento_fornitore_popup_usa_layout_e_chiude(self):
        scadenza, movimento = self._crea_scadenza_pagamento_test()
        pagamento = riconcilia_movimento_con_scadenza_fornitore(movimento, scadenza, utente=self.user)
        annulla_url = f"{reverse('elimina_pagamento_fornitore', kwargs={'pk': pagamento.pk})}?popup=1"

        response = self.client.get(annulla_url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="popup-page"', html=False)
        self.assertContains(response, "supplier-payment-shell is-popup")
        self.assertContains(response, '<input type="hidden" name="popup" value="1">', html=False)
        self.assertContains(response, 'onclick="window.close()"')
        self.assertNotContains(response, "NAVIGAZIONE")

        response = self.client.post(annulla_url, {"popup": "1"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pagamento fornitore annullato.")
        self.assertContains(response, r"gestione\u002Dfinanziaria/fatture\u002Dscadenze\u002Dfornitori")
        self.assertContains(response, "handleReloadToUrl")
        self.assertContains(response, "popup-close-fallback")
        self.assertContains(response, "handleReload")
        self.assertFalse(PagamentoFornitore.objects.filter(pk=pagamento.pk).exists())
        movimento.refresh_from_db()
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.NON_RICONCILIATO)

    def test_categoria_spesa_crud_pages(self):
        response = self.client.get(reverse("crea_categoria_spesa"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "category-expense-editor-shell")
        self.assertContains(response, "Categoria padre")

        response = self.client.post(
            reverse("crea_categoria_spesa"),
            {
                "nome": "Consulenze",
                "descrizione": "Consulenze e servizi professionali",
                "ordine": "1",
                "attiva": "on",
            },
        )

        self.assertRedirects(response, reverse("lista_categorie_spesa"))
        categoria = CategoriaFinanziaria.objects.get(nome="Consulenze", tipo=TipoCategoriaFinanziaria.SPESA)
        self.assertTrue(categoria.attiva)

        response = self.client.post(
            reverse("crea_categoria_spesa"),
            {
                "nome": "Consulenze legali",
                "parent": str(categoria.pk),
                "descrizione": "Sottocategoria per consulenze legali",
                "ordine": "2",
                "attiva": "on",
            },
        )

        self.assertRedirects(response, reverse("lista_categorie_spesa"))
        sottocategoria = CategoriaFinanziaria.objects.get(
            nome="Consulenze legali",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        self.assertEqual(sottocategoria.parent, categoria)

        response = self.client.get(reverse("lista_categorie_spesa"))
        self.assertContains(response, "Consulenze")
        self.assertContains(response, "Consulenze legali")
        self.assertContains(response, 'data-report-category-toggle="categoria-spesa-')
        self.assertContains(response, 'data-report-category-parent="categoria-spesa-')
        self.assertContains(response, "category-tree-badge-parent")
        self.assertContains(response, "Figlia")

    def test_categorie_finanziarie_list_renders_parent_child_tree(self):
        categoria_padre = CategoriaFinanziaria.objects.create(
            nome="Utenze",
            tipo=TipoCategoriaFinanziaria.SPESA,
            ordine=1,
        )
        CategoriaFinanziaria.objects.create(
            nome="Energia elettrica",
            tipo=TipoCategoriaFinanziaria.SPESA,
            parent=categoria_padre,
            ordine=1,
        )

        response = self.client.get(reverse("lista_categorie_finanziarie"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Utenze")
        self.assertContains(response, "Energia elettrica")
        self.assertContains(response, 'data-report-category-toggle="categoria-')
        self.assertContains(response, 'data-report-category-parent="categoria-')
        self.assertContains(response, "category-tree-badge-parent")
        self.assertContains(response, "Figlia")

    def test_categoria_finanziaria_form_has_color_picker_and_icon_library(self):
        response = self.client.get(reverse("crea_categoria_finanziaria"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'type="color"')
        self.assertContains(response, 'id="id_colore_picker"')
        self.assertContains(response, "data-category-icon-picker")
        self.assertContains(response, 'data-icon-value="banknote"')
        self.assertContains(response, "js/pages/categoria-finanziaria-form.js")

    def test_fornitore_uses_categoria_spesa(self):
        categoria = crea_categoria_spesa_test("Utenze")

        response = self.client.post(
            reverse("crea_fornitore"),
            {
                "denominazione": "Energia Srl",
                "tipo_soggetto": "azienda",
                "categoria_spesa": str(categoria.pk),
                "codice_fiscale": "",
                "partita_iva": "12345678901",
                "indirizzo": "Via Roma 1",
                "telefono": "051000000",
                "email": "amministrazione@energia.test",
                "pec": "",
                "codice_sdi": "ABC1234",
                "referente": "Mario Bianchi",
                "iban": "",
                "banca": "",
                "note": "",
                "attivo": "on",
            },
        )

        fornitore = Fornitore.objects.get(denominazione="Energia Srl")
        self.assertRedirects(response, reverse("modifica_fornitore", kwargs={"pk": fornitore.pk}))
        self.assertEqual(fornitore.categoria_spesa, categoria)

    def test_fornitore_form_renders_categoria_spesa_popup_controls(self):
        response = self.client.get(reverse("crea_fornitore"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Categoria di spesa")
        self.assertContains(response, 'id="add-categoria-spesa-btn"')
        self.assertContains(response, 'id="edit-categoria-spesa-btn"')
        self.assertContains(response, 'id="delete-categoria-spesa-btn"')
        self.assertContains(response, "supplier-profile-card")
        self.assertContains(response, "js/pages/fornitore-form.js")

    def test_fornitore_detail_renders_view_mode_and_sidebar_cards(self):
        categoria = crea_categoria_spesa_test("Cancelleria")
        fornitore = Fornitore.objects.create(
            denominazione="Carta Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
            partita_iva="12345678901",
            email="ordini@carta.test",
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            categoria_spesa=categoria,
            numero_documento="C-001",
            data_documento=date(2026, 5, 1),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            imponibile_ritenuta_acconto=Decimal("100.00"),
            ritenuta_acconto=Decimal("20.00"),
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("122.00"),
        )

        response = self.client.get(reverse("modifica_fornitore", kwargs={"pk": fornitore.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="fornitore-detail-form"')
        self.assertContains(response, "is-view-mode")
        self.assertContains(response, 'id="enable-edit-fornitore-btn"')
        self.assertContains(response, 'id="fornitore-main-fields"')
        self.assertContains(response, "supplier-profile-layout")
        self.assertContains(response, "supplier-profile-sidebar")
        self.assertContains(response, "Fatture recenti")
        self.assertContains(response, "Scadenze aperte")
        self.assertContains(response, "C-001")
        self.assertContains(response, "supplier-withholding-badge")
        self.assertContains(response, "31/05/2026")
        self.assertContains(response, 'class="btn btn-secondary btn-icon-text js-page-back-btn"')
        self.assertContains(response, "view-mode.js")

    def test_categoria_spesa_popup_create_returns_select_response(self):
        padre = crea_categoria_spesa_test("Servizi generali")
        response = self.client.get(f"{reverse('crea_categoria_spesa')}?popup=1&target_input_name=categoria_spesa")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "category-expense-editor-shell is-popup")
        self.assertContains(response, "Categoria padre")

        response = self.client.post(
            f"{reverse('crea_categoria_spesa')}?popup=1&target_input_name=categoria_spesa",
            {
                "popup": "1",
                "target_input_name": "categoria_spesa",
                "nome": "Servizi",
                "parent": str(padre.pk),
                "descrizione": "",
                "ordine": "",
                "attiva": "on",
            },
        )

        categoria = CategoriaFinanziaria.objects.get(nome="Servizi", tipo=TipoCategoriaFinanziaria.SPESA)
        self.assertEqual(categoria.parent, padre)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "dismissRelatedPopup")
        self.assertContains(response, "categoria_spesa")
        self.assertContains(response, str(categoria.pk))

    def test_conto_bancario_popup_usa_stile_globale(self):
        response = self.client.get(f"{reverse('crea_conto_bancario')}?popup=1&target_input_name=conto_bancario")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'body class="popup-page"', html=False)
        self.assertContains(response, "budget-voice-popup-card")
        self.assertContains(response, "budget-voice-input-shell")
        self.assertContains(response, "budget-voice-note-shell")
        self.assertContains(response, "btn-save-soft")
        self.assertContains(response, "Nuovo conto")
        self.assertContains(response, "Nome conto")
        self.assertContains(response, "IBAN")
        self.assertContains(response, '<input type="hidden" name="target_input_name" value="conto_bancario">', html=False)
        self.assertContains(response, 'placeholder="Aggiungi una nota..."', html=False)
        self.assertContains(response, 'data-rich-notes-skip="true"', html=False)
        self.assertNotContains(response, "form-table")
        self.assertNotContains(response, "detail-form")

    def test_documento_fornitore_importi_zero_default_restano_placeholder(self):
        form = DocumentoFornitoreForm()

        for field_name in ("imponibile", "iva", "totale"):
            with self.subTest(field_name=field_name):
                self.assertEqual(form.fields[field_name].initial, "")
                self.assertEqual(form[field_name].value(), "")
                self.assertEqual(form.fields[field_name].widget.attrs["placeholder"], "0,00")

    def test_documento_fornitore_creates_scadenza_and_calculates_totals(self):
        categoria = crea_categoria_spesa_test("Manutenzioni")
        fornitore = Fornitore.objects.create(
            denominazione="Tecnica Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )

        with patch("gestione_finanziaria.models.timezone.localdate", return_value=date(2026, 5, 1)):
            response = self.client.post(
                reverse("crea_documento_fornitore"),
                {
                    "fornitore": str(fornitore.pk),
                    "categoria_spesa": "",
                    "tipo_documento": "fattura",
                    "numero_documento": "F-001",
                    "data_documento": "2026-04-15",
                    "data_ricezione": "2026-04-16",
                    "anno_competenza": "",
                    "mese_competenza": "",
                    "descrizione": "Manutenzione ordinaria",
                    "imponibile": "1000.00",
                    "aliquota_iva": "22.00",
                    "iva": "",
                    "totale": "",
                    "stato": StatoDocumentoFornitore.DA_PAGARE,
                    "note": "",
                    "scadenze-TOTAL_FORMS": "1",
                    "scadenze-INITIAL_FORMS": "0",
                    "scadenze-MIN_NUM_FORMS": "0",
                    "scadenze-MAX_NUM_FORMS": "1000",
                    "scadenze-0-data_scadenza": "2026-05-31",
                    "scadenze-0-importo_previsto": "1220.00",
                    "scadenze-0-importo_pagato": "0.00",
                    "scadenze-0-data_pagamento": "",
                    "scadenze-0-stato": StatoScadenzaFornitore.PREVISTA,
                    "scadenze-0-conto_bancario": "",
                    "scadenze-0-movimento_finanziario": "",
                    "scadenze-0-note": "",
                },
            )

        documento = DocumentoFornitore.objects.get(numero_documento="F-001")
        self.assertRedirects(response, reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk}))
        self.assertEqual(documento.categoria_spesa, categoria)
        self.assertEqual(documento.iva, Decimal("220.00"))
        self.assertEqual(documento.totale, Decimal("1220.00"))
        scadenza = ScadenzaPagamentoFornitore.objects.get(documento=documento)
        self.assertEqual(scadenza.importo_previsto, Decimal("1220.00"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PREVISTA)

        response = self.client.get(reverse("scadenziario_fornitori"))
        self.assertContains(response, "Tecnica Srl")
        self.assertContains(response, "F-001")

    def test_documento_fornitore_con_ritenuta_calcola_netto_da_pagare(self):
        categoria = crea_categoria_spesa_test("Consulenze")
        fornitore = Fornitore.objects.create(
            denominazione="Studio Professionale Rossi",
            tipo_soggetto="professionista",
            categoria_spesa=categoria,
        )

        response = self.client.post(
            reverse("crea_documento_fornitore"),
            {
                "fornitore": str(fornitore.pk),
                "categoria_spesa": "",
                "tipo_documento": TipoDocumentoFornitore.FATTURA,
                "numero_documento": "PRO-001",
                "data_documento": "2025-12-15",
                "data_ricezione": "",
                "anno_competenza": "",
                "mese_competenza": "",
                "descrizione": "Parcella professionale",
                "imponibile": "1716.00",
                "aliquota_iva": "22.00",
                "iva": "",
                "totale": "",
                "imponibile_ritenuta_acconto": "1650.00",
                "aliquota_ritenuta_acconto": "20.00",
                "ritenuta_acconto": "",
                "stato": StatoDocumentoFornitore.DA_PAGARE,
                "note": "",
                "scadenze-TOTAL_FORMS": "1",
                "scadenze-INITIAL_FORMS": "0",
                "scadenze-MIN_NUM_FORMS": "0",
                "scadenze-MAX_NUM_FORMS": "1000",
                "scadenze-0-data_scadenza": "2025-12-15",
                "scadenze-0-importo_previsto": "1763.52",
                "scadenze-0-importo_pagato": "1763.52",
                "scadenze-0-data_pagamento": "2025-12-15",
                "scadenze-0-stato": StatoScadenzaFornitore.PREVISTA,
                "scadenze-0-conto_bancario": "",
                "scadenze-0-movimento_finanziario": "",
                "scadenze-0-note": "",
            },
        )

        documento = DocumentoFornitore.objects.get(numero_documento="PRO-001")
        self.assertRedirects(response, reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk}))
        self.assertEqual(documento.imponibile, Decimal("1716.00"))
        self.assertEqual(documento.iva, Decimal("377.52"))
        self.assertEqual(documento.totale, Decimal("2093.52"))
        self.assertEqual(documento.imponibile_ritenuta_acconto, Decimal("1650.00"))
        self.assertEqual(documento.ritenuta_acconto, Decimal("330.00"))
        self.assertEqual(documento.totale_da_pagare, Decimal("1763.52"))
        self.assertEqual(documento.residuo_da_pagare, Decimal("0.00"))
        self.assertEqual(documento.stato, StatoDocumentoFornitore.PAGATO)

        scadenza = ScadenzaPagamentoFornitore.objects.get(documento=documento)
        self.assertEqual(scadenza.importo_previsto, Decimal("1763.52"))
        self.assertEqual(scadenza.importo_pagato, Decimal("1763.52"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)

        response = self.client.get(reverse("scadenziario_fornitori"), {"stato": StatoScadenzaFornitore.PAGATA})
        self.assertContains(response, "PRO-001")
        self.assertContains(response, "supplier-withholding-badge")

    def test_documento_fornitore_popup_salva_scadenza_con_importo_senza_data(self):
        categoria = crea_categoria_spesa_test("Consulenze vecchie")
        fornitore = Fornitore.objects.create(
            denominazione="Studio Professionale Verdi",
            tipo_soggetto="professionista",
            categoria_spesa=categoria,
        )

        response = self.client.post(
            f"{reverse('crea_documento_fornitore')}?popup=1",
            {
                "popup": "1",
                "fornitore": str(fornitore.pk),
                "categoria_spesa": "",
                "tipo_documento": TipoDocumentoFornitore.FATTURA,
                "numero_documento": "PRO-NODATE",
                "data_documento": "2025-12-15",
                "data_ricezione": "",
                "anno_competenza": "",
                "mese_competenza": "",
                "descrizione": "Parcella da inserimento rapido",
                "imponibile": "1000.00",
                "aliquota_iva": "22.00",
                "iva": "",
                "totale": "",
                "imponibile_ritenuta_acconto": "1000.00",
                "aliquota_ritenuta_acconto": "20.00",
                "ritenuta_acconto": "",
                "stato": StatoDocumentoFornitore.DA_PAGARE,
                "note": "",
                "scadenze-TOTAL_FORMS": "1",
                "scadenze-INITIAL_FORMS": "0",
                "scadenze-MIN_NUM_FORMS": "0",
                "scadenze-MAX_NUM_FORMS": "1000",
                "scadenze-0-data_scadenza": "",
                "scadenze-0-importo_previsto": "1020.00",
                "scadenze-0-importo_pagato": "",
                "scadenze-0-data_pagamento": "",
                "scadenze-0-stato": StatoScadenzaFornitore.PREVISTA,
                "scadenze-0-conto_bancario": "",
                "scadenze-0-movimento_finanziario": "",
                "scadenze-0-note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fattura fornitore creata correttamente.")
        documento = DocumentoFornitore.objects.get(numero_documento="PRO-NODATE")
        scadenza = ScadenzaPagamentoFornitore.objects.get(documento=documento)
        self.assertEqual(scadenza.data_scadenza, date(2025, 12, 15))
        self.assertEqual(scadenza.importo_previsto, Decimal("1020.00"))
        self.assertEqual(scadenza.importo_pagato, Decimal("0.00"))

    def test_documento_fornitore_form_renders_search_popup_controls(self):
        response = self.client.get(reverse("crea_documento_fornitore"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-searchable-placeholder="Cerca un fornitore..."')
        self.assertContains(response, 'id="add-fornitore-btn"')
        self.assertContains(response, "Categoria di spesa")
        self.assertContains(response, 'id="add-documento-categoria-spesa-btn"')
        self.assertContains(response, 'id="edit-documento-categoria-spesa-btn"')
        self.assertContains(response, 'id="delete-documento-categoria-spesa-btn"')
        self.assertContains(response, 'data-related-type="conto_bancario"')
        self.assertContains(response, 'data-related-type="movimento_finanziario"')
        self.assertContains(response, "Gennaio")
        self.assertContains(response, 'name="imponibile_ritenuta_acconto"')
        self.assertContains(response, "Netto da pagare")
        self.assertContains(response, "js/pages/documento-fornitore-form.js")

    def test_documento_fornitore_non_popup_apre_pagamento_in_popup_e_torna_a_scadenze(self):
        fornitore = Fornitore.objects.create(denominazione="Pagamento Popup Srl")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="PAY-1",
            data_documento=date(2026, 5, 10),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("122.00"),
            importo_pagato=Decimal("40.00"),
        )
        documento_url = reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})
        pagamento_url = f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza.pk})}?popup=1"

        response = self.client.get(documento_url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'href="{reverse("fatture_scadenze_fornitori")}"')
        self.assertContains(response, pagamento_url)
        self.assertContains(response, "reload_url=")
        self.assertContains(response, f'data-popup-url="{pagamento_url}&reload_url=')
        self.assertContains(response, 'data-window-popup="1"')
        self.assertContains(response, 'data-popup-window-features="width=1120,height=820,resizable=yes,scrollbars=yes"')

    def test_documento_fornitore_popup_mostra_dati_effettivi(self):
        categoria = crea_categoria_spesa_test("Cancelleria")
        fornitore = Fornitore.objects.create(
            denominazione="Cartoleria Test",
            tipo_soggetto="azienda",
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.FATTURA,
            numero_documento="CAR-1",
            data_documento=date(2026, 5, 2),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            importato_at=timezone.make_aware(datetime(2026, 5, 4, 12, 0)),
        )
        fornitore.categoria_spesa = categoria
        fornitore.save(update_fields=["categoria_spesa", "data_aggiornamento"])

        response = self.client.get(f'{reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})}?popup=1')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Data ricezione")
        self.assertContains(response, "04 / 05 / 2026")
        self.assertContains(response, "Cancelleria")
        self.assertContains(response, "Maggio")

        response = self.client.get(reverse("lista_documenti_fornitori"), {"categoria": str(categoria.pk)})
        self.assertContains(response, "CAR-1")

    def test_documento_fornitore_popup_permette_eliminazione_con_doppia_conferma(self):
        fornitore = Fornitore.objects.create(denominazione="FIC Delete Supplier")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FIC-DEL-1",
            data_documento=date(2026, 5, 8),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            external_source="fatture_in_cloud",
            external_id="fic-del-1",
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("122.00"),
        )
        popup_url = f'{reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})}?popup=1'
        delete_url = reverse("elimina_documento_fornitore", kwargs={"pk": documento.pk})
        scadenze_url = reverse("fatture_scadenze_fornitori")

        response = self.client.get(popup_url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="supplier-document-delete-inline-form"')
        self.assertContains(response, f'action="{delete_url}"')
        self.assertContains(response, "Vuoi eliminare definitivamente questa fattura da Arboris?")
        self.assertContains(response, "Seconda conferma")
        self.assertContains(response, f'name="reload_url" value="{scadenze_url}"')
        self.assertContains(response, '<span class="btn-label">Elimina</span>', html=False)

        response = self.client.post(
            delete_url,
            {
                "popup": "1",
                "reload_url": scadenze_url,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "popup/popup_close.html")
        self.assertContains(response, "Fattura fornitore eliminata correttamente.")
        self.assertContains(response, r"gestione\u002Dfinanziaria/fatture\u002Dscadenze\u002Dfornitori")
        self.assertFalse(DocumentoFornitore.objects.filter(pk=documento.pk).exists())
        self.assertFalse(ScadenzaPagamentoFornitore.objects.filter(documento_id=documento.pk).exists())
        alias = DocumentoFornitoreImportAlias.objects.get(
            external_source="fatture_in_cloud",
            external_id="fic-del-1",
        )
        self.assertIsNone(alias.documento)
        self.assertTrue(alias.ignorato)

        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        result = importa_documento_fatture_in_cloud(
            connessione,
            {
                "id": "fic-del-1",
                "type": "expense",
                "description": "Fattura gia eliminata",
                "invoice_number": "FIC-DEL-1",
                "date": "2026-05-08",
                "amount_net": "100.00",
                "amount_vat": "22.00",
                "amount_gross": "122.00",
                "entity": {"name": "FIC Delete Supplier"},
                "payments_list": [{"due_date": "2026-05-31", "amount": "122.00"}],
            },
            pending=False,
            utente=self.user,
        )

        self.assertTrue(result["skipped"])
        self.assertEqual(result["skip_reason"], "alias_ignorato")
        self.assertFalse(DocumentoFornitore.objects.filter(external_id="fic-del-1").exists())

    def test_documento_fornitore_popup_salva_categoria_con_fornitore_inattivo(self):
        categoria = crea_categoria_spesa_test("Categoria aggiornata")
        fornitore = Fornitore.objects.create(
            denominazione="Fornitore importato inattivo",
            tipo_soggetto="azienda",
            attivo=False,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FIC-INATTIVO-1",
            data_documento=date(2026, 5, 9),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            stato=StatoDocumentoFornitore.DA_PAGARE,
            external_source="fatture_in_cloud",
            external_id="fic-inattivo-1",
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("122.00"),
            importo_pagato=Decimal("0.00"),
        )

        response = self.client.post(
            f'{reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})}?popup=1',
            {
                "popup": "1",
                "fornitore": str(fornitore.pk),
                "categoria_spesa": str(categoria.pk),
                "tipo_documento": TipoDocumentoFornitore.FATTURA,
                "numero_documento": documento.numero_documento,
                "data_documento": "2026-05-09",
                "data_ricezione": "",
                "anno_competenza": "2026",
                "mese_competenza": "5",
                "descrizione": "Fattura importata",
                "imponibile": "100.00",
                "aliquota_iva": "22.00",
                "iva": "22.00",
                "totale": "122.00",
                "stato": StatoDocumentoFornitore.DA_PAGARE,
                "note": "",
                "scadenze-TOTAL_FORMS": "1",
                "scadenze-INITIAL_FORMS": "1",
                "scadenze-MIN_NUM_FORMS": "0",
                "scadenze-MAX_NUM_FORMS": "1000",
                "scadenze-0-id": str(scadenza.pk),
                "scadenze-0-data_scadenza": "2026-05-31",
                "scadenze-0-importo_previsto": "122.00",
                "scadenze-0-importo_pagato": "0.00",
                "scadenze-0-data_pagamento": "",
                "scadenze-0-stato": StatoScadenzaFornitore.PREVISTA,
                "scadenze-0-conto_bancario": "",
                "scadenze-0-movimento_finanziario": "",
                "scadenze-0-note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "popup/popup_close.html")
        documento.refresh_from_db()
        self.assertEqual(documento.categoria_spesa, categoria)

    def test_documento_fornitore_popup_mostra_tutti_i_movimenti_collegabili(self):
        fornitore = Fornitore.objects.create(denominazione="Tecnica Srl")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="F-010",
            data_documento=date(2026, 5, 1),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )
        movimento_storico_collegato = MovimentoFinanziario.objects.create(
            data_contabile=date(2024, 1, 2),
            importo=Decimal("-122.00"),
            descrizione="Movimento storico collegato",
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2024, 1, 1),
            importo=Decimal("-50.00"),
            descrizione="Movimento storico non collegato",
        )
        MovimentoFinanziario.objects.bulk_create(
            [
                MovimentoFinanziario(
                    data_contabile=date(2026, 5, 1) + timedelta(days=index),
                    importo=Decimal("-10.00"),
                    descrizione=f"Movimento recente {index}",
                )
                for index in range(130)
            ]
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("122.00"),
            movimento_finanziario=movimento_storico_collegato,
        )

        response = self.client.get(f'{reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})}?popup=1')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "supplier-document-detail-shell is-popup")
        self.assertContains(response, "supplier-document-summary-grid")
        self.assertContains(response, "supplier-deadline-view-list")
        self.assertContains(response, "supplier-document-edit-table")
        self.assertContains(response, "is-view-mode")
        self.assertContains(response, 'id="enable-edit-documento-fornitore-btn"')
        self.assertContains(response, "Aggiungi pagamento")
        self.assertContains(response, "mode-edit-only-table-cell")
        self.assertContains(response, "Movimento storico collegato")
        self.assertContains(response, "Movimento storico non collegato")
        pagamento_url = f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': documento.scadenze.first().pk})}?popup=1"
        self.assertContains(response, pagamento_url)
        self.assertContains(response, "reload_url=")
        self.assertContains(response, f'data-popup-url="{pagamento_url}&reload_url=')
        self.assertContains(response, 'data-window-popup="1"')
        self.assertContains(response, 'data-popup-window-features="width=1120,height=820,resizable=yes,scrollbars=yes"')

    def test_notifica_fattura_ricevuta_apre_fattura_in_popup(self):
        fornitore = Fornitore.objects.create(denominazione="Cloud Supplier Srl")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FC-42",
            data_documento=date(2026, 4, 20),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )
        documento_url = reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})
        NotificaFinanziaria.objects.create(
            titolo="Nuova fattura fornitore ricevuta",
            messaggio="Cloud Supplier Srl - FC-42 - EUR 122.00",
            tipo="fattura_ricevuta",
            url=documento_url,
            documento=documento,
        )

        response = self.client.get(reverse("lista_notifiche_finanziarie"))

        popup_url = f"{documento_url}?popup=1"
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'href="{popup_url}"', count=2)
        self.assertContains(response, f'data-popup-url="{popup_url}"', count=2)
        self.assertContains(response, 'data-window-popup="1"', count=2)
        self.assertNotContains(response, f'href="{documento_url}"')

    def test_campanella_header_permette_di_segnare_notifiche_lette(self):
        prima = NotificaFinanziaria.objects.create(
            titolo="Nuovo movimento bancario",
            messaggio="Conto principale - EUR 25.00",
            tipo="movimento_bancario",
            url=reverse("lista_movimenti_finanziari"),
        )
        seconda = NotificaFinanziaria.objects.create(
            titolo="Secondo movimento bancario",
            messaggio="Conto principale - EUR -12.00",
            tipo="movimento_bancario",
            url=reverse("lista_movimenti_finanziari"),
        )
        dashboard_url = reverse("dashboard_gestione_finanziaria")

        response = self.client.get(dashboard_url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "header-notification-read-form", count=2)
        self.assertContains(response, reverse("segna_tutte_notifiche_finanziarie_lette"))
        self.assertContains(response, reverse("segna_notifica_finanziaria_letta", kwargs={"pk": prima.pk}))
        self.assertContains(response, '<input type="hidden" name="next" value="/gestione-finanziaria/">', html=False)

        response = self.client.post(
            reverse("segna_notifica_finanziaria_letta", kwargs={"pk": prima.pk}),
            {"next": dashboard_url},
        )
        self.assertRedirects(response, dashboard_url)
        self.assertTrue(NotificaFinanziariaLettura.objects.filter(notifica=prima, user=self.user).exists())
        self.assertFalse(NotificaFinanziariaLettura.objects.filter(notifica=seconda, user=self.user).exists())

        response = self.client.post(
            reverse("segna_tutte_notifiche_finanziarie_lette"),
            {"next": dashboard_url},
        )
        self.assertRedirects(response, dashboard_url)
        self.assertTrue(NotificaFinanziariaLettura.objects.filter(notifica=seconda, user=self.user).exists())

    def test_documento_fornitore_calculates_net_and_vat_from_total(self):
        categoria = crea_categoria_spesa_test("Servizi")
        fornitore = Fornitore.objects.create(
            denominazione="Servizi Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )

        response = self.client.post(
            reverse("crea_documento_fornitore"),
            {
                "fornitore": str(fornitore.pk),
                "categoria_spesa": "",
                "tipo_documento": "fattura",
                "numero_documento": "F-002",
                "data_documento": "2026-04-15",
                "data_ricezione": "",
                "anno_competenza": "",
                "mese_competenza": "4",
                "descrizione": "",
                "imponibile": "0.00",
                "aliquota_iva": "22.00",
                "iva": "",
                "totale": "122.00",
                "stato": StatoDocumentoFornitore.DA_PAGARE,
                "note": "",
                "scadenze-TOTAL_FORMS": "0",
                "scadenze-INITIAL_FORMS": "0",
                "scadenze-MIN_NUM_FORMS": "0",
                "scadenze-MAX_NUM_FORMS": "1000",
            },
        )

        documento = DocumentoFornitore.objects.get(numero_documento="F-002")
        self.assertRedirects(response, reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk}))
        self.assertEqual(documento.imponibile, Decimal("100.00"))
        self.assertEqual(documento.iva, Decimal("22.00"))
        self.assertEqual(documento.totale, Decimal("122.00"))
        self.assertEqual(documento.mese_competenza, 4)

    def test_documento_fornitore_accepts_italian_currency_format(self):
        categoria = crea_categoria_spesa_test("Pulizie")
        fornitore = Fornitore.objects.create(
            denominazione="Pulizie Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )

        response = self.client.post(
            reverse("crea_documento_fornitore"),
            {
                "fornitore": str(fornitore.pk),
                "categoria_spesa": "",
                "tipo_documento": "fattura",
                "numero_documento": "F-IT",
                "data_documento": "2026-04-15",
                "data_ricezione": "",
                "anno_competenza": "",
                "mese_competenza": "",
                "descrizione": "",
                "imponibile": "1.000,00",
                "aliquota_iva": "22.00",
                "iva": "",
                "totale": "",
                "stato": StatoDocumentoFornitore.DA_PAGARE,
                "note": "",
                "scadenze-TOTAL_FORMS": "0",
                "scadenze-INITIAL_FORMS": "0",
                "scadenze-MIN_NUM_FORMS": "0",
                "scadenze-MAX_NUM_FORMS": "1000",
            },
        )

        documento = DocumentoFornitore.objects.get(numero_documento="F-IT")
        self.assertRedirects(response, reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk}))
        self.assertEqual(documento.imponibile, Decimal("1000.00"))
        self.assertEqual(documento.iva, Decimal("220.00"))
        self.assertEqual(documento.totale, Decimal("1220.00"))

    def test_documento_fornitore_allegato_uses_supplier_prefix(self):
        categoria = crea_categoria_spesa_test("Materiali")
        fornitore = Fornitore.objects.create(
            denominazione="Upload Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )
        allegato = SimpleUploadedFile("fattura.pdf", b"pdf-content", content_type="application/pdf")

        response = self.client.post(
            reverse("crea_documento_fornitore"),
            {
                "fornitore": str(fornitore.pk),
                "categoria_spesa": "",
                "tipo_documento": "fattura",
                "numero_documento": "F-UP",
                "data_documento": "2026-04-15",
                "data_ricezione": "",
                "anno_competenza": "",
                "mese_competenza": "",
                "descrizione": "",
                "imponibile": "10.00",
                "aliquota_iva": "22.00",
                "iva": "",
                "totale": "",
                "stato": StatoDocumentoFornitore.DA_PAGARE,
                "allegato": allegato,
                "note": "",
                "scadenze-TOTAL_FORMS": "0",
                "scadenze-INITIAL_FORMS": "0",
                "scadenze-MIN_NUM_FORMS": "0",
                "scadenze-MAX_NUM_FORMS": "1000",
            },
        )

        documento = DocumentoFornitore.objects.get(numero_documento="F-UP")
        self.assertRedirects(response, reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk}))
        self.assertTrue(documento.allegato.name.startswith("documenti_fornitori/"))

    def test_scadenza_auto_status_allows_manual_override(self):
        categoria = crea_categoria_spesa_test("Utenze")
        fornitore = Fornitore.objects.create(
            denominazione="Acqua Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            categoria_spesa=categoria,
            numero_documento="SCAD-1",
            data_documento=date(2026, 4, 20),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )

        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2000, 1, 1),
            importo_previsto=Decimal("122.00"),
            importo_pagato=Decimal("0.00"),
            stato=StatoScadenzaFornitore.PREVISTA,
        )
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.SCADUTA)

        scadenza.stato = StatoScadenzaFornitore.PREVISTA
        scadenza._preserve_manual_stato = True
        scadenza.save()
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PREVISTA)

    def test_importa_documento_fatture_in_cloud_crea_documento_scadenza_notifica(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 987,
            "type": "expense",
            "description": "Consulenza mensile",
            "invoice_number": "FC-42",
            "date": "2026-04-20",
            "next_due_date": "2026-05-20",
            "amount_net": "1000.00",
            "amount_vat": "220.00",
            "amount_gross": "1220.00",
            "entity": {
                "name": "Cloud Supplier Srl",
                "vat_number": "IT12345678901",
                "tax_code": "12345678901",
                "address_street": "Via Nuvola 7",
                "address_postal_code": "40100",
                "address_city": "Bologna",
                "address_province": "BO",
                "email": "info@example.com",
                "certified_email": "cloud@examplepec.it",
                "phone": "051123456",
                "ei_code": "ABC1234",
                "bank_iban": "IT60X0542811101000000123456",
                "bank_name": "Banca Cloud",
            },
            "payments_list": [
                {
                    "due_date": "2026-05-20",
                    "amount": "1220.00",
                    "status": "not_paid",
                }
            ],
        }

        result = importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        self.assertTrue(result["created"])
        self.assertTrue(result["fornitore_created"])
        self.assertFalse(result["fornitore_updated"])
        documento = DocumentoFornitore.objects.get(external_id="987")
        self.assertEqual(documento.fornitore.denominazione, "Cloud Supplier Srl")
        self.assertEqual(documento.fornitore.partita_iva, "12345678901")
        self.assertEqual(documento.fornitore.codice_fiscale, "12345678901")
        self.assertEqual(documento.fornitore.indirizzo, "Via Nuvola 7 40100 Bologna BO")
        self.assertEqual(documento.fornitore.email, "info@example.com")
        self.assertEqual(documento.fornitore.pec, "cloud@examplepec.it")
        self.assertEqual(documento.fornitore.telefono, "051123456")
        self.assertEqual(documento.fornitore.codice_sdi, "ABC1234")
        self.assertEqual(documento.fornitore.iban, "IT60X0542811101000000123456")
        self.assertEqual(documento.fornitore.banca, "Banca Cloud")
        self.assertEqual(documento.numero_documento, "FC-42")
        self.assertEqual(documento.totale, Decimal("1220.00"))
        self.assertEqual(documento.origine, "fatture_in_cloud")
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.data_scadenza, date(2026, 5, 20))
        self.assertEqual(scadenza.importo_previsto, Decimal("1220.00"))
        self.assertTrue(NotificaFinanziaria.objects.filter(documento=documento).exists())

        result = importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)
        self.assertFalse(result["created"])
        self.assertFalse(result["fornitore_created"])
        self.assertFalse(result["fornitore_updated"])
        self.assertEqual(DocumentoFornitore.objects.filter(external_id="987").count(), 1)
        self.assertEqual(NotificaFinanziaria.objects.filter(documento=documento).count(), 1)

    def test_importa_documento_fatture_in_cloud_aggancia_duplicato_con_id_e_numero_varianti(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload_pending = {
            "id": "pending-42",
            "type": "agyo",
            "document_type": "invoice",
            "subject": "Fattura da registrare",
            "supplier_name": "Studio Rossi",
            "invoice_number": "42/PA",
            "emission_date": "2026-05-04",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "payments_list": [{"due_date": "2026-06-04", "amount": "122.00"}],
        }
        payload_registrato = {
            "id": "registered-42",
            "type": "expense",
            "description": "Fattura registrata",
            "invoice_number": "42 PA",
            "date": "2026-05-04",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {
                "name": "Studio Rossi S.r.l.",
                "vat_number": "IT12345678901",
            },
            "payments_list": [{"due_date": "2026-06-04", "amount": "122.00"}],
        }

        result = importa_documento_fatture_in_cloud(connessione, payload_pending, pending=True, utente=self.user)
        self.assertTrue(result["created"])
        documento = DocumentoFornitore.objects.get(external_id="pending-42")
        documento_id = documento.pk

        result = importa_documento_fatture_in_cloud(connessione, payload_registrato, pending=False, utente=self.user)

        self.assertFalse(result["created"])
        self.assertEqual(DocumentoFornitore.objects.count(), 1)
        self.assertEqual(Fornitore.objects.count(), 1)
        documento.refresh_from_db()
        self.assertEqual(documento.pk, documento_id)
        self.assertEqual(documento.external_id, "registered-42")
        self.assertEqual(documento.numero_documento, "42 PA")
        self.assertEqual(documento.external_type, "expense")
        self.assertEqual(documento.scadenze.count(), 1)
        self.assertEqual(documento.fornitore.partita_iva, "12345678901")

    def test_importa_documento_fatture_in_cloud_nota_credito_non_crea_scadenze(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 993,
            "type": "expense",
            "description": "Nota di credito",
            "invoice_number": "NC-1",
            "date": "2026-04-21",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Credit Supplier Srl", "vat_number": "IT12345678907"},
            "payments_list": [{"due_date": "2026-05-21", "amount": "122.00"}],
        }

        result = importa_documento_fatture_in_cloud(
            connessione,
            payload,
            pending=False,
            utente=self.user,
            source_doc_type="passive_credit_note",
        )

        self.assertTrue(result["created"])
        self.assertEqual(result["scadenze_create"], 0)
        self.assertEqual(result["pagamenti_auto"], 0)
        documento = DocumentoFornitore.objects.get(external_id="993")
        self.assertEqual(documento.tipo_documento, TipoDocumentoFornitore.NOTA_CREDITO)
        self.assertEqual(documento.stato, StatoDocumentoFornitore.PAGATO)
        self.assertFalse(documento.scadenze.exists())

    def test_importa_documento_fatture_in_cloud_arricchisce_fornitore_esistente(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        fornitore = Fornitore.objects.create(
            denominazione="Fornitore gia censito",
            tipo_soggetto="azienda",
            partita_iva="12345678906",
            email="manuale@example.com",
            attivo=False,
        )
        payload = {
            "id": 992,
            "type": "expense",
            "description": "Fattura con anagrafica completa",
            "invoice_number": "SUP-1",
            "date": "2026-04-24",
            "amount_net": "200.00",
            "amount_vat": "44.00",
            "amount_gross": "244.00",
            "entity": {
                "name": "Nome da Fatture in Cloud",
                "vat_number": "IT12345678906",
                "tax_code": "12345678906",
                "address_street": "Via Dati 10",
                "address_postal_code": "20100",
                "address_city": "Milano",
                "address_province": "MI",
                "email": "fic@example.com",
                "certified_email": "fornitore@examplepec.it",
                "phone": "02123456",
                "ei_code": "XYZ9876",
            },
            "payments_list": [
                {
                    "due_date": "2026-05-24",
                    "amount": "244.00",
                    "iban": "IT60 X054 2811 1010 0000 0123 456",
                    "bank_name": "Banca Test",
                }
            ],
        }

        result = importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        self.assertFalse(result["fornitore_created"])
        self.assertTrue(result["fornitore_updated"])
        self.assertEqual(Fornitore.objects.count(), 1)
        fornitore.refresh_from_db()
        self.assertEqual(fornitore.denominazione, "Fornitore gia censito")
        self.assertEqual(fornitore.email, "manuale@example.com")
        self.assertEqual(fornitore.codice_fiscale, "12345678906")
        self.assertEqual(fornitore.indirizzo, "Via Dati 10 20100 Milano MI")
        self.assertEqual(fornitore.pec, "fornitore@examplepec.it")
        self.assertEqual(fornitore.telefono, "02123456")
        self.assertEqual(fornitore.codice_sdi, "XYZ9876")
        self.assertEqual(fornitore.iban, "IT60X0542811101000000123456")
        self.assertEqual(fornitore.banca, "Banca Test")
        self.assertFalse(fornitore.attivo)

    @patch("gestione_finanziaria.fatture_in_cloud.download_bytes")
    def test_importa_documento_fatture_in_cloud_accetta_url_allegato_lunghi(self, mock_download_bytes):
        mock_download_bytes.return_value = (None, {"download_status": "request_error"})
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        attachment_url = "https://files.example.com/" + ("a" * 1200)
        payload = {
            "id": 988,
            "type": "expense",
            "description": "Documento con URL allegato lungo",
            "invoice_number": "FC-43",
            "date": "2026-04-21",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "attachment_url": attachment_url,
            "entity": {"name": "Long Link Supplier Srl", "vat_number": "IT12345678902"},
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        documento = DocumentoFornitore.objects.get(external_id="988")
        self.assertLessEqual(len(documento.external_url), 1000)
        self.assertEqual(documento.external_payload["attachment_url"], attachment_url)
        self.assertEqual(
            DocumentoFornitore._meta.get_field("external_url").max_length,
            1000,
        )

    @patch("gestione_finanziaria.fatture_in_cloud.download_bytes")
    def test_importa_documento_fatture_in_cloud_salva_allegato_in_fatture_fornitori(self, mock_download_bytes):
        mock_download_bytes.return_value = (
            b"%PDF-1.4\nfattura",
            {
                "download_status": "ok",
                "http_status": 200,
                "content_type": "application/pdf",
                "downloaded_bytes": 16,
                "truncated": False,
            },
        )
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 990,
            "type": "expense",
            "description": "Documento con allegato",
            "invoice_number": "FC/44",
            "date": "2026-04-21",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "attachment_url": "https://files.example.com/fattura.pdf",
            "entity": {"name": "Attachment Supplier Srl", "vat_number": "IT12345678904"},
        }

        result = importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        self.assertTrue(result["created"])
        documento = DocumentoFornitore.objects.get(external_id="990")
        self.assertTrue(documento.allegato.name.startswith("fatture_fornitori/"))
        self.assertTrue(documento.allegato.name.endswith(".pdf"))
        self.assertEqual(documento.allegato.read(), b"%PDF-1.4\nfattura")
        self.assertTrue(documento.external_payload["_arboris_attachment_import"]["saved"])
        self.assertEqual(documento.external_payload["_arboris_attachment_import"]["source"], "attachment_url")

    @patch("gestione_finanziaria.fatture_in_cloud.download_bytes")
    def test_importa_documento_fatture_in_cloud_non_sovrascrive_allegato_esistente(self, mock_download_bytes):
        mock_download_bytes.return_value = (
            b"%PDF-1.4\nprima-versione",
            {
                "download_status": "ok",
                "http_status": 200,
                "content_type": "application/pdf",
                "downloaded_bytes": 23,
                "truncated": False,
            },
        )
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        payload = {
            "id": 991,
            "type": "expense",
            "description": "Documento con allegato",
            "invoice_number": "FC-45",
            "date": "2026-04-21",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "attachment_url": "https://files.example.com/fattura.pdf",
            "entity": {"name": "Attachment Supplier Due Srl", "vat_number": "IT12345678905"},
        }
        importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)
        documento = DocumentoFornitore.objects.get(external_id="991")
        original_name = documento.allegato.name

        mock_download_bytes.reset_mock()
        importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        documento.refresh_from_db()
        self.assertEqual(documento.allegato.name, original_name)
        mock_download_bytes.assert_not_called()

    def test_importa_documento_fatture_in_cloud_legge_fornitore_e_scadenza_da_e_invoice(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 989,
            "type": "expense",
            "description": "Fattura elettronica ricevuta",
            "date": "2026-04-22",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "e_invoice": {
                "dati_generali": {
                    "dati_generali_documento": {
                        "numero": "EI-44",
                        "importo_totale_documento": "122.00",
                    }
                },
                "cedente_prestatore": {
                    "dati_anagrafici": {
                        "id_fiscale_iva": {"id_codice": "12345678903"},
                        "codice_fiscale": "12345678903",
                        "anagrafica": {"denominazione": "E Invoice Supplier Srl"},
                    },
                    "sede": {
                        "indirizzo": "Via Roma 1",
                        "cap": "40100",
                        "comune": "Bologna",
                        "provincia": "BO",
                    },
                    "contatti": {"email": "fatture@example.com"},
                },
                "dati_pagamento": [
                    {
                        "dettaglio_pagamento": [
                            {
                                "data_scadenza_pagamento": "2026-06-15",
                                "importo_pagamento": "122.00",
                            }
                        ]
                    }
                ],
            },
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=True, utente=self.user)

        documento = DocumentoFornitore.objects.get(external_id="989")
        self.assertEqual(documento.fornitore.denominazione, "E Invoice Supplier Srl")
        self.assertEqual(documento.fornitore.partita_iva, "12345678903")
        self.assertEqual(documento.numero_documento, "EI-44")
        self.assertEqual(documento.totale, Decimal("122.00"))
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.data_scadenza, date(2026, 6, 15))
        self.assertEqual(scadenza.importo_previsto, Decimal("122.00"))

    def test_importa_documento_fatture_in_cloud_riconosce_ritenuta_da_e_invoice(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 9981,
            "type": "expense",
            "description": "Consulenza con ritenuta",
            "date": "2026-05-10",
            "amount_net": "1500.00",
            "amount_vat": "330.00",
            "amount_gross": "1830.00",
            "e_invoice": {
                "FatturaElettronicaHeader": {
                    "CedentePrestatore": {
                        "DatiAnagrafici": {
                            "IdFiscaleIVA": {"IdPaese": "IT", "IdCodice": "12345678966"},
                            "Anagrafica": {"Denominazione": "Consulente Ritenuta"},
                        },
                    },
                },
                "FatturaElettronicaBody": {
                    "DatiGenerali": {
                        "DatiGeneraliDocumento": {
                            "Numero": "RA-1",
                            "ImportoTotaleDocumento": "1830.00",
                            "DatiRitenuta": {
                                "TipoRitenuta": "RT01",
                                "ImportoRitenuta": "300.00",
                                "AliquotaRitenuta": "20.00",
                                "CausalePagamento": "A",
                            },
                        }
                    },
                    "DatiPagamento": [
                        {
                            "DettaglioPagamento": [
                                {
                                    "DataScadenzaPagamento": "2026-06-10",
                                    "ImportoPagamento": "1530.00",
                                }
                            ]
                        }
                    ],
                },
            },
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=True, utente=self.user)

        documento = DocumentoFornitore.objects.get(external_id="9981")
        self.assertEqual(documento.ritenuta_acconto, Decimal("300.00"))
        self.assertEqual(documento.aliquota_ritenuta_acconto, Decimal("20.00"))
        self.assertEqual(documento.imponibile_ritenuta_acconto, Decimal("1500.00"))
        self.assertEqual(documento.totale_da_pagare, Decimal("1530.00"))
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.importo_previsto, Decimal("1530.00"))

        response = self.client.get(reverse("fatture_scadenze_fornitori"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "supplier-withholding-badge")
        self.assertContains(response, "R.A.")

    def test_importa_documento_fatture_in_cloud_usa_descrizione_dalle_righe_xml(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 995,
            "type": "expense",
            "date": "2026-05-04",
            "amount_net": "200.00",
            "amount_vat": "44.00",
            "e_invoice": {
                "FatturaElettronicaHeader": {
                    "CedentePrestatore": {
                        "DatiAnagrafici": {
                            "IdFiscaleIVA": {"IdPaese": "IT", "IdCodice": "12345678908"},
                            "Anagrafica": {"Denominazione": "Righe Supplier Srl"},
                        },
                    },
                },
                "FatturaElettronicaBody": {
                    "DatiGenerali": {
                        "DatiGeneraliDocumento": {
                            "Numero": "RIGHE-1",
                            "ImportoTotaleDocumento": "244.00",
                        }
                    },
                    "DatiBeniServizi": {
                        "DettaglioLinee": [
                            {"Descrizione": "Servizio mensa maggio"},
                            {"Descrizione": "Materiale didattico"},
                        ]
                    },
                    "DatiPagamento": [
                        {
                            "DettaglioPagamento": [
                                {
                                    "DataScadenzaPagamento": "2026-06-04",
                                    "ImportoPagamento": "244.00",
                                }
                            ]
                        }
                    ],
                },
            },
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=True, utente=self.user)

        documento = DocumentoFornitore.objects.get(external_id="995")
        self.assertEqual(documento.descrizione, "Servizio mensa maggio; Materiale didattico")
        self.assertEqual(documento.descrizione_righe_fattura, "Servizio mensa maggio\nMateriale didattico")

    def test_importa_documento_fatture_in_cloud_usa_causale_se_mancano_le_righe(self):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        payload = {
            "id": 996,
            "type": "expense",
            "date": "2026-05-05",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "e_invoice": {
                "FatturaElettronicaHeader": {
                    "CedentePrestatore": {
                        "DatiAnagrafici": {
                            "IdFiscaleIVA": {"IdPaese": "IT", "IdCodice": "12345678909"},
                            "Anagrafica": {"Denominazione": "Causale Supplier Srl"},
                        },
                    },
                },
                "FatturaElettronicaBody": {
                    "DatiGenerali": {
                        "DatiGeneraliDocumento": {
                            "Numero": "CAUS-1",
                            "ImportoTotaleDocumento": "122.00",
                            "Causale": ["Assistenza educativa mese di maggio"],
                        }
                    },
                    "DatiPagamento": [
                        {
                            "DettaglioPagamento": [
                                {
                                    "DataScadenzaPagamento": "2026-06-05",
                                    "ImportoPagamento": "122.00",
                                }
                            ]
                        }
                    ],
                },
            },
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=True, utente=self.user)

        documento = DocumentoFornitore.objects.get(external_id="996")
        self.assertEqual(documento.descrizione, "Assistenza educativa mese di maggio")

    def test_importa_documento_fatture_in_cloud_marca_pagata_da_status_pagamento(self):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        payload = {
            "id": 997,
            "type": "expense",
            "description": "Documento gia pagato in FIC",
            "invoice_number": "PAID-1",
            "date": "2026-05-06",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Paid Supplier Srl", "vat_number": "IT12345678910"},
            "payments_list": [
                {
                    "due_date": "2026-05-20",
                    "amount": "122.00",
                    "status": "paid",
                    "paid_date": "2026-05-18",
                }
            ],
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        documento = DocumentoFornitore.objects.get(external_id="997")
        self.assertEqual(documento.stato, StatoDocumentoFornitore.PAGATO)
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.importo_pagato, Decimal("122.00"))
        self.assertEqual(scadenza.data_pagamento, date(2026, 5, 18))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)

    def test_importa_documento_fatture_in_cloud_riconosce_movimento_bancario_pagato(self):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        conto = ContoBancario.objects.create(nome_conto="Conto operativo")
        movimento = MovimentoFinanziario.objects.create(
            conto=conto,
            data_contabile=date(2026, 5, 20),
            importo=Decimal("-122.00"),
            descrizione="Bonifico Auto Match Supplier Srl fattura AM-1",
            controparte="Auto Match Supplier Srl",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        payload = {
            "id": 998,
            "type": "expense",
            "description": "Servizio gia saldato tramite banca",
            "invoice_number": "AM-1",
            "date": "2026-05-07",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Auto Match Supplier Srl", "vat_number": "IT12345678911"},
            "payments_list": [{"due_date": "2026-05-20", "amount": "122.00", "status": "not_paid"}],
        }

        result = importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        self.assertEqual(result["pagamenti_auto"], 1)
        documento = DocumentoFornitore.objects.get(external_id="998")
        self.assertEqual(documento.stato, StatoDocumentoFornitore.PAGATO)
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.importo_pagato, Decimal("122.00"))
        movimento.refresh_from_db()
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(
            PagamentoFornitore.objects.filter(scadenza=scadenza, movimento_finanziario=movimento).count(),
            1,
        )

    def test_importa_documento_fatture_in_cloud_legge_fornitore_da_header_xml_standard(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 994,
            "type": "expense",
            "description": "Fattura elettronica standard",
            "date": "2026-05-02",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "e_invoice": {
                "FatturaElettronicaHeader": {
                    "CedentePrestatore": {
                        "DatiAnagrafici": {
                            "IdFiscaleIVA": {"IdPaese": "IT", "IdCodice": "12345678907"},
                            "CodiceFiscale": "ABCDEF12G34H567I",
                            "Anagrafica": {"Denominazione": "Header Supplier Srl"},
                        },
                        "Sede": {
                            "Indirizzo": "Via Header 4",
                            "CAP": "40121",
                            "Comune": "Bologna",
                            "Provincia": "BO",
                        },
                        "Contatti": {
                            "Telefono": "051999",
                            "Email": "header@example.com",
                            "PECMail": "header@examplepec.it",
                        },
                    },
                },
                "FatturaElettronicaBody": {
                    "DatiGenerali": {
                        "DatiGeneraliDocumento": {
                            "Numero": "STD-55",
                            "ImportoTotaleDocumento": "122.00",
                        }
                    },
                    "DatiPagamento": [
                        {
                            "DettaglioPagamento": [
                                {
                                    "DataScadenzaPagamento": "2026-06-30",
                                    "ImportoPagamento": "122.00",
                                    "IBAN": "IT60X0542811101000000123456",
                                    "IstitutoFinanziario": "Banca Header",
                                }
                            ]
                        }
                    ],
                },
            },
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=True, utente=self.user)

        documento = DocumentoFornitore.objects.get(external_id="994")
        fornitore = documento.fornitore
        self.assertEqual(fornitore.denominazione, "Header Supplier Srl")
        self.assertEqual(fornitore.partita_iva, "12345678907")
        self.assertEqual(fornitore.codice_fiscale, "ABCDEF12G34H567I")
        self.assertEqual(fornitore.indirizzo, "Via Header 4 40121 Bologna BO")
        self.assertEqual(fornitore.telefono, "051999")
        self.assertEqual(fornitore.email, "header@example.com")
        self.assertEqual(fornitore.pec, "header@examplepec.it")
        self.assertEqual(fornitore.codice_sdi, "")
        self.assertEqual(fornitore.iban, "IT60X0542811101000000123456")
        self.assertEqual(fornitore.banca, "Banca Header")
        self.assertEqual(documento.numero_documento, "STD-55")
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.data_scadenza, date(2026, 6, 30))

    def test_importa_documento_fatture_in_cloud_pending_usa_supplier_name_e_scadenza(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        payload = {
            "id": 993,
            "type": "agyo",
            "document_type": "expense",
            "subject": "Fattura da registrare",
            "supplier_name": "Fornitore Pending Srl",
            "invoice_number": "PEND-1",
            "emssion_date": "2026-04-25",
            "amount_net": "300.00",
            "amount_vat": "66.00",
            "amount_gross": "366.00",
            "payments_list": [
                {
                    "due_date": "2026-06-10",
                    "amount": "366.00",
                    "status": "not_paid",
                }
            ],
        }

        result = importa_documento_fatture_in_cloud(connessione, payload, pending=True, utente=self.user)

        self.assertTrue(result["created"])
        self.assertTrue(result["fornitore_created"])
        documento = DocumentoFornitore.objects.get(external_id="993")
        self.assertEqual(documento.fornitore.denominazione, "Fornitore Pending Srl")
        self.assertEqual(documento.tipo_documento, TipoDocumentoFornitore.FATTURA)
        self.assertEqual(documento.data_documento, date(2026, 4, 25))
        self.assertEqual(documento.numero_documento, "PEND-1")
        self.assertEqual(documento.totale, Decimal("366.00"))
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.data_scadenza, date(2026, 6, 10))
        self.assertEqual(scadenza.importo_previsto, Decimal("366.00"))

    def test_importa_documento_fatture_in_cloud_aggiorna_scadenza_importata_non_pagata(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        fornitore = Fornitore.objects.create(denominazione="Fornitore temporaneo")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="OLD-1",
            data_documento=date(2026, 4, 22),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            external_source="fatture_in_cloud",
            external_id="990",
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 4, 22),
            importo_previsto=Decimal("122.00"),
        )
        payload = {
            "id": 990,
            "type": "expense",
            "description": "Fattura aggiornata",
            "invoice_number": "NEW-1",
            "date": "2026-04-22",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Supplier Correct Srl", "vat_number": "IT12345678904"},
            "payments_list": [{"due_date": "2026-06-30", "amount": "122.00"}],
        }

        importa_documento_fatture_in_cloud(connessione, payload, pending=False, utente=self.user)

        documento.refresh_from_db()
        self.assertEqual(documento.fornitore.denominazione, "Supplier Correct Srl")
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.data_scadenza, date(2026, 6, 30))

    def test_importa_documento_fatture_in_cloud_nota_credito_rimuove_scadenza_precedente(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
        )
        fornitore = Fornitore.objects.create(denominazione="Fornitore nota")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="NC-OLD",
            data_documento=date(2026, 4, 22),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            external_source="fatture_in_cloud",
            external_id="994",
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 22),
            importo_previsto=Decimal("122.00"),
        )
        payload = {
            "id": 994,
            "type": "expense",
            "document_type": "expense",
            "description": "Nota di credito aggiornata",
            "invoice_number": "NC-OLD",
            "date": "2026-04-22",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Fornitore nota", "vat_number": "IT12345678908"},
            "payments_list": [{"due_date": "2026-05-22", "amount": "122.00"}],
        }

        importa_documento_fatture_in_cloud(
            connessione,
            payload,
            pending=False,
            utente=self.user,
            source_doc_type="passive_credit_note",
        )

        documento.refresh_from_db()
        self.assertEqual(documento.tipo_documento, TipoDocumentoFornitore.NOTA_CREDITO)
        self.assertEqual(documento.stato, StatoDocumentoFornitore.PAGATO)
        self.assertFalse(documento.scadenze.exists())

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_recupera_dettaglio_prima_di_importare(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_da_registrare=False,
        )
        client = Mock()
        client.list_received_documents.side_effect = [
            {"data": [{"id": 991}], "pagination": {"current_page": 1, "last_page": 1}},
            {"data": [], "pagination": {"current_page": 1, "last_page": 1}},
        ]
        client.get_received_document.return_value = {
            "id": 991,
            "type": "expense",
            "description": "Dettaglio completo",
            "invoice_number": "DET-1",
            "date": "2026-04-23",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Detailed Supplier Srl", "vat_number": "IT12345678905"},
            "payments_list": [{"due_date": "2026-07-01", "amount": "122.00"}],
        }
        mock_client_class.return_value = client

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user)

        self.assertEqual(stats["creati"], 1)
        self.assertEqual(stats["fornitori_creati"], 1)
        self.assertEqual(stats["fornitori_aggiornati"], 0)
        self.assertIn("Fornitori: 1 creati, 0 aggiornati.", stats["messaggi"][0])
        client.get_received_document.assert_called_once_with(991)
        documento = DocumentoFornitore.objects.get(external_id="991")
        self.assertEqual(documento.fornitore.denominazione, "Detailed Supplier Srl")
        self.assertEqual(documento.scadenze.get().data_scadenza, date(2026, 7, 1))

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_importa_passive_credit_note_come_nota(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_da_registrare=False,
        )
        client = Mock()

        def list_received(doc_type, *, page=1, per_page=50, data_inizio=None):
            if doc_type == "passive_credit_note":
                return {"data": [{"id": 992, "date": "2026-04-24"}], "pagination": {"current_page": 1, "last_page": 1}}
            return {"data": [], "pagination": {"current_page": 1, "last_page": 1}}

        client.list_received_documents.side_effect = list_received
        client.get_received_document.return_value = {
            "id": 992,
            "type": "expense",
            "document_type": "expense",
            "description": "Dettaglio nota credito",
            "invoice_number": "NC-DET-1",
            "date": "2026-04-24",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Credit Detail Supplier Srl", "vat_number": "IT12345678910"},
            "payments_list": [{"due_date": "2026-07-01", "amount": "122.00"}],
        }
        mock_client_class.return_value = client

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user)

        self.assertEqual(stats["creati"], 1)
        documento = DocumentoFornitore.objects.get(external_id="992")
        self.assertEqual(documento.tipo_documento, TipoDocumentoFornitore.NOTA_CREDITO)
        self.assertEqual(documento.stato, StatoDocumentoFornitore.PAGATO)
        self.assertFalse(documento.scadenze.exists())

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_filtra_da_data_inizio(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_da_registrare=False,
        )
        client = Mock()

        def list_received(doc_type, *, page=1, per_page=50, data_inizio=None):
            self.assertEqual(data_inizio, date(2026, 4, 1))
            if doc_type == "expense":
                return {
                    "data": [
                        {"id": 993, "date": "2026-04-02"},
                        {"id": 994, "date": "2026-03-31"},
                    ],
                    "pagination": {"current_page": 1, "last_page": 1},
                }
            return {"data": [], "pagination": {"current_page": 1, "last_page": 1}}

        client.list_received_documents.side_effect = list_received
        client.get_received_document.return_value = {
            "id": 993,
            "type": "expense",
            "invoice_number": "DATE-1",
            "date": "2026-04-02",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"name": "Date Supplier Srl", "vat_number": "IT12345678911"},
            "payments_list": [{"due_date": "2026-05-02", "amount": "122.00"}],
        }
        mock_client_class.return_value = client

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user, data_inizio=date(2026, 4, 1))

        self.assertEqual(stats["creati"], 1)
        client.get_received_document.assert_called_once_with(993)
        self.assertTrue(DocumentoFornitore.objects.filter(external_id="993").exists())
        self.assertFalse(DocumentoFornitore.objects.filter(external_id="994").exists())

    @patch("gestione_finanziaria.views.sincronizza_fatture_in_cloud")
    def test_sincronizza_fatture_in_cloud_view_accetta_data_inizio(self, mock_sync):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        mock_sync.return_value = {
            "creati": 0,
            "aggiornati": 0,
            "fornitori_creati": 0,
            "fornitori_aggiornati": 0,
            "esito": EsitoSincronizzazione.OK,
            "messaggi": [],
        }

        response = self.client.post(
            reverse("sincronizza_fatture_in_cloud", kwargs={"pk": connessione.pk}),
            {"data_inizio": "2026-01-01"},
        )

        self.assertRedirects(response, reverse("modifica_fatture_in_cloud", kwargs={"pk": connessione.pk}))
        mock_sync.assert_called_once_with(connessione, utente=self.user, data_inizio=date(2026, 1, 1))

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_arricchisce_fornitore_da_entity_supplier(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_da_registrare=False,
        )
        Fornitore.objects.create(denominazione="Supplier Basic Srl", tipo_soggetto="azienda")
        client = Mock()
        client.list_received_documents.side_effect = [
            {"data": [{"id": 995, "entity": {"id": 77, "name": "Supplier Basic Srl"}}], "pagination": {"current_page": 1, "last_page": 1}},
            {"data": [], "pagination": {"current_page": 1, "last_page": 1}},
        ]
        client.get_received_document.return_value = {
            "id": 995,
            "type": "expense",
            "description": "Dettaglio con fornitore minimo",
            "invoice_number": "DET-SUP-1",
            "date": "2026-05-04",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"id": 77, "name": "Supplier Basic Srl", "vat_number": "IT12345678909"},
            "payments_list": [{"due_date": "2026-06-04", "amount": "122.00"}],
        }
        client.get_supplier.return_value = {
            "id": 77,
            "name": "Supplier Basic Srl",
            "vat_number": "IT12345678909",
            "tax_code": "12345678909",
            "address_street": "Via Completa 8",
            "address_postal_code": "40122",
            "address_city": "Bologna",
            "address_province": "BO",
            "email": "fornitore@example.com",
            "certified_email": "fornitore@examplepec.it",
            "phone": "051888",
        }
        mock_client_class.return_value = client

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user)

        self.assertEqual(stats["creati"], 1)
        self.assertEqual(stats["fornitori_creati"], 0)
        self.assertEqual(stats["fornitori_aggiornati"], 1)
        client.get_supplier.assert_called_once_with("77")
        fornitore = Fornitore.objects.get(denominazione="Supplier Basic Srl")
        self.assertEqual(fornitore.partita_iva, "12345678909")
        self.assertEqual(fornitore.codice_fiscale, "12345678909")
        self.assertEqual(fornitore.indirizzo, "Via Completa 8 40122 Bologna BO")
        self.assertEqual(fornitore.email, "fornitore@example.com")
        self.assertEqual(fornitore.pec, "fornitore@examplepec.it")
        self.assertEqual(fornitore.telefono, "051888")

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_importa_documento_se_supplier_scope_manca(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_da_registrare=False,
        )
        client = Mock()
        client.list_received_documents.side_effect = [
            {"data": [{"id": 996, "entity": {"id": 78, "name": "Supplier Scope Srl"}}], "pagination": {"current_page": 1, "last_page": 1}},
            {"data": [], "pagination": {"current_page": 1, "last_page": 1}},
        ]
        client.get_received_document.return_value = {
            "id": 996,
            "type": "expense",
            "invoice_number": "NO-SCOPE-1",
            "date": "2026-05-04",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
            "entity": {"id": 78, "name": "Supplier Scope Srl"},
            "payments_list": [{"due_date": "2026-06-04", "amount": "122.00"}],
        }
        client.get_supplier.side_effect = FattureInCloudError("Errore API Fatture in Cloud 403")
        mock_client_class.return_value = client

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user)

        self.assertEqual(stats["creati"], 1)
        self.assertEqual(stats["esito"], EsitoSincronizzazione.PARZIALE)
        self.assertTrue(DocumentoFornitore.objects.filter(numero_documento="NO-SCOPE-1").exists())
        self.assertTrue(any("lettura dei fornitori" in message for message in stats["messaggi"]))

    @patch("gestione_finanziaria.fatture_in_cloud_xml.requests.get")
    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_arricchisce_fornitore_da_xml_allegato_pending(
        self,
        mock_client_class,
        mock_requests_get,
    ):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_registrati=False,
            sincronizza_documenti_da_registrare=True,
        )
        client = Mock()

        def list_pending(doc_type, *, page=1, per_page=50, data_inizio=None):
            if doc_type == "agyo":
                return {"data": [{"id": 998}], "pagination": {"current_page": 1, "last_page": 1}}
            return {"data": [], "pagination": {"current_page": 1, "last_page": 1}}

        client.list_pending_received_documents.side_effect = list_pending
        client.get_pending_received_document.return_value = {
            "id": 998,
            "type": "agyo",
            "document_type": "invoice",
            "ei_number": "42",
            "supplier_name": "Fornitore da XML",
            "subject": "Documento pending con XML",
            "filename": "fattura.xml",
            "attachment_url": "https://fic.example.test/download/998",
            "emission_date": "2026-05-04",
            "amount_net": "100.00",
            "amount_vat": "22.00",
            "amount_gross": "122.00",
        }
        mock_client_class.return_value = client
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
<FatturaElettronica>
  <FatturaElettronicaHeader>
    <CedentePrestatore>
      <DatiAnagrafici>
        <IdFiscaleIVA><IdPaese>IT</IdPaese><IdCodice>12345678901</IdCodice></IdFiscaleIVA>
        <CodiceFiscale>12345678901</CodiceFiscale>
        <Anagrafica><Nome>Mario</Nome><Cognome>Rossi</Cognome></Anagrafica>
      </DatiAnagrafici>
      <Sede><Indirizzo>Via Completa 9</Indirizzo><CAP>40100</CAP><Comune>Bologna</Comune><Provincia>BO</Provincia></Sede>
      <Contatti><Email>fornitore.xml@example.com</Email></Contatti>
    </CedentePrestatore>
  </FatturaElettronicaHeader>
  <FatturaElettronicaBody>
    <DatiGenerali>
      <DatiGeneraliDocumento>
        <TipoDocumento>TD01</TipoDocumento>
        <Data>2026-05-04</Data>
        <Numero>42</Numero>
        <DatiRitenuta>
          <TipoRitenuta>RT01</TipoRitenuta>
          <ImportoRitenuta>20.00</ImportoRitenuta>
          <AliquotaRitenuta>20.00</AliquotaRitenuta>
          <CausalePagamento>A</CausalePagamento>
        </DatiRitenuta>
        <ImportoTotaleDocumento>122.00</ImportoTotaleDocumento>
      </DatiGeneraliDocumento>
    </DatiGenerali>
    <DatiPagamento>
      <DettaglioPagamento>
        <DataScadenzaPagamento>2026-06-04</DataScadenzaPagamento>
        <ImportoPagamento>102.00</ImportoPagamento>
      </DettaglioPagamento>
    </DatiPagamento>
    <DatiBeniServizi>
      <DettaglioLinee>
        <NumeroLinea>1</NumeroLinea>
        <Descrizione>Affitto locali aprile 2025</Descrizione>
        <PrezzoTotale>102.00</PrezzoTotale>
      </DettaglioLinee>
    </DatiBeniServizi>
  </FatturaElettronicaBody>
</FatturaElettronica>"""
        attachment_response = Mock(status_code=200, headers={"Content-Type": "text/xml"})
        attachment_response.iter_content.return_value = [xml]
        mock_requests_get.return_value = attachment_response

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user)

        self.assertEqual(stats["creati"], 1)
        self.assertEqual(stats["fornitori_creati"], 1)
        self.assertEqual(stats["fornitori_aggiornati"], 0)
        fornitore = Fornitore.objects.get(denominazione="Fornitore da XML")
        self.assertEqual(fornitore.partita_iva, "12345678901")
        self.assertEqual(fornitore.codice_fiscale, "12345678901")
        self.assertEqual(fornitore.indirizzo, "Via Completa 9 40100 Bologna BO")
        self.assertEqual(fornitore.email, "fornitore.xml@example.com")
        documento = DocumentoFornitore.objects.get(external_id="998")
        self.assertEqual(documento.fornitore, fornitore)
        self.assertEqual(documento.descrizione, "Documento pending con XML")
        self.assertEqual(documento.descrizione_righe_fattura, "Affitto locali aprile 2025")
        self.assertEqual(documento.ritenuta_acconto, Decimal("20.00"))
        self.assertEqual(documento.totale_da_pagare, Decimal("102.00"))
        self.assertTrue(documento.allegato.name.startswith("fatture_fornitori/"))
        scadenza = documento.scadenze.get()
        self.assertEqual(scadenza.data_scadenza, date(2026, 6, 4))
        self.assertEqual(scadenza.importo_previsto, Decimal("102.00"))
        self.assertEqual(mock_requests_get.call_count, 2)

    @patch("gestione_finanziaria.management.commands.debug_fatture_in_cloud_payload.FattureInCloudClient")
    def test_debug_fatture_in_cloud_payload_maschera_dati_sensibili(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        client = Mock()
        client.get_received_document.return_value = {
            "id": 997,
            "entity": {"id": 77, "name": "Fornitore Segreto Srl"},
            "e_invoice": {
                "FatturaElettronicaHeader": {
                    "CedentePrestatore": {
                        "DatiAnagrafici": {
                            "IdFiscaleIVA": {"IdCodice": "12345678901"},
                            "Anagrafica": {"Denominazione": "Ragione Segreta Srl"},
                        },
                        "Sede": {"Indirizzo": "Via Segreta 1", "Comune": "Bologna"},
                    }
                }
            },
            "payments_list": [{"iban": "IT60X0000000000000000000000"}],
        }
        mock_client_class.return_value = client
        output = StringIO()

        call_command(
            "debug_fatture_in_cloud_payload",
            "--connessione",
            str(connessione.pk),
            "--document-id",
            "997",
            stdout=output,
        )

        text = output.getvalue()
        self.assertNotIn("Fornitore Segreto", text)
        self.assertNotIn("Ragione Segreta", text)
        self.assertNotIn("Via Segreta", text)
        self.assertNotIn("12345678901", text)
        self.assertNotIn("IT60X", text)
        report = json.loads(text)
        self.assertTrue(report["entity_supplier_fields_present"]["name"])
        self.assertTrue(report["e_invoice_supplier_fields_present"]["vat_number"])
        self.assertTrue(report["e_invoice_supplier_fields_present"]["address_street"])
        self.assertTrue(report["supplier_payment_fields_present"]["bank_iban"])

    @patch("gestione_finanziaria.views.FattureInCloudClient")
    def test_diagnostica_payload_fatture_in_cloud_via_browser_maschera_dati(self, mock_client_class):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        fornitore = Fornitore.objects.create(denominazione="Fornitore Gia Importato", tipo_soggetto="azienda")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FIC-1",
            data_documento=date(2026, 5, 4),
            totale=Decimal("100.00"),
            origine=OrigineDocumentoFornitore.FATTURE_IN_CLOUD,
            external_source="fatture_in_cloud",
            external_id="997",
        )
        client = Mock()
        client.get_received_document.return_value = {
            "id": 997,
            "entity": {"id": 77, "name": "Fornitore Segreto Srl"},
            "e_invoice": {
                "FatturaElettronicaHeader": {
                    "CedentePrestatore": {
                        "DatiAnagrafici": {
                            "IdFiscaleIVA": {"IdCodice": "12345678901"},
                            "Anagrafica": {"Denominazione": "Ragione Segreta Srl"},
                        },
                        "Sede": {"Indirizzo": "Via Segreta 1", "Comune": "Bologna"},
                    }
                }
            },
        }
        mock_client_class.return_value = client

        response = self.client.post(
            reverse("diagnostica_payload_fatture_in_cloud", kwargs={"pk": connessione.pk}),
            {"documento_fornitore": str(documento.pk), "source_type": "registered"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Report mascherato")
        self.assertContains(response, "e_invoice_supplier_fields_present")
        self.assertNotContains(response, "Fornitore Segreto")
        self.assertNotContains(response, "Ragione Segreta")
        self.assertNotContains(response, "Via Segreta")
        self.assertNotContains(response, "12345678901")
        client.get_received_document.assert_called_once_with("997")

    def test_diagnostica_payload_fatture_in_cloud_riservata_admin(self):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)

        response = self.client.get(reverse("diagnostica_payload_fatture_in_cloud", kwargs={"pk": connessione.pk}))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("modifica_fatture_in_cloud", kwargs={"pk": connessione.pk}))

    @patch("gestione_finanziaria.fatture_in_cloud_xml.requests.get")
    @patch("gestione_finanziaria.views.FattureInCloudClient")
    def test_diagnostica_payload_fatture_in_cloud_analizza_allegato_xml_mascherato(
        self,
        mock_client_class,
        mock_requests_get,
    ):
        self.user.is_superuser = True
        self.user.save(update_fields=["is_superuser"])
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        client = Mock()
        client.get_pending_received_document.return_value = {
            "id": 998,
            "supplier_name": "Nome Visibile Solo Nel Payload Reale",
            "attachment_url": "https://fic.example.test/download/998",
            "filename": "fattura-segreta.xml",
        }
        mock_client_class.return_value = client
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
<FatturaElettronica>
  <FatturaElettronicaHeader>
    <CedentePrestatore>
      <DatiAnagrafici>
        <IdFiscaleIVA><IdPaese>IT</IdPaese><IdCodice>12345678901</IdCodice></IdFiscaleIVA>
        <CodiceFiscale>12345678901</CodiceFiscale>
        <Anagrafica><Denominazione>Ragione Segreta Srl</Denominazione></Anagrafica>
      </DatiAnagrafici>
      <Sede><Indirizzo>Via Segreta 1</Indirizzo><CAP>40100</CAP><Comune>Bologna</Comune><Provincia>BO</Provincia></Sede>
      <Contatti><Telefono>051123456</Telefono><Email>segreta@example.com</Email></Contatti>
    </CedentePrestatore>
  </FatturaElettronicaHeader>
</FatturaElettronica>"""
        attachment_response = Mock(status_code=200, headers={"Content-Type": "application/xml"})
        attachment_response.iter_content.return_value = [xml]
        mock_requests_get.return_value = attachment_response

        response = self.client.post(
            reverse("diagnostica_payload_fatture_in_cloud", kwargs={"pk": connessione.pk}),
            {"document_id": "998", "source_type": "pending"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "attachment_analysis")
        self.assertContains(response, "xml_detected")
        self.assertContains(response, "cedente_prestatore_detected")
        self.assertContains(response, "attachment_supplier_fields_present")
        self.assertNotContains(response, "Ragione Segreta")
        self.assertNotContains(response, "Via Segreta")
        self.assertNotContains(response, "12345678901")
        self.assertNotContains(response, "segreta@example.com")
        mock_requests_get.assert_called_once()
        client.get_pending_received_document.assert_called_once_with("998")

    @override_settings(
        FATTURE_IN_CLOUD_API_CONNECT_TIMEOUT_SECONDS=2,
        FATTURE_IN_CLOUD_API_READ_TIMEOUT_SECONDS=6,
    )
    @patch("gestione_finanziaria.fatture_in_cloud.requests.request")
    def test_fatture_in_cloud_client_usa_timeout_api_breve(self, mock_request):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        response = Mock(status_code=200, content=b"{}", text="{}")
        response.json.return_value = {}
        mock_request.return_value = response
        client = FattureInCloudClient(connessione)
        client._headers = Mock(return_value={})

        client.request("GET", "/test")

        self.assertEqual(mock_request.call_args.kwargs["timeout"], (2.0, 6.0))

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_restituisce_parziale_su_errore_dettaglio(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_registrati=False,
            sincronizza_documenti_da_registrare=True,
        )
        client = Mock()

        def list_pending(doc_type, *, page=1, per_page=50, data_inizio=None):
            if doc_type == "agyo":
                return {"data": [{"id": 992}], "pagination": {"current_page": 1, "last_page": 1}}
            return {"data": [], "pagination": {"current_page": 1, "last_page": 1}}

        client.list_pending_received_documents.side_effect = list_pending
        client.get_pending_received_document.side_effect = FattureInCloudError("Timeout Fatture in Cloud")
        mock_client_class.return_value = client

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user)

        self.assertEqual(stats["esito"], EsitoSincronizzazione.PARZIALE)
        self.assertEqual(stats["creati"], 0)
        self.assertIn("documento 992", stats["messaggi"][0])
        connessione.refresh_from_db()
        self.assertEqual(connessione.ultimo_esito, EsitoSincronizzazione.PARZIALE)

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    def test_sincronizza_fatture_in_cloud_pending_usa_tipi_sorgente_fic(self, mock_client_class):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_registrati=False,
            sincronizza_documenti_da_registrare=True,
        )
        client = Mock()
        client.list_pending_received_documents.return_value = {
            "data": [],
            "pagination": {"current_page": 1, "last_page": 1},
        }
        mock_client_class.return_value = client

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user)

        self.assertEqual(stats["esito"], EsitoSincronizzazione.OK)
        called_types = [
            call.args[0]
            for call in client.list_pending_received_documents.call_args_list
        ]
        self.assertEqual(called_types, ["agyo", "mail", "browser"])
        client.get_pending_received_document.assert_not_called()

    def test_fatture_in_cloud_connessione_puo_essere_rimossa(self):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC da rifare",
            company_id=123,
            access_token_cifrato="token-cifrato",
            refresh_token_cifrato="refresh-cifrato",
        )
        FattureInCloudSyncLog.objects.create(
            connessione=connessione,
            esito=EsitoSincronizzazione.OK,
            documenti_creati=2,
        )
        delete_url = reverse("elimina_fatture_in_cloud", kwargs={"pk": connessione.pk})

        response = self.client.get(reverse("modifica_fatture_in_cloud", kwargs={"pk": connessione.pk}))
        self.assertContains(response, delete_url)
        self.assertContains(response, "Rimuovi connessione")

        confirm_response = self.client.get(delete_url)
        self.assertEqual(confirm_response.status_code, 200)
        self.assertContains(confirm_response, "fatture fornitori")
        self.assertContains(confirm_response, "Rimuovi connessione")

        response = self.client.post(delete_url)

        self.assertRedirects(response, reverse("lista_fatture_in_cloud"))
        self.assertFalse(FattureInCloudConnessione.objects.filter(pk=connessione.pk).exists())
        log = FattureInCloudSyncLog.objects.get()
        self.assertIsNone(log.connessione)

    @patch("gestione_finanziaria.fatture_in_cloud.FattureInCloudClient")
    @patch("gestione_finanziaria.fatture_in_cloud.time.monotonic")
    def test_sincronizza_fatture_in_cloud_si_interrompe_prima_del_timeout_worker(
        self,
        mock_monotonic,
        mock_client_class,
    ):
        connessione = FattureInCloudConnessione.objects.create(
            nome="FIC",
            company_id=123,
            sincronizza_documenti_registrati=True,
            sincronizza_documenti_da_registrare=True,
        )
        mock_monotonic.side_effect = [0, 30, 30]
        mock_client_class.return_value = Mock()

        stats = sincronizza_fatture_in_cloud(connessione, utente=self.user, max_seconds=10)

        self.assertEqual(stats["esito"], EsitoSincronizzazione.PARZIALE)
        self.assertTrue(stats["interrotta_per_tempo"])
        self.assertIn("Tempo massimo", stats["messaggi"][0])
        mock_client_class.return_value.list_received_documents.assert_not_called()
        connessione.refresh_from_db()
        self.assertEqual(connessione.ultimo_esito, EsitoSincronizzazione.PARZIALE)

    @override_settings(
        FATTURE_IN_CLOUD_OAUTH_CLIENT_ID="render-client",
        FATTURE_IN_CLOUD_OAUTH_CLIENT_SECRET="render-secret",
        FATTURE_IN_CLOUD_OAUTH_REDIRECT_URI="https://arboris-test.onrender.com/gestione-finanziaria/fatture-in-cloud/callback/",
    )
    def test_fatture_in_cloud_oauth_usa_credenziali_render(self):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC Render")

        self.assertTrue(has_oauth_credentials(connessione))
        auth_url = authorization_url(
            connessione,
            "https://arboris-test.onrender.com/gestione-finanziaria/fatture-in-cloud/callback/",
            "state-test",
        )

        self.assertIn("client_id=render-client", auth_url)
        self.assertIn("state=state-test", auth_url)

    @override_settings(
        FATTURE_IN_CLOUD_OAUTH_CLIENT_ID="render-client",
        FATTURE_IN_CLOUD_OAUTH_CLIENT_SECRET="render-secret",
        FATTURE_IN_CLOUD_OAUTH_REDIRECT_URI="https://arboris-test.onrender.com/gestione-finanziaria/fatture-in-cloud/callback/",
    )
    def test_avvia_oauth_fatture_in_cloud_con_credenziali_render(self):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC Render")

        response = self.client.get(reverse("avvia_oauth_fatture_in_cloud", kwargs={"pk": connessione.pk}))

        self.assertEqual(response.status_code, 302)
        self.assertIn("https://api-v2.fattureincloud.it/oauth/authorize", response["Location"])
        self.assertIn("client_id=render-client", response["Location"])
        self.assertIn("state=", response["Location"])
        self.assertIn("redirect_uri=https%3A%2F%2Farboris-test.onrender.com", response["Location"])

    @override_settings(
        FATTURE_IN_CLOUD_OAUTH_CLIENT_ID="render-client",
        FATTURE_IN_CLOUD_OAUTH_CLIENT_SECRET="render-secret",
        FATTURE_IN_CLOUD_OAUTH_REDIRECT_URI="https://arboris-test.onrender.com/gestione-finanziaria/fatture-in-cloud/callback/",
    )
    @patch("gestione_finanziaria.fatture_in_cloud.requests.request")
    @patch("gestione_finanziaria.fatture_in_cloud.requests.post")
    def test_callback_fatture_in_cloud_legge_company_id_da_data_companies(self, mock_post, mock_request):
        connessione = FattureInCloudConnessione.objects.create(nome="FIC Render", oauth_state="state-test")
        token_response = Mock(
            status_code=200,
            content=b"{}",
            text='{"access_token": "token"}',
        )
        token_response.json.return_value = {
            "access_token": "access-token",
            "refresh_token": "refresh-token",
            "expires_in": 86400,
        }
        companies_response = Mock(
            status_code=200,
            content=b"{}",
            text='{"data": {"companies": [{"id": 456}]}}',
        )
        companies_response.json.return_value = {"data": {"companies": [{"id": 456}]}}
        mock_post.return_value = token_response
        mock_request.return_value = companies_response

        response = self.client.get(
            reverse("callback_fatture_in_cloud"),
            {"code": "auth-code", "state": "state-test"},
        )

        self.assertEqual(response.status_code, 302)
        expected_url = f"{reverse('modifica_fatture_in_cloud', kwargs={'pk': connessione.pk})}?oauth=ok"
        self.assertEqual(response["Location"], expected_url)
        connessione.refresh_from_db()
        self.assertEqual(connessione.company_id, 456)
        self.assertEqual(connessione.oauth_state, "")
        self.assertTrue(connessione.access_token_cifrato)
        self.assertTrue(connessione.refresh_token_cifrato)
        page_response = self.client.get(expected_url)
        self.assertContains(page_response, "Collegamento OAuth completato")
        self.assertContains(page_response, "Company ID collegato: 456")

    def test_riconciliazione_fornitore_collega_movimento_in_uscita(self):
        categoria = crea_categoria_spesa_test("Utenze")
        fornitore = Fornitore.objects.create(
            denominazione="Energia Srl",
            tipo_soggetto="azienda",
            partita_iva="12345678901",
            categoria_spesa=categoria,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.FATTURA,
            numero_documento="E-001",
            data_documento=date(2026, 4, 1),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 4, 30),
            importo_previsto=Decimal("122.00"),
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 29),
            importo=Decimal("-122.00"),
            descrizione="Bonifico Energia Srl fattura E-001",
            controparte="Energia Srl",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidati = trova_scadenze_fornitori_candidate(movimento)
        self.assertEqual(candidati[0].scadenza, scadenza)

        pagamento = riconcilia_movimento_con_scadenza_fornitore(
            movimento,
            scadenza,
            utente=self.user,
        )

        self.assertEqual(pagamento.importo, Decimal("122.00"))
        scadenza.refresh_from_db()
        documento.refresh_from_db()
        movimento.refresh_from_db()
        self.assertEqual(scadenza.importo_pagato, Decimal("122.00"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(documento.stato, StatoDocumentoFornitore.PAGATO)
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(movimento.categoria, categoria)
        self.assertTrue(movimento.categorizzazione_automatica)
        self.assertEqual(importo_movimento_disponibile_fornitori(movimento), Decimal("0.00"))
        self.assertEqual(PagamentoFornitore.objects.count(), 1)

    def test_riconciliazione_fornitore_propone_e_applica_pagamento_cumulativo(self):
        categoria = crea_categoria_spesa_test("Materiali")
        fornitore = Fornitore.objects.create(
            denominazione="Beta Servizi",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )
        documento_a = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.FATTURA,
            numero_documento="B-001",
            data_documento=date(2026, 5, 1),
            imponibile=Decimal("60.00"),
            iva=Decimal("0.00"),
            totale=Decimal("60.00"),
        )
        documento_b = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.FATTURA,
            numero_documento="B-002",
            data_documento=date(2026, 5, 2),
            imponibile=Decimal("40.00"),
            iva=Decimal("0.00"),
            totale=Decimal("40.00"),
        )
        scadenza_a = ScadenzaPagamentoFornitore.objects.create(
            documento=documento_a,
            data_scadenza=date(2026, 5, 15),
            importo_previsto=Decimal("60.00"),
        )
        scadenza_b = ScadenzaPagamentoFornitore.objects.create(
            documento=documento_b,
            data_scadenza=date(2026, 5, 15),
            importo_previsto=Decimal("40.00"),
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 15),
            importo=Decimal("-100.00"),
            descrizione="Bonifico Beta Servizi saldo fatture B-001 B-002",
            controparte="Beta Servizi",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        candidati = trova_scadenze_fornitori_cumulative_candidate(movimento)

        self.assertEqual(len(candidati), 1)
        self.assertEqual({scadenza.pk for scadenza in candidati[0].scadenze}, {scadenza_a.pk, scadenza_b.pk})
        proposte = proposte_riconciliazione_da_movimento(movimento)
        proposta_cumulativa = next(proposta for proposta in proposte if proposta.tipo == "cumulativa")
        self.assertEqual(proposta_cumulativa.kind, "fornitore")
        self.assertEqual(proposta_cumulativa.direction, "movimento_to_targets")
        self.assertEqual(proposta_cumulativa.movimenti, [movimento])
        self.assertEqual(
            {scadenza.pk for scadenza in proposta_cumulativa.targets},
            {scadenza_a.pk, scadenza_b.pk},
        )
        self.assertEqual(proposta_cumulativa.importo_totale, Decimal("100.00"))

        pagamenti = riconcilia_movimento_con_scadenze_fornitore(
            movimento,
            candidati[0].allocazioni,
            utente=self.user,
        )

        self.assertEqual(len(pagamenti), 2)
        movimento.refresh_from_db()
        scadenza_a.refresh_from_db()
        scadenza_b.refresh_from_db()
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(scadenza_a.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(scadenza_b.stato, StatoScadenzaFornitore.PAGATA)
        self.assertEqual(importo_movimento_disponibile_fornitori(movimento), Decimal("0.00"))

    def test_annulla_pagamento_fornitore_rende_movimento_riconciliabile(self):
        fornitore = Fornitore.objects.create(denominazione="Energia Srl", tipo_soggetto="azienda")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.FATTURA,
            numero_documento="E-002",
            data_documento=date(2026, 4, 1),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 4, 30),
            importo_previsto=Decimal("122.00"),
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 29),
            importo=Decimal("-122.00"),
            descrizione="Bonifico Energia Srl fattura E-002",
            controparte="Energia Srl",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )
        pagamento = riconcilia_movimento_con_scadenza_fornitore(movimento, scadenza, utente=self.user)

        annulla_pagamento_fornitore(pagamento)

        movimento.refresh_from_db()
        scadenza.refresh_from_db()
        documento.refresh_from_db()
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.NON_RICONCILIATO)
        self.assertEqual(importo_movimento_disponibile_fornitori(movimento), Decimal("122.00"))
        self.assertEqual(scadenza.importo_pagato, Decimal("0.00"))
        self.assertNotEqual(documento.stato, StatoDocumentoFornitore.PAGATO)
        self.assertEqual(PagamentoFornitore.objects.count(), 0)

    def test_anteprima_riconciliazione_fornitori_non_scrive_prima_della_conferma(self):
        fornitore = Fornitore.objects.create(denominazione="Energia Srl", tipo_soggetto="azienda")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.FATTURA,
            numero_documento="E-003",
            data_documento=date(2026, 4, 1),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 4, 30),
            importo_previsto=Decimal("122.00"),
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 29),
            importo=Decimal("-122.00"),
            descrizione="Bonifico Energia Srl fattura E-003",
            controparte="Energia Srl",
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO,
        )

        preview = anteprima_riconcilia_fornitori_automaticamente()

        movimento.refresh_from_db()
        scadenza.refresh_from_db()
        self.assertEqual(preview["stats"]["proposti"], 1)
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.NON_RICONCILIATO)
        self.assertEqual(scadenza.importo_pagato, Decimal("0.00"))

        risultato = applica_anteprima_riconciliazione_fornitori(
            preview["dettagli"],
            [preview["dettagli"][0]["key"]],
            utente=self.user,
        )

        movimento.refresh_from_db()
        scadenza.refresh_from_db()
        self.assertEqual(risultato["stats"]["riconciliati"], 1)
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(scadenza.importo_pagato, Decimal("122.00"))

    def test_fornitori_pages_render(self):
        categoria = crea_categoria_spesa_test("Materiali")
        dipendente = Dipendente.objects.create(
            nome="Laura",
            cognome="Bianchi",
            codice_fiscale="BNCLRA80A01A944K",
        )
        fornitore = Fornitore.objects.create(
            denominazione="Carta Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
            dipendente_collegato=dipendente,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            categoria_spesa=categoria,
            numero_documento="MAT-1",
            data_documento=date(2026, 4, 20),
            imponibile=Decimal("50.00"),
            iva=Decimal("11.00"),
            totale=Decimal("61.00"),
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("61.00"),
        )

        urls = [
            reverse("dashboard_gestione_finanziaria"),
            reverse("lista_fornitori"),
            reverse("modifica_fornitore", kwargs={"pk": fornitore.pk}),
            reverse("lista_documenti_fornitori"),
            reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk}),
            reverse("scadenziario_fornitori"),
            reverse("fatture_scadenze_fornitori"),
        ]

        for url in urls:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)

        response = self.client.get(reverse("lista_documenti_fornitori"))
        self.assertContains(response, "Fatture fornitori")
        self.assertContains(response, "Data di scadenza")
        self.assertContains(response, "31/05/2026")
        self.assertContains(response, f'{reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})}?popup=1')
        self.assertContains(response, 'data-live-list-form')
        self.assertContains(response, 'data-live-list-input')
        self.assertContains(response, 'id="documenti-fornitori-results"')
        self.assertContains(response, 'title="Seleziona tutto"')
        self.assertNotContains(response, ">Seleziona Tutto<")
        self.assertContains(response, "live-list-search.js")

        response = self.client.get(reverse("lista_fornitori"))
        self.assertContains(response, "Bianchi Laura")
        response = self.client.get(reverse("modifica_fornitore", kwargs={"pk": fornitore.pk}))
        self.assertContains(response, "Dipendente / educatore collegato")
        self.assertContains(response, "Bianchi Laura")

    def test_fatture_scadenze_fornitori_fonde_fatture_e_scadenziario(self):
        categoria = crea_categoria_spesa_test("Servizi")
        fornitore = Fornitore.objects.create(
            denominazione="Fusioni Srl",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )
        documento_da_pagare = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FUS-1",
            data_documento=date(2026, 5, 1),
            descrizione="Canone sede",
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            imponibile_ritenuta_acconto=Decimal("100.00"),
            ritenuta_acconto=Decimal("20.00"),
            stato=StatoDocumentoFornitore.DA_PAGARE,
        )
        scadenza_da_pagare = ScadenzaPagamentoFornitore.objects.create(
            documento=documento_da_pagare,
            data_scadenza=date(2020, 1, 31),
            importo_previsto=Decimal("122.00"),
            importo_pagato=Decimal("0.00"),
        )
        documento_pagato = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FUS-2",
            data_documento=date(2026, 5, 2),
            descrizione="Materiali",
            imponibile=Decimal("50.00"),
            iva=Decimal("0.00"),
            totale=Decimal("50.00"),
            stato=StatoDocumentoFornitore.PAGATO,
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento_pagato,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("50.00"),
            importo_pagato=Decimal("50.00"),
        )
        documento_parziale = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="FUS-3",
            data_documento=date(2026, 5, 3),
            descrizione="Pulizie",
            imponibile=Decimal("100.00"),
            iva=Decimal("0.00"),
            totale=Decimal("100.00"),
            stato=StatoDocumentoFornitore.PARZIALMENTE_PAGATO,
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento_parziale,
            data_scadenza=date(2026, 5, 30),
            importo_previsto=Decimal("100.00"),
            importo_pagato=Decimal("40.00"),
        )

        response = self.client.get(reverse("fatture_scadenze_fornitori"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fatture e scadenze")
        self.assertContains(response, "Tutte le fatture")
        self.assertContains(response, "Solo fatture insolute")
        self.assertContains(response, f'{reverse("fatture_scadenze_fornitori")}?vista=insolute')
        self.assertContains(response, "Totale previsto")
        self.assertContains(response, "Totale pagato")
        self.assertContains(response, "Totale residuo")
        self.assertContains(response, "<th>Scadenza</th>", html=False)
        self.assertContains(response, "<th>Fornitore</th>", html=False)
        self.assertContains(response, "<th>Categoria</th>", html=False)
        self.assertContains(response, "Previsto")
        self.assertContains(response, "Pagato")
        self.assertContains(response, "Residuo")
        self.assertNotContains(response, "<th>Fattura</th>", html=False)
        self.assertNotContains(response, "<th>IVA</th>", html=False)
        self.assertContains(response, "Da pagare")
        self.assertContains(response, "Parziale")
        self.assertContains(response, "Pagata")
        self.assertContains(response, "Scaduta")
        self.assertContains(response, "supplier-withholding-badge")
        self.assertContains(response, "supplier-invoice-row-unpaid", count=1)
        self.assertContains(response, "supplier-invoice-row-partial", count=1)
        self.assertContains(response, "supplier-invoice-row-paid", count=1)
        self.assertContains(
            response,
            f'data-row-popup-url="{reverse("modifica_documento_fornitore", kwargs={"pk": documento_da_pagare.pk})}?popup=1"',
        )
        pagamento_url = (
            f"{reverse('registra_pagamento_scadenza_fornitore', kwargs={'pk': scadenza_da_pagare.pk})}"
            f"?popup=1&reload_url={reverse('fatture_scadenze_fornitori')}"
        )
        self.assertContains(response, "Registra pagamento")
        self.assertContains(response, pagamento_url)

        response = self.client.get(reverse("fatture_scadenze_fornitori"), {"vista": "insolute"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["totale_previsto"], Decimal("272.00"))
        self.assertEqual(response.context["totale_pagato"], Decimal("90.00"))
        self.assertEqual(response.context["totale_residuo"], Decimal("182.00"))
        self.assertContains(response, "Solo fatture insolute")
        self.assertContains(response, "supplier-invoice-row-unpaid", count=1)
        self.assertContains(response, "supplier-invoice-row-partial", count=1)
        self.assertNotContains(response, "supplier-invoice-row-paid")
        self.assertNotContains(response, "Materiali")

    def test_compensa_documento_fornitore_con_nota_credito_esclude_dai_conteggi(self):
        categoria = crea_categoria_spesa_test("Storni fornitori")
        connessione = FattureInCloudConnessione.objects.create(nome="FIC", company_id=123)
        fornitore = Fornitore.objects.create(
            denominazione="Fornitore Compensato",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="COMP-1",
            data_documento=date(2026, 5, 10),
            descrizione="Servizio annullato",
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            stato=StatoDocumentoFornitore.DA_PAGARE,
            external_source="fatture_in_cloud",
            external_id="fic-comp-1",
        )
        scadenza = ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("122.00"),
        )
        nota_credito = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            tipo_documento=TipoDocumentoFornitore.NOTA_CREDITO,
            numero_documento="NC-COMP-1",
            data_documento=date(2026, 5, 12),
            descrizione="Storno servizio annullato",
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            stato=StatoDocumentoFornitore.PAGATO,
        )

        response = self.client.get(f'{reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})}?popup=1')
        self.assertContains(response, "Compensa")
        self.assertContains(response, "NC-COMP-1")

        detail_url = f'{reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})}?popup=1'
        response = self.client.post(
            reverse("compensa_documento_fornitore", kwargs={"pk": documento.pk}),
            {
                "popup": "1",
                "next": detail_url,
                "nota_credito": str(nota_credito.pk),
            },
        )

        self.assertRedirects(response, detail_url)
        documento.refresh_from_db()
        scadenza.refresh_from_db()
        self.assertEqual(documento.stato, StatoDocumentoFornitore.COMPENSATO)
        self.assertEqual(documento.nota_credito_compensazione, nota_credito)
        self.assertIsNotNone(documento.compensata_at)
        self.assertEqual(documento.totale_da_pagare, Decimal("0.00"))
        self.assertEqual(documento.residuo_da_pagare, Decimal("0.00"))
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.ANNULLATA)
        self.assertIn("NC-COMP-1", scadenza.note)

        response = self.client.get(reverse("fatture_scadenze_fornitori"))
        self.assertEqual(response.context["totale_previsto"], Decimal("0.00"))
        self.assertEqual(response.context["totale_pagato"], Decimal("0.00"))
        self.assertEqual(response.context["totale_residuo"], Decimal("0.00"))
        self.assertNotContains(response, "COMP-1")

        response = self.client.get(
            reverse("spese_mensili_dashboard"),
            {"periodo": "solare", "anno": "2026", "mese": "2026-05"},
        )
        selected_month = next(month for month in response.context["month_stats"] if month["key"] == "2026-05")
        self.assertEqual(selected_month["totale_spese"], Decimal("0.00"))
        self.assertEqual(selected_month["residuo"], Decimal("0.00"))
        self.assertEqual(selected_month["spese_count"], 0)
        self.assertNotContains(response, "Servizio annullato")

        result = importa_documento_fatture_in_cloud(
            connessione,
            {
                "id": "fic-comp-1",
                "type": "expense",
                "description": "Servizio annullato aggiornato",
                "invoice_number": "COMP-1",
                "date": "2026-05-10",
                "amount_net": "100.00",
                "amount_vat": "22.00",
                "amount_gross": "122.00",
                "entity": {"name": "Fornitore Compensato"},
                "payments_list": [{"due_date": "2026-06-30", "amount": "122.00"}],
            },
            pending=False,
            utente=self.user,
        )

        documento.refresh_from_db()
        scadenza.refresh_from_db()
        self.assertFalse(result["created"])
        self.assertEqual(result["scadenze_create"], 0)
        self.assertEqual(documento.stato, StatoDocumentoFornitore.COMPENSATO)
        self.assertEqual(scadenza.stato, StatoScadenzaFornitore.ANNULLATA)

    def test_spese_mensili_dashboard_unisce_fatture_spese_e_introiti(self):
        categoria = crea_categoria_spesa_test("Servizi generali")
        fornitore = Fornitore.objects.create(
            denominazione="Supermercato Verde",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="MAG-1",
            data_documento=date(2026, 5, 3),
            descrizione="Materiale didattico da Supermercato Verde",
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            stato=StatoDocumentoFornitore.DA_PAGARE,
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("122.00"),
            importo_pagato=Decimal("0.00"),
        )
        SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.CONTANTI,
            descrizione="Spesa supermercato",
            categoria=categoria,
            fornitore=fornitore,
            data_scadenza=date(2026, 5, 12),
            importo_previsto=Decimal("48.50"),
            importo_pagato=Decimal("48.50"),
        )
        SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.F24,
            descrizione="F24 contributi maggio",
            categoria=categoria,
            data_scadenza=date(2026, 5, 16),
            importo_previsto=Decimal("300.00"),
            importo_pagato=Decimal("120.00"),
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 20),
            importo=Decimal("850.00"),
            descrizione="Incasso rette maggio",
            controparte="Famiglie",
            origine=OrigineMovimento.BANCA,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 21),
            importo=Decimal("-210.50"),
            descrizione="Pagamento utenze maggio",
            controparte="Fornitore Utenze",
            origine=OrigineMovimento.BANCA,
        )

        response = self.client.get(
            reverse("spese_mensili_dashboard"),
            {"periodo": "solare", "anno": "2026", "mese": "2026-05"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Spese mensili")
        self.assertContains(response, "Mag 2026")
        self.assertContains(response, "Materiale didattico")
        self.assertNotContains(response, "da Supermercato Verde")
        self.assertContains(response, "Spesa supermercato")
        self.assertContains(response, "F24 contributi maggio")
        self.assertContains(response, "Incasso rette maggio")
        self.assertContains(response, "Totale introiti del mese")
        self.assertContains(response, "Totale movimenti in uscita del mese")
        self.assertContains(response, "Bilancio del mese")
        self.assertContains(response, "Totale spese e fatture mensili")
        self.assertContains(response, "Residuo da pagare")
        self.assertContains(response, "3 fatture - 2 insolute")
        self.assertContains(response, "monthly-expense-month-value-income")
        self.assertContains(response, "Parziale")
        documento_url = reverse("modifica_documento_fornitore", kwargs={"pk": documento.pk})
        self.assertContains(response, f'data-row-href="{documento_url}"')
        self.assertContains(response, f'data-row-popup-url="{documento_url}?popup=1"')
        self.assertContains(response, "js/core/list-row-links.js")
        selected_month = next(month for month in response.context["month_stats"] if month["key"] == "2026-05")
        self.assertEqual(selected_month["totale_spese"], Decimal("470.50"))
        self.assertEqual(selected_month["residuo"], Decimal("302.00"))
        self.assertEqual(selected_month["uscite_movimenti"], Decimal("210.50"))
        self.assertEqual(selected_month["bilancio"], Decimal("639.50"))
        self.assertEqual(selected_month["bilancio_segno"], "+")
        self.assertEqual(selected_month["bilancio_tone"], "positive")
        self.assertEqual(selected_month["spese_count"], 3)
        self.assertEqual(selected_month["insolute_count"], 2)
        self.assertEqual(len(response.context["selected_category_summary"]), 1)
        categoria_summary = response.context["selected_category_summary"][0]
        self.assertEqual(categoria_summary["categoria"], "Servizi generali")
        self.assertEqual(categoria_summary["count"], 3)
        self.assertEqual(categoria_summary["totale_previsto"], Decimal("470.50"))
        self.assertEqual(categoria_summary["totale_pagato"], Decimal("168.50"))
        self.assertEqual(categoria_summary["totale_residuo"], Decimal("302.00"))
        self.assertEqual(response.context["period_summary"]["totale_entrate"], Decimal("850.00"))
        self.assertEqual(response.context["period_summary"]["totale_uscite"], Decimal("210.50"))
        self.assertEqual(response.context["period_summary"]["differenza"], Decimal("639.50"))
        self.assertEqual(response.context["period_summary"]["differenza_segno"], "+")
        self.assertEqual(response.context["period_summary"]["differenza_tone"], "positive")
        self.assertContains(response, "Totale periodo")
        self.assertContains(response, "Riepilogo totale anno solare 2026")
        self.assertContains(response, "+639,50")
        self.assertContains(response, "monthly-expense-period-delta is-positive")
        self.assertContains(response, "Riepilogo spese per categoria")
        self.assertContains(response, "Spese di Mag 2026 raggruppate per categoria.")
        self.assertContains(response, "supplier-invoice-row-unpaid", count=1)
        self.assertContains(response, "supplier-invoice-row-partial", count=1)
        self.assertContains(response, "supplier-invoice-row-paid", count=1)

        response = self.client.get(
            reverse("spese_mensili_dashboard"),
            {"periodo": "solare", "anno": "2026", "mese": "2026-05", "vista": "insolute"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Materiale didattico")
        self.assertContains(response, "F24 contributi maggio")
        self.assertNotContains(response, "Spesa supermercato")

    def test_spese_mensili_dashboard_ordina_spese_e_fatture(self):
        categoria = crea_categoria_spesa_test("Ordinamento")
        fornitore = Fornitore.objects.create(
            denominazione="Consulenze Blu",
            tipo_soggetto="azienda",
            categoria_spesa=categoria,
        )
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="ORD-1",
            data_documento=date(2026, 5, 5),
            descrizione="Fattura consulenza",
            imponibile=Decimal("200.00"),
            iva=Decimal("0.00"),
            totale=Decimal("200.00"),
            stato=StatoDocumentoFornitore.PARZIALMENTE_PAGATO,
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento,
            data_scadenza=date(2026, 5, 25),
            importo_previsto=Decimal("200.00"),
            importo_pagato=Decimal("20.00"),
        )
        SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.MANUALE,
            descrizione="Spesa importo alto",
            categoria=categoria,
            data_scadenza=date(2026, 5, 20),
            importo_previsto=Decimal("300.00"),
            importo_pagato=Decimal("10.00"),
        )
        SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.MANUALE,
            descrizione="Spesa pagamento alto",
            categoria=categoria,
            data_scadenza=date(2026, 5, 10),
            importo_previsto=Decimal("100.00"),
            importo_pagato=Decimal("80.00"),
        )
        SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.MANUALE,
            descrizione="Spesa saldata",
            categoria=categoria,
            data_scadenza=date(2026, 5, 15),
            importo_previsto=Decimal("50.00"),
            importo_pagato=Decimal("50.00"),
        )

        def get_descriptions(**params):
            response = self.client.get(
                reverse("spese_mensili_dashboard"),
                {"periodo": "solare", "anno": "2026", "mese": "2026-05", **params},
            )
            self.assertEqual(response.status_code, 200)
            return [row["descrizione"] for row in response.context["selected_rows"]], response

        descriptions, response = get_descriptions(ordina="previsto", direzione="desc")
        self.assertEqual(
            descriptions,
            ["Spesa importo alto", "Fattura consulenza", "Spesa pagamento alto", "Spesa saldata"],
        )
        self.assertEqual(response.context["spese_sort_key"], "previsto")
        self.assertEqual(response.context["spese_sort_direction"], "desc")
        self.assertTrue(response.context["spese_sort_links"]["previsto"]["active"])
        self.assertIn("ordina=previsto", response.context["vista_insolute_url"])
        self.assertIn("direzione=desc", response.context["vista_insolute_url"])
        self.assertContains(response, "Importo previsto")
        self.assertContains(response, "DESC")

        descriptions, _ = get_descriptions(ordina="pagato", direzione="asc")
        self.assertEqual(
            descriptions,
            ["Spesa importo alto", "Fattura consulenza", "Spesa saldata", "Spesa pagamento alto"],
        )

        descriptions, _ = get_descriptions(ordina="residuo", direzione="desc")
        self.assertEqual(
            descriptions,
            ["Spesa importo alto", "Fattura consulenza", "Spesa pagamento alto", "Spesa saldata"],
        )

        descriptions, _ = get_descriptions(ordina="scadenza", direzione="desc")
        self.assertEqual(
            descriptions,
            ["Fattura consulenza", "Spesa importo alto", "Spesa saldata", "Spesa pagamento alto"],
        )

        descriptions, response = get_descriptions(ordina="campo-non-valido", direzione="lato")
        self.assertEqual(
            descriptions,
            ["Spesa pagamento alto", "Spesa saldata", "Spesa importo alto", "Fattura consulenza"],
        )
        self.assertEqual(response.context["spese_sort_key"], "scadenza")
        self.assertEqual(response.context["spese_sort_direction"], "asc")

    def test_spese_mensili_dashboard_esclude_ricariche_prepagate_dagli_introiti(self):
        conto_corrente = ContoBancario.objects.create(
            nome_conto="Conto operativo",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
        )
        prepagata = ContoBancario.objects.create(
            nome_conto="Carta web",
            tipo_conto=TipoContoFinanziario.CARTA_PREPAGATA,
        )
        incasso = MovimentoFinanziario.objects.create(
            conto=conto_corrente,
            data_contabile=date(2026, 1, 10),
            importo=Decimal("1000.00"),
            descrizione="Incasso rette gennaio",
            origine=OrigineMovimento.BANCA,
            incide_su_saldo_banca=True,
        )
        ricarica = MovimentoFinanziario.objects.create(
            conto=prepagata,
            data_contabile=date(2026, 1, 12),
            importo=Decimal("300.00"),
            descrizione="Ricarica carta prepagata",
            origine=OrigineMovimento.BANCA,
            canale=CanaleMovimento.PREPAGATA,
            incide_su_saldo_banca=True,
        )
        uscita_prepagata = MovimentoFinanziario.objects.create(
            conto=prepagata,
            data_contabile=date(2026, 1, 13),
            importo=Decimal("-45.00"),
            descrizione="Amazon marketplace",
            origine=OrigineMovimento.BANCA,
            canale=CanaleMovimento.PREPAGATA,
            incide_su_saldo_banca=True,
        )

        response = self.client.get(
            reverse("spese_mensili_dashboard"),
            {"periodo": "solare", "anno": "2026", "mese": "2026-01"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["period_summary"]["totale_entrate"], Decimal("1000.00"))
        self.assertEqual(response.context["selected_introiti_total"], Decimal("1000.00"))
        self.assertEqual(response.context["selected_introiti"], [incasso])
        self.assertContains(response, "Incasso rette gennaio")
        self.assertNotContains(response, ricarica.descrizione)
        self.assertEqual(response.context["selected_uscite_movimenti"], [uscita_prepagata])
        self.assertEqual(response.context["selected_uscite_movimenti_total"], Decimal("45.00"))
        self.assertContains(response, "Movimenti in uscita di Gen 2026")
        self.assertContains(response, "Totale movimenti in uscita registrati")
        self.assertContains(response, "Amazon marketplace")

    def test_spese_mensili_dashboard_riepilogo_periodo_mostra_perdita(self):
        categoria = crea_categoria_spesa_test("Utenze")
        SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.MANUALE,
            descrizione="Energia elettrica",
            categoria=categoria,
            data_scadenza=date(2026, 6, 12),
            importo_previsto=Decimal("180.00"),
            importo_pagato=Decimal("0.00"),
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 6, 20),
            importo=Decimal("50.00"),
            descrizione="Rimborso",
            origine=OrigineMovimento.BANCA,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 6, 21),
            importo=Decimal("-80.00"),
            descrizione="Pagamento materiali",
            origine=OrigineMovimento.BANCA,
        )

        response = self.client.get(
            reverse("spese_mensili_dashboard"),
            {"periodo": "solare", "anno": "2026", "mese": "2026-06"},
        )

        self.assertEqual(response.status_code, 200)
        selected_month = next(month for month in response.context["month_stats"] if month["key"] == "2026-06")
        self.assertEqual(selected_month["uscite_movimenti"], Decimal("80.00"))
        self.assertEqual(selected_month["bilancio"], Decimal("-30.00"))
        self.assertEqual(selected_month["bilancio_segno"], "-")
        self.assertEqual(selected_month["bilancio_tone"], "negative")
        self.assertEqual(response.context["period_summary"]["totale_entrate"], Decimal("50.00"))
        self.assertEqual(response.context["period_summary"]["totale_uscite"], Decimal("80.00"))
        self.assertEqual(response.context["period_summary"]["differenza"], Decimal("-30.00"))
        self.assertEqual(response.context["period_summary"]["differenza_segno"], "-")
        self.assertEqual(response.context["period_summary"]["differenza_tone"], "negative")
        self.assertContains(response, "-30,00")
        self.assertContains(response, "monthly-expense-period-delta is-negative")
        self.assertContains(response, "monthly-expense-month-value-balance is-negative")

    def test_spese_mensili_dashboard_prepara_click_destro_categoria(self):
        padre = CategoriaFinanziaria.objects.create(
            nome="Spese di Gestione",
            tipo=TipoCategoriaFinanziaria.SPESA,
            icona="briefcase",
        )
        figlia = CategoriaFinanziaria.objects.create(
            nome="Utenze e Servizi",
            tipo=TipoCategoriaFinanziaria.SPESA,
            parent=padre,
            icona="bolt",
        )
        entrata = CategoriaFinanziaria.objects.create(
            nome="Rette",
            tipo=TipoCategoriaFinanziaria.ENTRATA,
            icona="coins",
        )
        spesa = SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.MANUALE,
            descrizione="Pagamento energia",
            categoria=figlia,
            data_scadenza=date(2026, 5, 14),
            importo_previsto=Decimal("82.96"),
        )
        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 20),
            importo=Decimal("480.00"),
            descrizione="Incasso retta",
            origine=OrigineMovimento.BANCA,
            categoria=entrata,
        )

        response = self.client.get(
            reverse("spese_mensili_dashboard"),
            {"periodo": "solare", "anno": "2026", "mese": "2026-05"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="monthly-expense-category-options-template"', html=False)
        self.assertContains(response, 'id="monthly-all-category-options-template"', html=False)
        self.assertContains(response, 'data-category-options-template="monthly-expense-category-options-template"', html=False)
        self.assertContains(response, 'data-category-options-template="monthly-all-category-options-template"', html=False)
        self.assertContains(response, reverse("aggiorna_categoria_spesa_operativa", args=[spesa.pk]))
        self.assertContains(response, reverse("aggiorna_categoria_movimento", args=[movimento.pk]))
        self.assertContains(response, 'data-category-name="Spese di Gestione"', html=False)
        self.assertContains(response, 'data-category-has-children="1"', html=False)
        self.assertContains(response, 'data-category-name="Utenze e Servizi"', html=False)
        self.assertContains(response, 'data-category-level="1"', html=False)
        self.assertContains(response, 'data-category-parent="Spese di Gestione"', html=False)
        self.assertContains(response, 'data-category-icon="bolt"', html=False)
        self.assertContains(response, "js/pages/movimenti-list.js")
        self.assertContains(response, "Clic destro sulla cella Categoria")

    def test_aggiorna_categoria_spesa_operativa_da_spese_mensili(self):
        categoria_iniziale = crea_categoria_spesa_test("Cancelleria")
        categoria_nuova = crea_categoria_spesa_test("Utenze")
        spesa = SpesaOperativa.objects.create(
            tipo=TipoSpesaOperativa.MANUALE,
            descrizione="Acquisto materiale",
            categoria=categoria_iniziale,
            data_scadenza=date(2026, 5, 12),
            importo_previsto=Decimal("35.00"),
        )

        response = self.client.post(
            reverse("aggiorna_categoria_spesa_operativa", args=[spesa.pk]),
            {"categoria": categoria_nuova.pk},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["category_id"], str(categoria_nuova.pk))
        self.assertEqual(response.json()["category_label"], "Utenze")
        spesa.refresh_from_db()
        self.assertEqual(spesa.categoria, categoria_nuova)

    def test_aggiorna_categoria_documento_fornitore_da_spese_mensili(self):
        categoria_iniziale = crea_categoria_spesa_test("Materiali")
        categoria_nuova = crea_categoria_spesa_test("Servizi")
        fornitore = Fornitore.objects.create(denominazione="Fornitore Test")
        documento = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="DOC-1",
            data_documento=date(2026, 5, 3),
            totale=Decimal("122.00"),
            categoria_spesa=categoria_iniziale,
        )

        response = self.client.post(
            reverse("aggiorna_categoria_documento_fornitore", args=[documento.pk]),
            {"categoria": categoria_nuova.pk},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["category_id"], str(categoria_nuova.pk))
        self.assertEqual(response.json()["category_label"], "Servizi")
        documento.refresh_from_db()
        self.assertEqual(documento.categoria_spesa, categoria_nuova)

    def test_spese_mensili_dashboard_rende_modificabile_categoria_busta_paga(self):
        categoria = crea_categoria_spesa_test("Dipendenti")
        dipendente = Dipendente.objects.create(
            nome="Mario",
            cognome="Rossi",
            codice_fiscale="RSSMRA80A01H501U",
        )
        busta = BustaPagaDipendente.objects.create(
            dipendente=dipendente,
            anno=2026,
            mese=4,
            costo_azienda_previsto=Decimal("2186.94"),
            categoria=categoria,
        )

        response = self.client.get(
            reverse("spese_mensili_dashboard"),
            {"periodo": "solare", "anno": "2026", "mese": "2026-04", "vista": "tutte"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Busta paga")
        self.assertContains(response, "Rossi Mario")
        self.assertContains(response, "Dipendenti")
        self.assertContains(response, 'data-category-options-template="monthly-expense-category-options-template"', html=False)
        self.assertContains(response, reverse("aggiorna_categoria_busta_paga", args=[busta.pk]))
        self.assertContains(response, f'data-category-id="{categoria.pk}"', html=False)

    def test_aggiorna_categoria_busta_paga_da_spese_mensili(self):
        categoria_iniziale = crea_categoria_spesa_test("Personale")
        categoria_nuova = crea_categoria_spesa_test("Dipendenti")
        dipendente = Dipendente.objects.create(
            nome="Mario",
            cognome="Rossi",
            codice_fiscale="RSSMRA80A01H501U",
        )
        busta = BustaPagaDipendente.objects.create(
            dipendente=dipendente,
            anno=2026,
            mese=4,
            costo_azienda_previsto=Decimal("2186.94"),
            categoria=categoria_iniziale,
        )

        response = self.client.post(
            reverse("aggiorna_categoria_busta_paga", args=[busta.pk]),
            {"categoria": categoria_nuova.pk},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["category_id"], str(categoria_nuova.pk))
        self.assertEqual(response.json()["category_label"], "Dipendenti")
        busta.refresh_from_db()
        self.assertEqual(busta.categoria, categoria_nuova)

    def test_spese_mensili_dashboard_apre_nuova_spesa_in_popup(self):
        response = self.client.get(reverse("spese_mensili_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'{reverse("crea_spesa_operativa")}?popup=1')
        self.assertContains(response, 'data-window-popup="1"')
        self.assertContains(response, 'data-popup-window-name="arboris-spesa-mensile-popup"')
        self.assertContains(response, f'{reverse("crea_piano_rateale_spesa")}?popup=1')
        self.assertContains(response, 'data-popup-window-name="arboris-piano-rateale-spesa-popup"')

    def test_spesa_operativa_form_usa_layout_popup_e_controlli_related(self):
        response = self.client.get(
            f'{reverse("crea_spesa_operativa")}?popup=1&next={reverse("spese_mensili_dashboard")}'
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'body class="popup-page"', html=False)
        self.assertContains(response, "budget-voice-popup-card")
        self.assertContains(response, '<input type="hidden" name="popup" value="1">', html=False)
        self.assertContains(response, 'id="add-spesa-categoria-btn"')
        self.assertContains(response, 'id="edit-spesa-categoria-btn"')
        self.assertContains(response, 'id="delete-spesa-categoria-btn"')
        self.assertContains(response, 'relatedType: "categoria_spesa"')
        self.assertContains(response, 'id="add-spesa-fornitore-btn"')
        self.assertContains(response, 'relatedType: "fornitore"')
        self.assertContains(response, 'id="add-spesa-conto-btn"')
        self.assertContains(response, 'id="edit-spesa-conto-btn"')
        self.assertContains(response, 'id="delete-spesa-conto-btn"')
        self.assertContains(response, 'relatedType: "conto_bancario"')
        self.assertContains(response, "budget-voice-currency-field", count=2)
        self.assertContains(response, 'name="importo_pagato"', html=False)
        self.assertContains(response, 'placeholder="0,00"', html=False)
        self.assertNotContains(response, 'name="importo_pagato" value="0,00"', html=False)
        self.assertNotContains(response, 'name="importo_pagato" value="0.00"', html=False)
        self.assertContains(response, 'name="note"', html=False)
        self.assertContains(response, 'data-rich-notes-skip="true"', html=False)
        self.assertContains(response, 'placeholder="Aggiungi una nota..."', html=False)
        self.assertNotContains(response, "id_dipendente")
        self.assertNotContains(response, "Dipendente")

    def test_crea_spesa_operativa_popup_chiude_e_ricarica_riepilogo(self):
        categoria = crea_categoria_spesa_test("Spese manuali")

        response = self.client.post(
            f'{reverse("crea_spesa_operativa")}?popup=1',
            {
                "popup": "1",
                "next": reverse("spese_mensili_dashboard"),
                "tipo": TipoSpesaOperativa.MANUALE,
                "descrizione": "Acquisto cancelleria",
                "categoria": categoria.pk,
                "fornitore": "",
                "data_scadenza": "2026-05-13",
                "importo_previsto": "25.00",
                "importo_pagato": "",
                "data_pagamento": "",
                "conto_bancario": "",
                "movimento_finanziario": "",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "popup/popup_close.html")
        self.assertEqual(response.context["reload_url"], reverse("spese_mensili_dashboard"))
        spesa = SpesaOperativa.objects.get(descrizione="Acquisto cancelleria")
        self.assertEqual(spesa.importo_pagato, Decimal("0.00"))

    def test_piano_rateale_spesa_form_usa_layout_popup_e_controlli_related(self):
        response = self.client.get(
            f'{reverse("crea_piano_rateale_spesa")}?popup=1&next={reverse("spese_mensili_dashboard")}'
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'body class="popup-page"', html=False)
        self.assertContains(response, "budget-voice-popup-card")
        self.assertContains(response, '<input type="hidden" name="popup" value="1">', html=False)
        self.assertContains(response, 'id="add-piano-categoria-btn"')
        self.assertContains(response, 'id="edit-piano-categoria-btn"')
        self.assertContains(response, 'id="delete-piano-categoria-btn"')
        self.assertContains(response, 'relatedType: "categoria_spesa"')
        self.assertContains(response, 'id="add-piano-fornitore-btn"')
        self.assertContains(response, 'relatedType: "fornitore"')
        self.assertContains(response, "data-plan-supplier-field")
        self.assertContains(response, 'typeSelect.value === "fornitore"')
        self.assertContains(response, "supplierSelect.disabled = !canUseSupplier")
        self.assertContains(response, 'data-rich-notes-skip="true"', html=False)

    def test_crea_piano_rateale_popup_chiude_e_pulisce_fornitore_se_non_richiesto(self):
        categoria = crea_categoria_spesa_test("Finanziamenti")
        fornitore = Fornitore.objects.create(denominazione="Banca Test")

        response = self.client.post(
            f'{reverse("crea_piano_rateale_spesa")}?popup=1',
            {
                "popup": "1",
                "next": reverse("spese_mensili_dashboard"),
                "tipo": TipoPianoRatealeSpesa.FINANZIAMENTO,
                "descrizione": "Finanziamento laboratorio",
                "categoria": categoria.pk,
                "fornitore": fornitore.pk,
                "importo_totale": "100.00",
                "numero_rate": "2",
                "frequenza_mesi": "1",
                "data_prima_scadenza": "2026-05-31",
                "giorno_scadenza": "31",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "popup/popup_close.html")
        self.assertEqual(response.context["reload_url"], reverse("spese_mensili_dashboard"))
        piano = PianoRatealeSpesa.objects.get(descrizione="Finanziamento laboratorio")
        self.assertIsNone(piano.fornitore)
        self.assertEqual(piano.rate.count(), 2)

    def test_crea_piano_rateale_spesa_genera_rate(self):
        categoria = crea_categoria_spesa_test("Finanziamenti")

        response = self.client.post(
            reverse("crea_piano_rateale_spesa"),
            {
                "tipo": TipoPianoRatealeSpesa.FINANZIAMENTO,
                "descrizione": "Finanziamento cucina",
                "categoria": categoria.pk,
                "fornitore": "",
                "importo_totale": "100.00",
                "numero_rate": "3",
                "frequenza_mesi": "1",
                "data_prima_scadenza": "2026-05-31",
                "giorno_scadenza": "31",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        piano = PianoRatealeSpesa.objects.get(descrizione="Finanziamento cucina")
        rate = list(piano.rate.order_by("numero_rata"))
        self.assertEqual(len(rate), 3)
        self.assertEqual([rata.importo_previsto for rata in rate], [Decimal("33.33"), Decimal("33.33"), Decimal("33.34")])
        self.assertEqual([rata.data_scadenza for rata in rate], [date(2026, 5, 31), date(2026, 6, 30), date(2026, 7, 31)])
        self.assertEqual(rate[0].tipo, TipoSpesaOperativa.FINANZIAMENTO)

    def test_lista_documenti_fornitori_mostra_riepilogo_e_colori_stato(self):
        fornitore = Fornitore.objects.create(
            denominazione="Riepilogo Srl",
            tipo_soggetto="azienda",
        )
        DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="APER-1",
            data_documento=date(2026, 5, 1),
            imponibile=Decimal("100.00"),
            iva=Decimal("22.00"),
            totale=Decimal("122.00"),
            stato=StatoDocumentoFornitore.DA_PAGARE,
        )
        documento_parziale = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="PARZ-1",
            data_documento=date(2026, 5, 2),
            imponibile=Decimal("200.00"),
            iva=Decimal("0.00"),
            totale=Decimal("200.00"),
            stato=StatoDocumentoFornitore.PARZIALMENTE_PAGATO,
        )
        documento_pagato = DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="PAG-1",
            data_documento=date(2026, 5, 3),
            imponibile=Decimal("80.00"),
            iva=Decimal("0.00"),
            totale=Decimal("80.00"),
            stato=StatoDocumentoFornitore.PAGATO,
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento_parziale,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("200.00"),
            importo_pagato=Decimal("50.00"),
        )
        ScadenzaPagamentoFornitore.objects.create(
            documento=documento_pagato,
            data_scadenza=date(2026, 5, 31),
            importo_previsto=Decimal("80.00"),
            importo_pagato=Decimal("80.00"),
        )

        response = self.client.get(reverse("lista_documenti_fornitori"))

        self.assertEqual(response.context["totale_documenti_non_saldati"], Decimal("272.00"))
        self.assertEqual(response.context["numero_documenti_non_saldati"], 2)
        self.assertContains(response, "Totale fatture non saldate")
        self.assertContains(response, "272,00")
        self.assertContains(response, "supplier-invoice-row-unpaid", count=2)
        self.assertContains(response, "supplier-invoice-row-paid", count=1)

    def test_lista_documenti_fornitori_riepilogo_usa_netto_da_pagare(self):
        fornitore = Fornitore.objects.create(
            denominazione="Professionista Riepilogo",
            tipo_soggetto="professionista",
        )
        DocumentoFornitore.objects.create(
            fornitore=fornitore,
            numero_documento="RIT-1",
            data_documento=date(2025, 12, 15),
            imponibile=Decimal("1716.00"),
            iva=Decimal("377.52"),
            totale=Decimal("2093.52"),
            imponibile_ritenuta_acconto=Decimal("1650.00"),
            aliquota_ritenuta_acconto=Decimal("20.00"),
            ritenuta_acconto=Decimal("330.00"),
            stato=StatoDocumentoFornitore.DA_PAGARE,
        )

        response = self.client.get(reverse("lista_documenti_fornitori"))

        self.assertEqual(response.context["totale_documenti_non_saldati"], Decimal("1763.52"))
        self.assertContains(response, "1.763,52")
        self.assertContains(response, "2.093,52")
        self.assertContains(response, "supplier-withholding-badge")
        self.assertContains(response, "R.A.")

    def test_eliminazione_multipla_documenti_fornitori_con_conferma(self):
        fornitore = Fornitore.objects.create(
            denominazione="Fornitore bulk",
            tipo_soggetto="azienda",
        )
        documenti = [
            DocumentoFornitore.objects.create(
                fornitore=fornitore,
                numero_documento=f"BULK-{index}",
                data_documento=date(2026, 4, index),
                imponibile=Decimal("100.00"),
                iva=Decimal("22.00"),
                totale=Decimal("122.00"),
            )
            for index in (1, 2, 3)
        ]
        ScadenzaPagamentoFornitore.objects.create(
            documento=documenti[0],
            data_scadenza=date(2026, 5, 1),
            importo_previsto=Decimal("122.00"),
        )
        next_url = reverse("lista_documenti_fornitori") + "?stato=da_pagare"

        response = self.client.get(reverse("lista_documenti_fornitori"))
        self.assertContains(response, reverse("elimina_documenti_fornitori_multipla"))
        self.assertContains(response, "data-bulk-form")

        response = self.client.post(
            reverse("elimina_documenti_fornitori_multipla"),
            {
                "selected_ids": [str(documenti[0].pk), str(documenti[1].pk)],
                "next": next_url,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Elimina fatture fornitori")
        self.assertContains(response, "BULK-1")
        self.assertContains(response, "BULK-2")

        response = self.client.post(
            reverse("elimina_documenti_fornitori_multipla"),
            {
                "selected_ids": [str(documenti[0].pk), str(documenti[1].pk)],
                "next": next_url,
                "conferma": "1",
            },
        )

        self.assertRedirects(response, next_url)
        self.assertFalse(DocumentoFornitore.objects.filter(pk__in=[documenti[0].pk, documenti[1].pk]).exists())
        self.assertTrue(DocumentoFornitore.objects.filter(pk=documenti[2].pk).exists())
        self.assertFalse(ScadenzaPagamentoFornitore.objects.filter(documento=documenti[0]).exists())

    def test_cbi_csv_autodetect_parses_movements(self):
        detection = detect_csv_import_config(CBI_CSV_SAMPLE.encode("utf-8"))

        self.assertEqual(detection.formato_rilevato, "CSV CBI")
        self.assertEqual(detection.config.delimiter, ";")
        self.assertEqual(detection.abi, "05034")
        self.assertEqual(detection.cab, "37060")
        self.assertEqual(detection.numero_conto, "000000003228")

        movimenti = list(CsvImporter(detection.config).parse(CBI_CSV_SAMPLE.encode("utf-8")))

        self.assertEqual(len(movimenti), 2)
        self.assertEqual(movimenti[0].data_contabile, date(2026, 4, 24))
        self.assertEqual(movimenti[0].importo, Decimal("300.00"))
        self.assertIn("Gheduzzi Sofia", movimenti[0].descrizione)
        self.assertIn("BONIF. VS. FAVORE", movimenti[0].descrizione)
        self.assertEqual(movimenti[0].provider_transaction_id, "")
        self.assertEqual(movimenti[1].importo, Decimal("-24.40"))

    def test_import_estratto_conto_preview_and_confirm_with_cbi_csv(self):
        provider = ProviderBancario.objects.create(
            nome="Import file test",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto CBI",
            iban="IT00X0503437060000000003228",
            provider=provider,
            attivo=True,
        )
        uploaded = SimpleUploadedFile(
            "movimenti_cbi.csv",
            CBI_CSV_SAMPLE.encode("utf-8"),
            content_type="text/csv",
        )

        preview_response = self.client.post(
            reverse("import_estratto_conto"),
            {
                "import_action": "preview",
                "formato": "auto",
                "conto": "",
                "file": uploaded,
            },
        )

        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "Anteprima import")
        self.assertContains(preview_response, "CSV CBI")
        self.assertEqual(preview_response.context["selected_conto"], conto)
        token = preview_response.context["import_token"]
        self.assertTrue(token)

        confirm_response = self.client.post(
            reverse("import_estratto_conto"),
            {
                "import_action": "confirm",
                "import_token": token,
                "conto": str(conto.pk),
            },
        )

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(MovimentoFinanziario.objects.filter(conto=conto).count(), 2)
        movimento = MovimentoFinanziario.objects.get(importo=Decimal("300.00"))
        self.assertIn("Gheduzzi Sofia", movimento.descrizione)

    def test_import_estratto_conto_preview_stima_duplicato_storico_senza_hash(self):
        provider = ProviderBancario.objects.create(
            nome="Import preview dedup",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto preview dedup",
            provider=provider,
            attivo=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.IMPORT_FILE,
            data_contabile=date(2026, 6, 3),
            importo=Decimal("-18.90"),
            descrizione="COMM.SU BONIFICI AREA SEPA",
            controparte="Banca Test",
            incide_su_saldo_banca=True,
        )
        uploaded = SimpleUploadedFile(
            "movimenti_dedup.csv",
            (
                "Data;Importo;Descrizione;Controparte\n"
                "03/06/2026;-18,90;Comm su bonifici   area sepa;Banca Test\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            reverse("import_estratto_conto"),
            {
                "import_action": "preview",
                "formato": "auto",
                "conto": str(conto.pk),
                "file": uploaded,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["preview"]["duplicati_stimati"], 1)
        self.assertEqual(response.context["preview"]["nuovi_stimati"], 0)

    def test_excel_autodetect_parses_movements(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Data", "Data valuta", "Importo", "Descrizione", "ID transazione"])
        sheet.append([date(2026, 4, 24), date(2026, 4, 24), 315.50, "Bonifico retta Rossi", "TX-1"])
        sheet.append([date(2026, 4, 25), date(2026, 4, 25), -12.40, "Commissione banca", "TX-2"])
        buffer = BytesIO()
        workbook.save(buffer)
        raw_excel = buffer.getvalue()

        detection = detect_excel_import_config(raw_excel)

        self.assertEqual(detection.formato_rilevato, "Excel")
        self.assertEqual(detection.config.colonna_data_contabile, "data")
        self.assertEqual(detection.config.colonna_importo, "importo")

        movimenti = list(ExcelImporter(detection.config).parse(raw_excel))

        self.assertEqual(len(movimenti), 2)
        self.assertEqual(movimenti[0].data_contabile, date(2026, 4, 24))
        self.assertEqual(movimenti[0].importo, Decimal("315.50"))
        self.assertIn("Rossi", movimenti[0].descrizione)
        self.assertEqual(movimenti[0].provider_transaction_id, "TX-1")
        self.assertEqual(movimenti[1].importo, Decimal("-12.40"))

    def test_excel_html_xls_autodetect_parses_movements(self):
        raw_excel = (
            "<html><body><table>"
            "<tr><th>Data</th><th>Importo</th><th>Descrizione</th></tr>"
            "<tr><td>02/05/2026</td><td>98,40</td><td>Incasso mensa</td></tr>"
            "<tr><td>03/05/2026</td><td>-15,20</td><td>Commissione</td></tr>"
            "</table></body></html>"
        ).encode("utf-8")

        detection = detect_excel_import_config(raw_excel)
        movimenti = list(ExcelImporter(detection.config).parse(raw_excel))

        self.assertEqual(detection.formato_rilevato, "Excel")
        self.assertEqual(len(movimenti), 2)
        self.assertEqual(movimenti[0].data_contabile, date(2026, 5, 2))
        self.assertEqual(movimenti[0].importo, Decimal("98.40"))
        self.assertIn("mensa", movimenti[0].descrizione)
        self.assertEqual(movimenti[1].importo, Decimal("-15.20"))

    def test_excel_unicredit_two_line_header_autodetect_parses_movements(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Rapporto IT 40 A 02008 13030 000107342804 - SCUOLA TEST"])
        sheet.append(["Data", "", "Descrizione", "EUR", "Caus."])
        sheet.append(["Operaz.", "Valuta"])
        sheet.append(["03/12/2025", "03/12/2025", "EROGAZIONE FINANZIAMENTO", "39800", "061"])
        sheet.append(["04/12/2025", "04/12/2025", "DISPOSIZIONE DI BONIFICO", "-15005,75", "208"])
        buffer = BytesIO()
        workbook.save(buffer)
        raw_excel = buffer.getvalue()

        detection = detect_excel_import_config(raw_excel)
        movimenti = list(ExcelImporter(detection.config).parse(raw_excel))

        self.assertEqual(detection.formato_rilevato, "Excel UniCredit")
        self.assertEqual(detection.config.righe_da_saltare, 3)
        self.assertFalse(detection.config.ha_intestazione)
        self.assertEqual(detection.config.colonna_data_contabile, 0)
        self.assertEqual(detection.config.colonna_data_valuta, 1)
        self.assertEqual(detection.config.colonna_importo, 3)
        self.assertEqual(detection.abi, "02008")
        self.assertEqual(detection.cab, "13030")
        self.assertEqual(detection.numero_conto, "000107342804")
        self.assertEqual(len(movimenti), 2)
        self.assertEqual(movimenti[0].data_contabile, date(2025, 12, 3))
        self.assertEqual(movimenti[0].data_valuta, date(2025, 12, 3))
        self.assertEqual(movimenti[0].importo, Decimal("39800"))
        self.assertEqual(movimenti[1].importo, Decimal("-15005.75"))

    def test_import_estratto_conto_preview_and_confirm_with_xlsx(self):
        from openpyxl import Workbook

        provider = ProviderBancario.objects.create(
            nome="Import Excel test",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto Excel",
            iban="IT00X0503437060000000003228",
            provider=provider,
            attivo=True,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Data", "Data valuta", "Importo", "Descrizione", "ID transazione"])
        sheet.append([date(2026, 5, 1), date(2026, 5, 1), 120.75, "Bonifico laboratorio", "XLSX-1"])
        buffer = BytesIO()
        workbook.save(buffer)
        uploaded = SimpleUploadedFile(
            "movimenti_unicredit.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        preview_response = self.client.post(
            reverse("import_estratto_conto"),
            {
                "import_action": "preview",
                "formato": "auto",
                "conto": "",
                "file": uploaded,
            },
        )

        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "Anteprima import")
        self.assertContains(preview_response, "Excel")
        self.assertEqual(preview_response.context["selected_conto"], conto)
        token = preview_response.context["import_token"]
        self.assertTrue(token)

        confirm_response = self.client.post(
            reverse("import_estratto_conto"),
            {
                "import_action": "confirm",
                "import_token": token,
                "conto": str(conto.pk),
            },
        )

        self.assertEqual(confirm_response.status_code, 200)
        movimento = MovimentoFinanziario.objects.get(conto=conto)
        self.assertEqual(movimento.importo, Decimal("120.75"))
        self.assertIn("laboratorio", movimento.descrizione)

    def test_regole_categorizzazione_supportano_condizioni_testuali_avanzate(self):
        categoria = CategoriaFinanziaria.objects.create(nome="Incassi rette")
        RegolaCategorizzazione.objects.create(
            nome="Quote e commissioni",
            condizione_tipo=CondizioneRegolaCategorizzazione.DESCRIZIONE_CONTIENE,
            pattern="COMM. SU BONIFICI | quota + maggio",
            categoria_da_assegnare=categoria,
        )

        movimento_or = MovimentoFinanziario(
            data_contabile=date(2026, 4, 24),
            importo=Decimal("-2.00"),
            descrizione="COMM.SU BONIFICI AREA SEPA",
        )
        regola_or = applica_regole_a_movimento(movimento_or)

        self.assertIsNotNone(regola_or)
        self.assertEqual(movimento_or.categoria_id, categoria.pk)

        movimento_and = MovimentoFinanziario(
            data_contabile=date(2026, 5, 10),
            importo=Decimal("100.00"),
            descrizione="Versamento quota retta mese di maggio",
        )
        regola_and = applica_regole_a_movimento(movimento_and)

        self.assertIsNotNone(regola_and)
        self.assertEqual(movimento_and.categoria_id, categoria.pk)

        movimento_no_match = MovimentoFinanziario(
            data_contabile=date(2026, 5, 11),
            importo=Decimal("100.00"),
            descrizione="Versamento quota generica",
        )
        self.assertIsNone(applica_regole_a_movimento(movimento_no_match))

    @skip("Legacy test basato sulla tabella anagrafica.Famiglia rimossa.")
    def test_import_movimenti_riconcilia_automaticamente_retta_studente(self):
        provider = ProviderBancario.objects.create(
            nome="Import rette test",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto rette",
            iban="IT00X0000000000000000000000",
            provider=provider,
            attivo=True,
        )
        stato_relazione = StatoRelazioneFamiglia.objects.create(stato="Iscritta")
        famiglia = Famiglia.objects.create(
            cognome_famiglia="Bianchi",
            stato_relazione_famiglia=stato_relazione,
        )
        studente = Studente.objects.create(
            famiglia=famiglia,
            nome="Luca",
            cognome="Bianchi",
            data_nascita=date(2020, 5, 5),
        )
        anno = AnnoScolastico.objects.create(
            nome_anno_scolastico="2025/2026",
            data_inizio=date(2025, 9, 1),
            data_fine=date(2026, 6, 30),
        )
        classe = Classe.objects.create(
            nome_classe="Materna",
            ordine_classe=1,
        )
        stato_iscrizione = StatoIscrizione.objects.create(stato_iscrizione="Iscritto")
        condizione = CondizioneIscrizione.objects.create(
            anno_scolastico=anno,
            nome_condizione_iscrizione="Retta standard",
            numero_mensilita_default=10,
            mese_prima_retta=9,
            giorno_scadenza_rate=10,
        )
        TariffaCondizioneIscrizione.objects.create(
            condizione_iscrizione=condizione,
            ordine_figlio_da=1,
            retta_annuale=Decimal("1000.00"),
            preiscrizione=Decimal("0.00"),
        )
        iscrizione = Iscrizione.objects.create(
            studente=studente,
            anno_scolastico=anno,
            classe=classe,
            stato_iscrizione=stato_iscrizione,
            condizione_iscrizione=condizione,
            data_iscrizione=date(2025, 9, 1),
            data_fine_iscrizione=date(2026, 6, 30),
        )
        self.assertEqual(iscrizione.sync_rate_schedule(), "created")
        rata = RataIscrizione.objects.get(iscrizione=iscrizione, numero_rata=1)
        self.assertEqual(rata.importo_finale, Decimal("100.00"))

        raw_csv = (
            "Data;Importo;Descrizione\n"
            "10/09/2025;100,00;Bonifico retta settembre Luca Bianchi\n"
        ).encode("utf-8")
        config = CsvImporterConfig(
            delimiter=";",
            ha_intestazione=True,
            colonna_data_contabile="Data",
            colonna_importo="Importo",
            colonna_descrizione="Descrizione",
        )

        risultato = importa_movimenti_da_file(
            parser=CsvImporter(config),
            raw_bytes=raw_csv,
            conto=conto,
            provider=provider,
            nome_file="rette.csv",
        )

        self.assertEqual(risultato.inseriti, 1)
        self.assertEqual(risultato.riconciliati, 1)

        movimento = MovimentoFinanziario.objects.get(conto=conto)
        rata.refresh_from_db()
        self.assertEqual(movimento.rata_iscrizione_id, rata.pk)
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertTrue(rata.pagata)
        self.assertEqual(rata.importo_pagato, Decimal("100.00"))
        self.assertEqual(rata.data_pagamento, date(2025, 9, 10))

    @skip("Legacy test basato sulla tabella anagrafica.Famiglia rimossa.")
    def test_import_movimenti_riconcilia_pagamento_cumulativo_rette(self):
        provider = ProviderBancario.objects.create(
            nome="Import rette cumulative test",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto rette",
            iban="IT00X0000000000000000000000",
            provider=provider,
            attivo=True,
        )
        stato_relazione = StatoRelazioneFamiglia.objects.create(stato="Iscritta")
        famiglia = Famiglia.objects.create(
            cognome_famiglia="Rossi",
            stato_relazione_famiglia=stato_relazione,
        )
        anno = AnnoScolastico.objects.create(
            nome_anno_scolastico="2025/2026",
            data_inizio=date(2025, 9, 1),
            data_fine=date(2026, 6, 30),
        )
        classe = Classe.objects.create(nome_classe="Materna", ordine_classe=1)
        stato_iscrizione = StatoIscrizione.objects.create(stato_iscrizione="Iscritto")
        condizione = CondizioneIscrizione.objects.create(
            anno_scolastico=anno,
            nome_condizione_iscrizione="Retta standard",
            numero_mensilita_default=10,
            mese_prima_retta=9,
            giorno_scadenza_rate=10,
        )
        TariffaCondizioneIscrizione.objects.create(
            condizione_iscrizione=condizione,
            ordine_figlio_da=1,
            retta_annuale=Decimal("1000.00"),
            preiscrizione=Decimal("0.00"),
        )
        iscrizioni = []
        for nome in ["Luca", "Marta"]:
            studente = Studente.objects.create(
                famiglia=famiglia,
                nome=nome,
                cognome="Rossi",
                data_nascita=date(2020, 5, 5),
            )
            iscrizione = Iscrizione.objects.create(
                studente=studente,
                anno_scolastico=anno,
                classe=classe,
                stato_iscrizione=stato_iscrizione,
                condizione_iscrizione=condizione,
                data_iscrizione=date(2025, 9, 1),
                data_fine_iscrizione=date(2026, 6, 30),
            )
            self.assertEqual(iscrizione.sync_rate_schedule(), "created")
            iscrizioni.append(iscrizione)

        rate = [
            iscrizione.rate.get(tipo_rata=RataIscrizione.TIPO_MENSILE, numero_rata=1)
            for iscrizione in iscrizioni
        ]
        raw_csv = (
            "Data;Importo;Descrizione\n"
            "10/09/2025;200,00;Bonifico rette settembre Luca e Marta Rossi\n"
        ).encode("utf-8")
        config = CsvImporterConfig(
            delimiter=";",
            ha_intestazione=True,
            colonna_data_contabile="Data",
            colonna_importo="Importo",
            colonna_descrizione="Descrizione",
        )

        risultato = importa_movimenti_da_file(
            parser=CsvImporter(config),
            raw_bytes=raw_csv,
            conto=conto,
            provider=provider,
            nome_file="rette-cumulative.csv",
        )

        self.assertEqual(risultato.inseriti, 1)
        self.assertEqual(risultato.riconciliati, 1)
        movimento = MovimentoFinanziario.objects.get(conto=conto)
        movimento.refresh_from_db()
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertIsNone(movimento.rata_iscrizione_id)
        self.assertEqual(movimento.riconciliazioni_rate.count(), 2)
        for rata in rate:
            rata.refresh_from_db()
            self.assertTrue(rata.pagata)
            self.assertEqual(rata.importo_pagato, Decimal("100.00"))
            self.assertEqual(rata.data_pagamento, date(2025, 9, 10))

    @skip("Legacy test basato sulla tabella anagrafica.Famiglia rimossa.")
    def test_riconcilia_movimento_con_rate_supporta_pagamento_cumulativo(self):
        stato_relazione = StatoRelazioneFamiglia.objects.create(stato="Iscritta")
        famiglia = Famiglia.objects.create(
            cognome_famiglia="Rossi",
            stato_relazione_famiglia=stato_relazione,
        )
        anno = AnnoScolastico.objects.create(
            nome_anno_scolastico="2025/2026",
            data_inizio=date(2025, 9, 1),
            data_fine=date(2026, 6, 30),
        )
        classe = Classe.objects.create(nome_classe="Materna", ordine_classe=1)
        stato_iscrizione = StatoIscrizione.objects.create(stato_iscrizione="Iscritto")
        condizione = CondizioneIscrizione.objects.create(
            anno_scolastico=anno,
            nome_condizione_iscrizione="Retta standard",
            numero_mensilita_default=10,
        )
        TariffaCondizioneIscrizione.objects.create(
            condizione_iscrizione=condizione,
            ordine_figlio_da=1,
            retta_annuale=Decimal("1000.00"),
        )
        rate = []
        for nome in ["Luca", "Marta"]:
            studente = Studente.objects.create(
                famiglia=famiglia,
                nome=nome,
                cognome="Rossi",
                data_nascita=date(2020, 5, 5),
            )
            iscrizione = Iscrizione.objects.create(
                studente=studente,
                anno_scolastico=anno,
                classe=classe,
                stato_iscrizione=stato_iscrizione,
                condizione_iscrizione=condizione,
                data_iscrizione=date(2025, 9, 1),
                data_fine_iscrizione=date(2026, 6, 30),
            )
            rate.append(
                RataIscrizione.objects.create(
                    iscrizione=iscrizione,
                    famiglia=famiglia,
                    numero_rata=1,
                    mese_riferimento=9,
                    anno_riferimento=2025,
                    importo_dovuto=Decimal("100.00"),
                    importo_finale=Decimal("100.00"),
                    data_scadenza=date(2025, 9, 10),
                )
            )

        movimento = MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 11),
            importo=Decimal("200.00"),
            descrizione="Bonifico rette Luca e Marta Rossi",
        )

        riconcilia_movimento_con_rate(
            movimento,
            [(rate[0], Decimal("100.00")), (rate[1], Decimal("100.00"))],
            utente=self.user,
        )

        movimento.refresh_from_db()
        for rata in rate:
            rata.refresh_from_db()
            self.assertTrue(rata.pagata)
            self.assertEqual(rata.importo_pagato, Decimal("100.00"))
            self.assertEqual(rata.data_pagamento, date(2025, 9, 11))
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertIsNone(movimento.rata_iscrizione_id)
        self.assertEqual(RiconciliazioneRataMovimento.objects.filter(movimento=movimento).count(), 2)

    def test_lista_movimenti_colora_entrate_e_uscite(self):
        categoria = CategoriaFinanziaria.objects.create(
            nome="Rette",
            tipo=TipoCategoriaFinanziaria.ENTRATA,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 1),
            importo=Decimal("100.00"),
            descrizione="Incasso retta",
            categoria=categoria,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 2),
            importo=Decimal("-25.00"),
            descrizione="Spesa bancaria",
            categoria=categoria,
        )

        response = self.client.get(reverse("lista_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "finance-movement-filters")
        self.assertContains(response, "finance-movement-row-incoming")
        self.assertContains(response, "finance-movement-row-outgoing")

        response = self.client.get(reverse("dashboard_gestione_finanziaria"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "finance-movement-row-incoming")
        self.assertContains(response, "finance-movement-row-outgoing")

    def test_lista_movimenti_filtra_per_tipo_e_intervallo_date(self):
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 10),
            importo=Decimal("120.00"),
            descrizione="Incasso filtro aprile",
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 4, 12),
            importo=Decimal("-35.00"),
            descrizione="Pagamento filtro aprile",
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 5, 5),
            importo=Decimal("-20.00"),
            descrizione="Pagamento filtro maggio",
        )

        response = self.client.get(reverse("lista_movimenti_finanziari"), {"tipo": "entrate"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Solo entrate")
        self.assertContains(response, "Incasso filtro aprile")
        self.assertNotContains(response, "Pagamento filtro aprile")

        response = self.client.get(
            reverse("lista_movimenti_finanziari"),
            {
                "tipo": "uscite",
                "data_da": "2026-04-01",
                "data_a": "2026-04-30",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="data_da" value="2026-04-01"', html=False)
        self.assertContains(response, 'name="data_a" value="2026-04-30"', html=False)
        self.assertContains(response, "Pagamento filtro aprile")
        self.assertNotContains(response, "Incasso filtro aprile")
        self.assertNotContains(response, "Pagamento filtro maggio")

    def test_lista_movimenti_mostra_stato_riconciliazione_effettivo(self):
        scadenza_completa, movimento_completo = self._crea_scadenza_pagamento_test(importo=Decimal("100.00"))
        registra_pagamento_fornitore(
            scadenza_completa,
            importo=Decimal("100.00"),
            data_pagamento=movimento_completo.data_contabile,
            movimento=movimento_completo,
            metodo=MetodoPagamentoFornitore.BANCA,
            conto=movimento_completo.conto,
            utente=self.user,
        )
        MovimentoFinanziario.objects.filter(pk=movimento_completo.pk).update(
            stato_riconciliazione=StatoRiconciliazione.NON_RICONCILIATO
        )

        scadenza_parziale, movimento_parziale = self._crea_scadenza_pagamento_test(importo=Decimal("100.00"))
        movimento_parziale.importo = Decimal("-150.00")
        movimento_parziale.descrizione = "Pagamento parziale Beta Servizi"
        movimento_parziale.save(update_fields=["importo", "descrizione"])
        registra_pagamento_fornitore(
            scadenza_parziale,
            importo=Decimal("100.00"),
            data_pagamento=movimento_parziale.data_contabile,
            movimento=movimento_parziale,
            metodo=MetodoPagamentoFornitore.BANCA,
            conto=movimento_parziale.conto,
            utente=self.user,
        )

        response = self.client.get(reverse("lista_movimenti_finanziari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Riconciliato")
        self.assertContains(response, "Parzialmente riconciliato")
        movimento_completo.refresh_from_db()
        movimento_parziale.refresh_from_db()
        self.assertEqual(movimento_completo.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)
        self.assertEqual(movimento_parziale.stato_riconciliazione, StatoRiconciliazione.NON_RICONCILIATO)

    def test_eliminazione_multipla_movimenti_conferma_e_ricalcola_saldo(self):
        conto = ContoBancario.objects.create(
            nome_conto="Conto bulk",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
            attivo=True,
        )
        SaldoConto.objects.create(
            conto=conto,
            data_riferimento=timezone.make_aware(datetime(2026, 4, 1, 23, 59)),
            saldo_contabile=Decimal("1000.00"),
            fonte=FonteSaldo.MANUALE,
        )
        movimento_da_eliminare = MovimentoFinanziario.objects.create(
            conto=conto,
            canale=CanaleMovimento.BANCA,
            data_contabile=date(2026, 4, 2),
            importo=Decimal("-100.00"),
            descrizione="Movimento da eliminare",
            incide_su_saldo_banca=True,
        )
        movimento_da_mantenere = MovimentoFinanziario.objects.create(
            conto=conto,
            canale=CanaleMovimento.BANCA,
            data_contabile=date(2026, 4, 3),
            importo=Decimal("-25.00"),
            descrizione="Movimento da mantenere",
            incide_su_saldo_banca=True,
        )
        next_url = reverse("lista_movimenti_finanziari") + f"?conto={conto.pk}"

        response = self.client.get(reverse("lista_movimenti_finanziari"))
        self.assertContains(response, reverse("elimina_movimenti_finanziari_multipla"))
        self.assertContains(response, "data-bulk-form")

        response = self.client.post(
            reverse("elimina_movimenti_finanziari_multipla"),
            {
                "selected_ids": [str(movimento_da_eliminare.pk)],
                "next": next_url,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Elimina movimenti")
        self.assertContains(response, "Movimento da eliminare")

        response = self.client.post(
            reverse("elimina_movimenti_finanziari_multipla"),
            {
                "selected_ids": [str(movimento_da_eliminare.pk)],
                "next": next_url,
                "conferma": "1",
            },
        )

        self.assertRedirects(response, next_url)
        self.assertFalse(MovimentoFinanziario.objects.filter(pk=movimento_da_eliminare.pk).exists())
        self.assertTrue(MovimentoFinanziario.objects.filter(pk=movimento_da_mantenere.pk).exists())
        conto.refresh_from_db()
        self.assertEqual(conto.saldo_corrente, Decimal("975.00"))

    def test_saldo_conto_manuale_alimenta_saldo_corrente_con_movimenti_successivi(self):
        conto = ContoBancario.objects.create(
            nome_conto="Cassa contanti",
            tipo_conto=TipoContoFinanziario.CASSA_CONTANTI,
            attivo=True,
        )

        response = self.client.post(
            reverse("crea_saldo_conto"),
            {
                "conto": str(conto.pk),
                "data_riferimento": "2026-04-01T23:59",
                "saldo_contabile": "1000.00",
                "saldo_disponibile": "",
                "valuta": "EUR",
                "fonte": FonteSaldo.MANUALE,
                "note": "Saldo iniziale cassa",
            },
        )

        self.assertRedirects(response, reverse("lista_saldi_conti"))
        self.assertEqual(SaldoConto.objects.filter(conto=conto).count(), 1)
        conto.refresh_from_db()
        self.assertEqual(conto.saldo_corrente, Decimal("1000.00"))

        response = self.client.post(
            reverse("crea_movimento_manuale"),
            {
                "conto": str(conto.pk),
                "canale": CanaleMovimento.CONTANTI,
                "data_contabile": "2026-04-02",
                "data_valuta": "",
                "importo": "-100.00",
                "valuta": "EUR",
                "descrizione": "Acquisto contanti",
                "controparte": "",
                "iban_controparte": "",
                "categoria": "",
                "incide_su_saldo_banca": "on",
                "sostenuta_da_terzi": "",
                "rimborsabile": "",
                "sostenitore": "",
                "note": "",
            },
        )

        self.assertRedirects(response, reverse("lista_movimenti_finanziari"))
        conto.refresh_from_db()
        self.assertEqual(conto.saldo_corrente, Decimal("900.00"))

    def test_saldo_manuale_e_accessibile_dalle_impostazioni_con_conto_preselezionato(self):
        conto = ContoBancario.objects.create(
            nome_conto="Conto operativo",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
            attivo=True,
        )

        response = self.client.get(reverse("lista_conti_bancari"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Inserisci saldo manuale")
        self.assertContains(response, f"{reverse('crea_saldo_conto')}?conto={conto.pk}")

        response = self.client.get(f"{reverse('crea_saldo_conto')}?conto={conto.pk}")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Inserisci qui il saldo rilevato")
        self.assertContains(response, f'<option value="{conto.pk}" selected>{conto.nome_conto}</option>', html=True)

    def test_movimento_personale_usa_badge_e_non_incide_sul_saldo(self):
        conto = ContoBancario.objects.create(
            nome_conto="Conto operativo",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
            saldo_corrente=Decimal("500.00"),
            saldo_corrente_aggiornato_al=timezone.now(),
            attivo=True,
        )

        response = self.client.post(
            reverse("crea_movimento_manuale"),
            {
                "conto": str(conto.pk),
                "canale": CanaleMovimento.PERSONALE,
                "data_contabile": "2026-04-03",
                "data_valuta": "",
                "importo": "-35.00",
                "valuta": "EUR",
                "descrizione": "Materiale pagato da genitore",
                "controparte": "Genitore",
                "iban_controparte": "",
                "categoria": "",
                "incide_su_saldo_banca": "",
                "sostenuta_da_terzi": "",
                "rimborsabile": "",
                "sostenitore": "Genitore",
                "note": "",
            },
        )

        self.assertRedirects(response, reverse("lista_movimenti_finanziari"))
        movimento = MovimentoFinanziario.objects.get(descrizione="Materiale pagato da genitore")
        self.assertTrue(movimento.sostenuta_da_terzi)
        self.assertFalse(movimento.incide_su_saldo_banca)

        response = self.client.get(reverse("lista_movimenti_finanziari"))
        self.assertContains(response, "finance-channel-badge-personale")
        self.assertContains(response, "senza rimborso")

        response = self.client.get(reverse("crea_movimento_manuale"))
        self.assertContains(response, "movimento-finanziario-form.js")

    def test_movimento_popup_accetta_prefill_da_busta_paga(self):
        response = self.client.get(
            reverse("crea_movimento_manuale"),
            {
                "popup": "1",
                "target_input_name": "movimento_pagamento",
                "data_contabile": "2025-10-31",
                "importo": "-1302.00",
                "descrizione": "Pagamento busta paga Mario Rossi",
                "valuta": "EUR",
                "canale": CanaleMovimento.BANCA,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="target_input_name" value="movimento_pagamento"', html=False)
        self.assertContains(response, "Pagamento busta paga Mario Rossi")
        self.assertContains(response, 'value="-1302.00"', html=False)

    def test_movimento_popup_collega_pagamento_a_busta_paga(self):
        dipendente = Dipendente.objects.create(
            nome="Mario",
            cognome="Rossi",
            codice_fiscale="RSSMRA80A01H501U",
        )
        busta = BustaPagaDipendente.objects.create(
            dipendente=dipendente,
            anno=2025,
            mese=10,
            netto_effettivo=Decimal("1302.00"),
        )

        response = self.client.post(
            f"{reverse('crea_movimento_manuale')}?popup=1",
            {
                "popup": "1",
                "target_input_name": "movimento_pagamento",
                "busta_paga_pagamento": str(busta.pk),
                "conto": "",
                "canale": CanaleMovimento.BANCA,
                "data_contabile": "2025-10-31",
                "data_valuta": "",
                "importo": "-1302.00",
                "valuta": "EUR",
                "descrizione": "Pagamento busta paga Mario Rossi",
                "controparte": "Mario Rossi",
                "iban_controparte": "",
                "categoria": "",
                "incide_su_saldo_banca": "",
                "sostenuta_da_terzi": "",
                "rimborsabile": "",
                "sostenitore": "",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        movimento = MovimentoFinanziario.objects.get(descrizione="Pagamento busta paga Mario Rossi")
        busta.refresh_from_db()
        movimento.refresh_from_db()
        self.assertEqual(busta.movimento_pagamento, movimento)
        self.assertEqual(busta.data_pagamento_effettiva, date(2025, 10, 31))
        self.assertEqual(movimento.stato_riconciliazione, StatoRiconciliazione.RICONCILIATO)

    def test_dashboard_mostra_saldi_per_tipo_conto(self):
        conto = ContoBancario.objects.create(
            nome_conto="Cassa",
            tipo_conto=TipoContoFinanziario.CASSA_CONTANTI,
            attivo=True,
        )
        SaldoConto.objects.create(
            conto=conto,
            data_riferimento=timezone.now(),
            saldo_contabile=Decimal("250.00"),
            fonte=FonteSaldo.MANUALE,
        )

        response = self.client.get(reverse("dashboard_gestione_finanziaria"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Saldi per tipo")
        self.assertContains(response, "Cassa contanti")
        self.assertContains(response, "250,00")

    def test_template_import_saldi_conti_csv(self):
        response = self.client.get(reverse("scarica_template_saldi_conti_csv"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertContains(response, "nome_conto;data_riferimento;saldo_contabile")

    def test_import_saldi_banco_bpm_cbi_usa_iban_e_colonne_banca(self):
        conto = ContoBancario.objects.create(
            nome_conto="Conto Banco BPM",
            iban="IT67C0503437060000000003228",
            attivo=True,
        )
        raw_csv = (
            '"Ragione sociale";"Banca";"Rapporto";"IBAN";"Data";"Saldo divisa";"Saldo liquido";"Div."\n'
            '"IL SOLE E L\'ALTRE STELLE SRL IMPRESA SOCIALE";"05034 - BANCO BPM S.P.A.";'
            '"37060 - 000000003228";"IT67C0503437060000000003228";"28/04/2026";'
            '"980,89";"980,89";"EUR"\n'
        )
        uploaded = SimpleUploadedFile(
            "RiepilogoSaldiCBI_30_04_2026_01.24.53.csv",
            raw_csv.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(reverse("import_saldi_conti"), {"file": uploaded})

        self.assertRedirects(response, reverse("lista_saldi_conti"))
        saldo = SaldoConto.objects.get(conto=conto)
        self.assertEqual(saldo.data_riferimento.date(), date(2026, 4, 28))
        self.assertEqual(saldo.saldo_contabile, Decimal("980.89"))
        self.assertEqual(saldo.saldo_disponibile, Decimal("980.89"))
        conto.refresh_from_db()
        self.assertEqual(conto.saldo_corrente, Decimal("980.89"))

    def test_import_saldi_banco_bpm_online_deduce_data_da_nome_file_e_crea_conto(self):
        raw_csv = (
            '"Ragione sociale";"Banca";"Rapporto";"IBAN";"Saldo finale";"Saldo disponibile";"Div."\n'
            '"IL SOLE E L\'ALTRE STELLE SRL IMPRESA SOCIALE";"05034 - BANCO BPM S.P.A.";'
            '"37060 - 056300003228";"IT67C0503437060000000003228";"980,89";"980,89";"EUR"\n'
        )
        uploaded = SimpleUploadedFile(
            "SaldiCC_OnLine_30_04_2026_01.24.38.csv",
            raw_csv.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(reverse("import_saldi_conti"), {"file": uploaded})

        self.assertRedirects(response, reverse("lista_saldi_conti"))
        conto = ContoBancario.objects.get(iban="IT67C0503437060000000003228")
        saldo = SaldoConto.objects.get(conto=conto)
        self.assertEqual(timezone.localtime(saldo.data_riferimento).date(), date(2026, 4, 30))
        self.assertEqual(saldo.saldo_contabile, Decimal("980.89"))
        self.assertEqual(conto.banca, "05034 - BANCO BPM S.P.A.")

    def test_pulizia_movimenti_automatici_elimina_import_non_manuali(self):
        provider = ProviderBancario.objects.create(
            nome="Import test",
            tipo=TipoProviderBancario.IMPORT_FILE,
        )
        conto = ContoBancario.objects.create(
            nome_conto="Conto operativo",
            iban="IT00X0000000000000000000000",
            provider=provider,
            attivo=True,
            saldo_corrente=Decimal("75.00"),
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.IMPORT_FILE,
            data_contabile=date(2026, 4, 1),
            importo=Decimal("100.00"),
            descrizione="Import file",
            incide_su_saldo_banca=True,
        )
        MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.BANCA,
            data_contabile=date(2026, 4, 2),
            importo=Decimal("-25.00"),
            descrizione="Sync banca",
            incide_su_saldo_banca=True,
        )
        manuale = MovimentoFinanziario.objects.create(
            conto=conto,
            origine=OrigineMovimento.MANUALE,
            data_contabile=date(2026, 4, 3),
            importo=Decimal("50.00"),
            descrizione="Manuale",
            incide_su_saldo_banca=False,
        )

        response = self.client.get(reverse("lista_movimenti_finanziari"))
        self.assertContains(response, reverse("pulizia_movimenti_finanziari"))

        response = self.client.get(reverse("pulizia_movimenti_finanziari"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ripulisci movimenti")
        self.assertEqual(response.context["statistiche"]["totale"], 3)
        self.assertEqual(response.context["statistiche"]["automatici"], 2)
        self.assertEqual(response.context["statistiche"]["manuali"], 1)

        response = self.client.post(
            reverse("pulizia_movimenti_finanziari"),
            {
                "ambito": "automatici",
                "conferma": "ELIMINA",
            },
        )

        self.assertRedirects(response, reverse("lista_movimenti_finanziari"))
        self.assertEqual(MovimentoFinanziario.objects.count(), 1)
        self.assertTrue(MovimentoFinanziario.objects.filter(pk=manuale.pk).exists())
        conto.refresh_from_db()
        self.assertEqual(conto.saldo_corrente, Decimal("0"))

    def test_pulizia_movimenti_richiede_conferma_testuale(self):
        MovimentoFinanziario.objects.create(
            origine=OrigineMovimento.MANUALE,
            data_contabile=date(2026, 4, 1),
            importo=Decimal("10.00"),
            descrizione="Manuale",
        )

        response = self.client.post(
            reverse("pulizia_movimenti_finanziari"),
            {
                "ambito": "manuali",
                "conferma": "elimina tutto",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(MovimentoFinanziario.objects.count(), 1)
        self.assertContains(response, "Per confermare devi digitare")

    def test_report_categorie_filtra_per_anno_scolastico(self):
        anno = AnnoScolastico.objects.create(
            nome_anno_scolastico="2025/2026",
            data_inizio=date(2025, 9, 1),
            data_fine=date(2026, 8, 31),
        )
        CondizioneIscrizione.objects.create(
            anno_scolastico=anno,
            nome_condizione_iscrizione="Retta standard",
            numero_mensilita_default=10,
            mese_prima_retta=9,
            giorno_scadenza_rate=10,
        )
        categoria = CategoriaFinanziaria.objects.create(
            nome="Rette",
            tipo=TipoCategoriaFinanziaria.ENTRATA,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 8, 31),
            importo=Decimal("999.00"),
            descrizione="Fuori anno scolastico",
            categoria=categoria,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2025, 9, 1),
            importo=Decimal("100.00"),
            descrizione="Inizio anno scolastico",
            categoria=categoria,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 6, 30),
            importo=Decimal("50.00"),
            descrizione="Fine anno scolastico",
            categoria=categoria,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 7, 1),
            importo=Decimal("999.00"),
            descrizione="Dopo anno scolastico",
            categoria=categoria,
        )

        response = self.client.get(
            reverse("report_categorie_annuale"),
            {"periodo": "scolastico", "anno_scolastico": str(anno.pk)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["periodo_tipo"], "scolastico")
        self.assertEqual(response.context["periodo_label"], "anno scolastico 2025/2026")
        self.assertEqual(response.context["totale_entrate"], Decimal("150.00"))

        response = self.client.get(
            reverse("report_categorie_mensile"),
            {"periodo": "scolastico", "anno_scolastico": str(anno.pk)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["periodo_tipo"], "scolastico")
        self.assertEqual(response.context["mesi"][0], "Set 2025")
        self.assertEqual(response.context["mesi"][-1], "Giu 2026")
        self.assertEqual(response.context["totale_generale"], Decimal("150.00"))
        self.assertContains(response, "Sintesi")
        self.assertContains(response, "EUR 150,00")
        self.assertContains(response, "report-category-entrata")

    def test_report_categorie_esclude_ricariche_prepagate_e_trasferimenti(self):
        conto_corrente = ContoBancario.objects.create(
            nome_conto="Conto operativo",
            tipo_conto=TipoContoFinanziario.CONTO_CORRENTE,
        )
        prepagata = ContoBancario.objects.create(
            nome_conto="Carta web",
            tipo_conto=TipoContoFinanziario.CARTA_PREPAGATA,
        )
        entrate = CategoriaFinanziaria.objects.create(
            nome="Rette",
            tipo=TipoCategoriaFinanziaria.ENTRATA,
        )
        trasferimenti = CategoriaFinanziaria.objects.create(
            nome="Giroconti",
            tipo=TipoCategoriaFinanziaria.TRASFERIMENTO,
        )
        MovimentoFinanziario.objects.create(
            conto=conto_corrente,
            data_contabile=date(2026, 1, 10),
            importo=Decimal("150.00"),
            descrizione="Incasso retta",
            categoria=entrate,
        )
        MovimentoFinanziario.objects.create(
            conto=prepagata,
            data_contabile=date(2026, 1, 11),
            importo=Decimal("300.00"),
            descrizione="Ricarica carta",
            categoria=entrate,
            canale=CanaleMovimento.PREPAGATA,
        )
        MovimentoFinanziario.objects.create(
            conto=conto_corrente,
            data_contabile=date(2026, 1, 11),
            importo=Decimal("-300.00"),
            descrizione="Giroconto verso carta",
            categoria=trasferimenti,
        )

        response = self.client.get(
            reverse("report_categorie_annuale"),
            {"periodo": "solare", "anno": "2026"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["totale_entrate"], Decimal("150.00"))
        self.assertEqual(response.context["totale_uscite"], Decimal("0"))
        self.assertEqual(response.context["saldo_netto"], Decimal("150.00"))

        response = self.client.get(
            reverse("report_categorie_mensile"),
            {"periodo": "solare", "anno": "2026"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["totale_entrate"], Decimal("150.00"))
        self.assertEqual(response.context["totale_generale"], Decimal("150.00"))

    def test_report_categorie_annuale_mostra_categorie_figlie(self):
        categoria_padre = CategoriaFinanziaria.objects.create(
            nome="Utenze",
            tipo=TipoCategoriaFinanziaria.SPESA,
        )
        categoria_figlia = CategoriaFinanziaria.objects.create(
            nome="Energia elettrica",
            tipo=TipoCategoriaFinanziaria.SPESA,
            parent=categoria_padre,
        )
        MovimentoFinanziario.objects.create(
            data_contabile=date(2026, 1, 10),
            importo=Decimal("-1000.00"),
            descrizione="Bolletta luce",
            categoria=categoria_figlia,
        )

        response = self.client.get(
            reverse("report_categorie_annuale"),
            {"periodo": "solare", "anno": "2026"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Utenze")
        self.assertContains(response, "Energia elettrica")
        self.assertContains(response, "2026")
        self.assertNotContains(response, "2.026")
        self.assertContains(response, 'data-report-category-toggle="categoria-')
        self.assertContains(response, "report-category-spesa")
        self.assertContains(response, "-1.000,00")
        self.assertEqual(response.context["totale_uscite"], Decimal("-1000.00"))
